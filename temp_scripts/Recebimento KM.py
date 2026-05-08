import re
from pathlib import Path
from typing import List, Dict, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


PASTA_PDFS = Path(
    r"\\virm-rgr022\FILESERVER\Projetos\05_HANDYMAX\09. Doc Control\15 - Documentos KM\0 Transmittal Letters\Transmittal Letters"
)

ARQUIVO_EXCEL_NOVO = Path(
    r"\\virm-rgr022\FILESERVER\Projetos\05_HANDYMAX\09. Doc Control\15 - Documentos KM\0 Transmittal Letters\Transmittal Letters\Lista de Docs recebidos KM - NOVA.xlsx"
)

ABA_PLANILHA = "Planilha1"
ABA_LOG = "LOG"

CABECALHOS = [
    "Documento",
    "Titulo",
    "Pasta",
    "Emissão",
    "Proposito de Emissão",
    "Data Envio",
    "Transmittal N°",
]

PROPOSITOS_CONHECIDOS = [
    "Send Update without Approval",
    "Send for Re-Approval",
    "Send for Approval",
    "Send for Information",
    "Send Preliminary",
    "For Approval",
    "For Information",
]

PREENCHIMENTO_AMARELO = PatternFill(fill_type="solid", fgColor="FFF2CC")
PREENCHIMENTO_VERMELHO = PatternFill(fill_type="solid", fgColor="F4CCCC")
FONTE_LINK = Font(color="0563C1", underline="single")


def normalizar_data(texto: str) -> str:
    if not texto:
        return ""
    return re.sub(r"\b(\d{2})\.(\d{2})\.(\d{4})\b", r"\1-\2-\3", texto.strip())


def limpar_valor(valor: str) -> str:
    if not valor:
        return ""
    valor = valor.strip(" -;\n\t")
    valor = re.sub(r"\s+", " ", valor)
    return valor.strip()


def extrair_texto_pdf(caminho_pdf: Path) -> str:
    textos = []
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if texto:
                    textos.append(texto)
    except Exception as e:
        print(f"[ERRO] Falha ao ler PDF {caminho_pdf.name}: {e}")
        return ""
    return "\n".join(textos)


def normalizar_texto(texto: str) -> str:
    texto = texto.replace("\r", "\n").replace("\xa0", " ")
    texto = re.sub(r"[ \t]+", " ", texto)

    correcoes = [
        (r"Send Update without\s*\n\s*Approval", "Send Update without Approval"),
        (r"Send for Re-\s*\n\s*Approval", "Send for Re-Approval"),
        (r"Send for\s*\n\s*Approval", "Send for Approval"),
        (r"Send for\s*\n\s*Information", "Send for Information"),
        (r"Send\s*\n\s*Preliminary", "Send Preliminary"),
        (r"For\s*\n\s*Approval", "For Approval"),
        (r"For\s*\n\s*Information", "For Information"),
    ]

    for padrao, substituicao in correcoes:
        texto = re.sub(padrao, substituicao, texto, flags=re.IGNORECASE)

    texto = re.sub(r"\n+", "\n", texto)
    return texto.strip()


def extrair_transmittal(texto: str, nome_arquivo: str) -> str:
    m = re.search(r"Transmittal number:\s*([0-9]+)", texto, re.IGNORECASE)
    if m:
        return f"T-{m.group(1).strip()}"

    m = re.search(r"^T-([0-9]+)", nome_arquivo, re.IGNORECASE)
    if m:
        return f"T-{m.group(1).strip()}"

    return ""


def extrair_data_envio(texto: str) -> str:
    m = re.search(r"Sent By:\s*.*?,\s*(\d{2}\.\d{2}\.\d{4})", texto, re.IGNORECASE)
    if m:
        return normalizar_data(m.group(1))

    datas = re.findall(r"\b\d{2}\.\d{2}\.\d{4}\b", texto)
    return normalizar_data(datas[0]) if datas else ""


def extrair_bloco_document_information(texto: str) -> str:
    padroes_fim = [
        r"\nSent By:",
        r"\nTotal attachments:",
        r"\nPage \d+ of \d+",
        r"\nPlease find attached",
        r"\nThis transmittal",
    ]

    inicio = re.search(r"Document Information", texto, re.IGNORECASE)
    if not inicio:
        return texto

    trecho = texto[inicio.end():]
    fim_pos = None

    for padrao in padroes_fim:
        m = re.search(padrao, trecho, re.IGNORECASE)
        if m:
            pos = m.start()
            if fim_pos is None or pos < fim_pos:
                fim_pos = pos

    if fim_pos is not None:
        trecho = trecho[:fim_pos]

    return trecho.strip()


def identificar_proposito_em_texto(texto: str) -> str:
    for prop in sorted(PROPOSITOS_CONHECIDOS, key=len, reverse=True):
        if re.search(re.escape(prop), texto, re.IGNORECASE):
            return prop
    return ""


def separar_emissao_e_proposito(comment_texto: str) -> Tuple[str, str]:
    comment_texto = limpar_valor(comment_texto)

    proposito = identificar_proposito_em_texto(comment_texto)
    if proposito:
        idx = re.search(re.escape(proposito), comment_texto, re.IGNORECASE)
        if idx:
            emissao = limpar_valor(comment_texto[:idx.start()])
            return emissao, proposito

    for prop in PROPOSITOS_CONHECIDOS:
        if comment_texto.lower() == prop.lower():
            return "", prop

    if comment_texto.lower() == "not applicable":
        return "Not applicable", ""

    return comment_texto, ""


def encontrar_documentos_no_bloco(bloco: str) -> List[Dict[str, str]]:
    resultados = []
    linhas = [limpar_valor(l) for l in bloco.splitlines() if limpar_valor(l)]

    for linha in linhas:
        if re.search(
            r"^(Comment:|Rev\.?|Revision|Purpose|Scale|Format|Total|Sent By:)",
            linha,
            re.IGNORECASE,
        ):
            continue

        m = re.match(
            r"^([0-9]{3,4}(?:-[0-9]{2,4}){1,4})(?:-ETS)?[- ]+(.+?)\s+\(([^)]+)\)$",
            linha,
            re.IGNORECASE,
        )
        if m:
            resultados.append({
                "Documento": limpar_valor(m.group(1)),
                "Titulo": limpar_valor(m.group(2)),
                "Pasta": limpar_valor(m.group(3)),
            })

    vistos = set()
    unicos = []
    for item in resultados:
        chave = (item["Documento"], item["Titulo"], item["Pasta"])
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(item)

    return unicos


def extrair_comment_global(bloco: str) -> str:
    m = re.search(r"Comment:\s*(.+)", bloco, re.IGNORECASE | re.DOTALL)
    if not m:
        return ""

    comment = m.group(1)

    for marcador in ["Rev.", "Revision", "Scale", "Format", "Size", "A3", "A4", "A1"]:
        pos = re.search(rf"\b{re.escape(marcador)}\b", comment, re.IGNORECASE)
        if pos:
            comment = comment[:pos.start()]
            break

    return limpar_valor(comment)


def extrair_registros_pdf(texto: str, nome_arquivo: str, caminho_pdf: Path) -> List[Dict[str, str]]:
    texto = normalizar_texto(texto)
    bloco = extrair_bloco_document_information(texto)

    transmittal = extrair_transmittal(texto, nome_arquivo)
    data_envio = extrair_data_envio(texto)

    documentos = encontrar_documentos_no_bloco(bloco)
    comment_global = extrair_comment_global(bloco)
    emissao_global, proposito_global = separar_emissao_e_proposito(comment_global)

    registros = []

    if documentos:
        for doc in documentos:
            registros.append({
                "Documento": doc["Documento"],
                "Titulo": doc["Titulo"],
                "Pasta": doc["Pasta"],
                "Emissão": emissao_global,
                "Proposito de Emissão": proposito_global,
                "Data Envio": data_envio,
                "Transmittal N°": transmittal,
                "Arquivo PDF": str(caminho_pdf),
                "Status Parse": "OK",
                "Observação Parse": "",
            })
        return registros

    m_nome = re.search(
        r"\d{2}-\d{4}-\d{2}-([0-9]{3,4}(?:-[0-9]{2,4}){1,4})(?:-ETS)?[- ]+(.+?);\s*(.+?)\.pdf$",
        nome_arquivo,
        re.IGNORECASE,
    )
    if m_nome:
        emissao_nome, proposito_nome = separar_emissao_e_proposito(limpar_valor(m_nome.group(3)))
        registros.append({
            "Documento": limpar_valor(m_nome.group(1)),
            "Titulo": limpar_valor(m_nome.group(2)),
            "Pasta": "",
            "Emissão": emissao_nome or emissao_global,
            "Proposito de Emissão": proposito_nome or proposito_global,
            "Data Envio": data_envio,
            "Transmittal N°": transmittal,
            "Arquivo PDF": str(caminho_pdf),
            "Status Parse": "PARCIAL",
            "Observação Parse": "Registro montado com fallback pelo nome do arquivo.",
        })
        return registros

    registros.append({
        "Documento": "",
        "Titulo": Path(nome_arquivo).stem,
        "Pasta": "",
        "Emissão": emissao_global,
        "Proposito de Emissão": proposito_global,
        "Data Envio": data_envio,
        "Transmittal N°": transmittal,
        "Arquivo PDF": str(caminho_pdf),
        "Status Parse": "FALHA",
        "Observação Parse": "Não foi possível identificar o bloco de documentos.",
    })
    return registros


def criar_planilha_nova():
    wb = Workbook()
    ws = wb.active
    ws.title = ABA_PLANILHA
    ws_log = wb.create_sheet(ABA_LOG)

    for i, nome in enumerate(CABECALHOS, start=1):
        ws.cell(row=1, column=i).value = nome

    cab_log = ["Arquivo PDF", "Transmittal N°", "Status", "Mensagem"]
    for i, nome in enumerate(cab_log, start=1):
        ws_log.cell(row=1, column=i).value = nome

    return wb, ws, ws_log


def aplicar_link_nativo(celula, texto_exibido: str, arquivo_pdf: str):
    celula.value = texto_exibido if texto_exibido else ""
    if arquivo_pdf and texto_exibido:
        celula.hyperlink = arquivo_pdf
        celula.font = FONTE_LINK


def destacar_linha(ws, linha: int, dados: dict):
    campos_criticos = {
        1: dados.get("Documento", ""),
        2: dados.get("Titulo", ""),
        3: dados.get("Pasta", ""),
        5: dados.get("Proposito de Emissão", ""),
        6: dados.get("Data Envio", ""),
        7: dados.get("Transmittal N°", ""),
    }

    for coluna, valor in campos_criticos.items():
        if not str(valor).strip():
            ws.cell(linha, coluna).fill = PREENCHIMENTO_AMARELO

    if dados.get("Status Parse") == "FALHA":
        for coluna in range(1, 8):
            ws.cell(linha, coluna).fill = PREENCHIMENTO_VERMELHO


def adicionar_linha(ws, dados: dict):
    linha = ws.max_row + 1
    arquivo_pdf = dados.get("Arquivo PDF", "")

    aplicar_link_nativo(ws.cell(linha, 1), dados["Documento"], arquivo_pdf)
    ws.cell(linha, 2).value = dados["Titulo"]
    ws.cell(linha, 3).value = dados["Pasta"]
    ws.cell(linha, 4).value = dados["Emissão"]
    ws.cell(linha, 5).value = dados["Proposito de Emissão"]

    cel_data = ws.cell(linha, 6)
    cel_data.value = normalizar_data(dados["Data Envio"])
    cel_data.number_format = "@"

    aplicar_link_nativo(ws.cell(linha, 7), dados["Transmittal N°"], arquivo_pdf)

    destacar_linha(ws, linha, dados)


def registrar_log(ws_log, arquivo_pdf: str, transmittal: str, status: str, mensagem: str):
    linha = ws_log.max_row + 1
    ws_log.cell(linha, 1).value = Path(arquivo_pdf).name if arquivo_pdf else ""
    ws_log.cell(linha, 2).value = transmittal
    ws_log.cell(linha, 3).value = status
    ws_log.cell(linha, 4).value = mensagem

    if status in {"FALHA", "ERRO"}:
        for coluna in range(1, 5):
            ws_log.cell(linha, coluna).fill = PREENCHIMENTO_VERMELHO
    elif status in {"PARCIAL", "AVISO"}:
        for coluna in range(1, 5):
            ws_log.cell(linha, coluna).fill = PREENCHIMENTO_AMARELO


def ajustar_largura(ws):
    larguras = {
        "A": 20,
        "B": 65,
        "C": 22,
        "D": 25,
        "E": 32,
        "F": 15,
        "G": 18,
    }
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura


def ajustar_largura_log(ws_log):
    larguras = {
        "A": 28,
        "B": 18,
        "C": 14,
        "D": 90,
    }
    for col, largura in larguras.items():
        ws_log.column_dimensions[col].width = largura


def processar():
    if not PASTA_PDFS.exists():
        print(f"[ERRO] Pasta não encontrada: {PASTA_PDFS}")
        return

    pdfs = sorted(PASTA_PDFS.glob("*.pdf"))
    if not pdfs:
        print(f"[AVISO] Nenhum PDF encontrado em: {PASTA_PDFS}")
        return

    wb, ws, ws_log = criar_planilha_nova()

    total_pdfs_lidos = 0
    total_registros = 0

    vistos = set()

    for pdf in pdfs:
        print(f"[INFO] Processando: {pdf.name}")
        texto = extrair_texto_pdf(pdf)

        if not texto.strip():
            registrar_log(
                ws_log,
                str(pdf),
                "",
                "ERRO",
                "Não foi possível extrair texto do PDF.",
            )
            continue

        total_pdfs_lidos += 1

        try:
            registros = extrair_registros_pdf(texto, pdf.name, pdf)
        except Exception as e:
            registrar_log(
                ws_log,
                str(pdf),
                "",
                "ERRO",
                f"Erro ao interpretar conteúdo: {e}",
            )
            continue

        for dados in registros:
            chave = (dados["Documento"], dados["Transmittal N°"])
            if chave in vistos:
                registrar_log(
                    ws_log,
                    dados.get("Arquivo PDF", ""),
                    dados.get("Transmittal N°", ""),
                    "AVISO",
                    f"Duplicado ignorado para documento {dados.get('Documento', '')}.",
                )
                continue

            vistos.add(chave)
            adicionar_linha(ws, dados)
            total_registros += 1

            if dados.get("Status Parse", "OK") != "OK":
                registrar_log(
                    ws_log,
                    dados.get("Arquivo PDF", ""),
                    dados.get("Transmittal N°", ""),
                    dados.get("Status Parse", ""),
                    dados.get("Observação Parse", ""),
                )

    ajustar_largura(ws)
    ajustar_largura_log(ws_log)
    wb.save(ARQUIVO_EXCEL_NOVO)

    print("\n=== RESUMO ===")
    print(f"PDFs lidos: {total_pdfs_lidos}")
    print(f"Linhas gravadas: {total_registros}")
    print(f"Arquivo gerado: {ARQUIVO_EXCEL_NOVO}")


if __name__ == "__main__":
    processar()