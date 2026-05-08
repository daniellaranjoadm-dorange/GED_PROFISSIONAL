# -*- coding: utf-8 -*-
"""
Relatorio PCFs.py - versão ANTIBRANCO + LOG + VALIDAÇÃO FLEXÍVEL

Correções desta versão:
- Lê os arquivos PCF com openpyxl sem read_only, evitando falhas silenciosas em algumas PCFs.
- Não salva a Timeline se nenhuma linha válida for gerada.
- Não limpa uma aba antes de confirmar que há linhas novas.
- Grava em arquivo temporário e só substitui a Timeline no final, após validação.
- STATUS FINAL = último valor preenchido na coluna E, a partir de E9.
- Log detalhado no terminal para arquivos lidos, ignorados e erros.
"""

import os
import re
import shutil
import tempfile
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import IconSetRule


ARQUIVO_XLSX = r"\\virm-rgr022\FILESERVER\Projetos\05_HANDYMAX\09. Doc Control\9 - PCFs Transpetro\Timeline PCFs Transpetro.xlsx"

PASTAS_PCF = {
    "PCFs Recebidas TP": r"\\virm-rgr022\FILESERVER\Projetos\05_HANDYMAX\09. Doc Control\9 - PCFs Transpetro",
    "PCFs Respondidas CMN": r"\\virm-rgr022\FILESERVER\Projetos\05_HANDYMAX\09. Doc Control\9 - PCFs Transpetro\Respostas PCFs MARENOVA",
}

BACKUP_DIR = r"\\virm-rgr022\FILESERVER\Projetos\05_HANDYMAX\09. Doc Control\9 - PCFs Transpetro\_BACKUPS"

SHEET_SRC_RECEBIDAS = "PCFs Recebidas TP"
SHEET_SRC_RESPONDIDAS = "PCFs Respondidas CMN"
SHEET_OUT_RECEBIDAS = "Evolução - Recebidas"
SHEET_OUT_RESPONDIDAS = "Evolução - Respondidas"
SHEET_COMO_USAR = "Como usar"

ORDER_RECEBIDAS = ["0", "B", "D", "F"]
ORDER_RESPONDIDAS = ["0", "A", "B", "C", "E"]

EXTENSOES_LIVEIS = (".xlsx", ".xlsm")
EXTENSOES_EXCEL = (".xlsx", ".xlsm", ".xls", ".xlsb")

HEADERS_PCF = [
    "Caminho", "PCF LINK", "N º PCF", "Nº DOCUMENTO", "TITULO",
    "Revisão da PCF", "Data Recebimento", "Open Comments",
    "Qtd Comentarios", "SRC_OPEN", "SRC_TOTAL", "STATUS FINAL",
]


# =========================
# LOG NO TERMINAL
# =========================
LOG_DETALHADO = True
MOSTRAR_CADA_PCF_LIDA = True
MOSTRAR_CADA_PCF_IGNORADA = True
MOSTRAR_CADA_ERRO = True


def agora_log():
    return datetime.now().strftime("%H:%M:%S")


def log(msg=""):
    print(f"[{agora_log()}] {msg}")


def log_secao(titulo):
    print("\n" + "=" * 90)
    log(titulo)
    print("=" * 90)


def log_ok(msg):
    log(f"OK  | {msg}")


def log_aviso(msg):
    log(f"AVISO | {msg}")


def log_erro(msg):
    log(f"ERRO | {msg}")


def log_debug(msg):
    if LOG_DETALHADO:
        log(f"DEBUG | {msg}")


def safe_str(v):
    return "" if v is None else str(v).strip()


def norm_text(v):
    return re.sub(r"\s+", " ", safe_str(v).upper())


def norm_rev(v):
    s = safe_str(v).upper()
    if s.endswith(".0"):
        s = s[:-2]
    if s.startswith("R") and len(s) > 1:
        s = s[1:]
    return s


def make_backup(path):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    base = os.path.splitext(os.path.basename(path))[0]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"{base}_BACKUP_{ts}.xlsx")
    shutil.copy2(path, backup_path)
    return backup_path


def restaurar_backup(backup_path, destino):
    shutil.copy2(backup_path, destino)


def delete_if_exists(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])


def existing_headers(ws):
    headers = []
    for c in range(1, ws.max_column + 1):
        h = safe_str(ws.cell(row=1, column=c).value)
        if h:
            headers.append(h)
    return headers


def ensure_headers(ws, required_headers):
    headers = existing_headers(ws)
    if not headers:
        headers = required_headers[:]
    for h in required_headers:
        if h not in headers:
            headers.append(h)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    return headers


def find_col_by_header(ws, header_name):
    target = safe_str(header_name).lower()
    for c in range(1, ws.max_column + 1):
        if safe_str(ws.cell(row=1, column=c).value).lower() == target:
            return c
    return None


def eh_pasta_ignorada(dirname):
    n = dirname.upper()
    return n in {"_BACKUPS", "BACKUP", "BACKUPS"} or "BACKUP" in n


def listar_arquivos_pcf(pasta, aba):
    encontrados = []
    ignorados = []
    total_arquivos = 0
    excel_encontrados = 0

    pasta_respostas = os.path.normcase(os.path.abspath(PASTAS_PCF["PCFs Respondidas CMN"]))

    log_secao(f"Varredura da pasta - {aba}")
    log(f"Pasta: {pasta}")

    for root, dirs, files in os.walk(pasta):
        dirs[:] = [d for d in dirs if not eh_pasta_ignorada(d)]

        if aba == "PCFs Recebidas TP":
            dirs[:] = [
                d for d in dirs
                if os.path.normcase(os.path.abspath(os.path.join(root, d))) != pasta_respostas
            ]

        for nome in files:
            total_arquivos += 1
            caminho = os.path.join(root, nome)
            nome_up = nome.upper()
            ext = os.path.splitext(nome)[1].lower()

            motivo_ignorado = ""

            if nome.startswith("~$"):
                motivo_ignorado = "temporário do Excel"
            elif ext not in EXTENSOES_EXCEL:
                continue
            else:
                excel_encontrados += 1

                if os.path.normcase(os.path.abspath(caminho)) == os.path.normcase(os.path.abspath(ARQUIVO_XLSX)):
                    motivo_ignorado = "é a própria Timeline"
                elif "TIMELINE PCFS TRANSPETRO" in nome_up:
                    motivo_ignorado = "nome de Timeline"
                elif "PCF" not in nome_up:
                    motivo_ignorado = "Excel sem PCF no nome"
                elif ext not in EXTENSOES_LIVEIS:
                    motivo_ignorado = f"extensão {ext} não lida pelo openpyxl"
                else:
                    encontrados.append(caminho)
                    if LOG_DETALHADO:
                        log_debug(f"Candidata: {nome}")
                    continue

            if motivo_ignorado:
                ignorados.append((nome, motivo_ignorado))
                if MOSTRAR_CADA_PCF_IGNORADA and LOG_DETALHADO:
                    log_debug(f"Ignorado na varredura: {nome} :: {motivo_ignorado}")

    encontrados = sorted(set(encontrados), key=lambda x: x.lower())

    log(f"Total de arquivos vistos: {total_arquivos}")
    log(f"Arquivos Excel vistos: {excel_encontrados}")
    log(f"PCFs candidatas para leitura: {len(encontrados)}")
    log(f"Ignorados na varredura: {len(ignorados)}")

    if encontrados[:10]:
        log("Primeiros candidatos:")
        for x in encontrados[:10]:
            print("  +", x)

    if ignorados[:15]:
        log("Primeiros ignorados:")
        for nome, motivo in ignorados[:15]:
            print(f"  - {nome} :: {motivo}")

    return encontrados

def primeira_aba_util(wb_pcf):
    for ws in wb_pcf.worksheets:
        if getattr(ws, "sheet_state", "visible") == "visible":
            return ws
    return wb_pcf.worksheets[0]


def get_by_label(ws, label, max_row=15, max_col=15):
    alvo = norm_text(label)
    for row in range(1, min(ws.max_row, max_row) + 1):
        for col in range(1, min(ws.max_column, max_col) + 1):
            if norm_text(ws.cell(row=row, column=col).value) == alvo:
                # na PCF os valores costumam estar 1, 2 ou 3 colunas à direita
                for offset in (2, 1, 3, 4):
                    v = ws.cell(row=row, column=col + offset).value
                    if safe_str(v):
                        return v
    return ""


def extrair_status_final(ws):
    status_final = ""
    for row in range(9, ws.max_row + 1):
        v = ws.cell(row=row, column=5).value  # coluna E
        if safe_str(v):
            status_final = safe_str(v)
    return status_final


def extrair_qtd_e_open_comments(ws):
    comment_status_col = None
    header_row = None

    for r in range(1, min(ws.max_row, 60) + 1):
        for c in range(1, min(ws.max_column, 30) + 1):
            if norm_text(ws.cell(r, c).value) == "COMMENT STATUS":
                comment_status_col = c
                header_row = r
                break
        if comment_status_col:
            break

    open_count = 0
    total_status = 0

    if comment_status_col:
        for r in range(header_row + 1, ws.max_row + 1):
            st = norm_text(ws.cell(r, comment_status_col).value)
            if st:
                total_status += 1
                if st == "OPEN":
                    open_count += 1

    total_items = 0
    for r in range(21, ws.max_row + 1):
        item = ws.cell(r, 1).value
        if isinstance(item, (int, float)) and item > 0:
            total_items += 1

    return open_count, max(total_status, total_items)


def extrair_data_recebimento(ws):
    data = get_by_label(ws, "Date", max_row=20, max_col=20)
    if data:
        return data

    ultima = ""
    for row in range(9, min(ws.max_row, 16) + 1):
        for col in (9, 12):
            v = ws.cell(row=row, column=col).value
            if safe_str(v):
                ultima = v
    return ultima


def extrair_revisao_pcf(ws, caminho):
    rev = get_by_label(ws, "PCF Rev.", max_row=20, max_col=20)
    if safe_str(rev):
        return norm_rev(rev)

    nome = os.path.splitext(os.path.basename(caminho))[0].upper()
    m = re.search(r"_R([A-Z0-9]+)$", nome)
    if m:
        raw = m.group(1)
        if raw == "0":
            return "0"
        return raw[-1]

    return ""


def parece_pcf_valida(ws, caminho):
    """
    Validação flexível:
    - Se o nome contém PCF, considera candidata válida.
    - Se não contém PCF no nome, tenta validar pelo conteúdo em uma área maior da planilha.
    Isso evita ignorar PCFs novas com cabeçalho deslocado.
    """
    nome = os.path.basename(caminho).upper()

    if nome.startswith("LISTA "):
        return False

    if "PCF" in nome:
        return True

    tem_pcf_no = bool(safe_str(get_by_label(ws, "PCF No.", max_row=20, max_col=20)))
    tem_plan_no = bool(safe_str(get_by_label(ws, "Plan No.", max_row=20, max_col=20)))
    tem_status_em_e = bool(extrair_status_final(ws))

    return tem_pcf_no or tem_plan_no or tem_status_em_e

def extrair_dados_pcf(caminho):
    wb_pcf = load_workbook(caminho, data_only=True, read_only=False)
    try:
        ws = primeira_aba_util(wb_pcf)

        if not parece_pcf_valida(ws, caminho):
            return None

        # Busca ampliada para aceitar PCFs com cabeçalho deslocado.
        plan_no = safe_str(get_by_label(ws, "Plan No.", max_row=20, max_col=20))
        pcf_no = safe_str(get_by_label(ws, "PCF No.", max_row=20, max_col=20))
        titulo = safe_str(get_by_label(ws, "Plan Title", max_row=20, max_col=20))

        if not pcf_no:
            base = os.path.splitext(os.path.basename(caminho))[0]
            pcf_no = re.sub(r"_R[A-Z0-9]+$", "", base, flags=re.IGNORECASE)

        if not plan_no and pcf_no.upper().startswith("PCF-"):
            plan_no = pcf_no[4:]

        open_comments, qtd_comentarios = extrair_qtd_e_open_comments(ws)
        status_final = extrair_status_final(ws)

        row = {
            "Caminho": caminho,
            "PCF LINK": os.path.splitext(os.path.basename(caminho))[0],
            "N º PCF": pcf_no,
            "Nº DOCUMENTO": plan_no,
            "TITULO": titulo,
            "Revisão da PCF": extrair_revisao_pcf(ws, caminho),
            "Data Recebimento": extrair_data_recebimento(ws),
            "Open Comments": open_comments,
            "Qtd Comentarios": qtd_comentarios,
            "SRC_OPEN": "Comment Status = OPEN",
            "SRC_TOTAL": "Comment Status/Item count",
            "STATUS FINAL": status_final,
        }

        return row
    finally:
        wb_pcf.close()

def thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def aplicar_estilo_tabela(ws, headers):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill("solid", fgColor="1F4E79")
    border = thin_border()

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "Caminho": 60, "PCF LINK": 35, "N º PCF": 30, "Nº DOCUMENTO": 32,
        "TITULO": 55, "Revisão da PCF": 16, "Data Recebimento": 18,
        "Open Comments": 16, "Qtd Comentarios": 16, "SRC_OPEN": 26,
        "SRC_TOTAL": 26, "STATUS FINAL": 20,
    }
    for c, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = widths.get(h, max(14, min(40, len(safe_str(h)) + 4)))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)



def aplicar_hyperlink_pcf(cell, caminho):
    """
    Aplica hyperlink real do Excel para o arquivo PCF.
    O texto exibido fica como o nome da PCF, mas o clique abre o caminho completo.
    """
    caminho = safe_str(caminho)
    if not caminho:
        return False

    cell.hyperlink = caminho
    cell.style = "Hyperlink"
    return True


def limpar_e_preencher_aba(ws, rows, headers):
    # Só é chamada depois de rows > 0.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    col_map = {h: i + 1 for i, h in enumerate(headers)}
    links_aplicados = 0

    for r_idx, row in enumerate(rows, start=2):
        for h in headers:
            cell = ws.cell(row=r_idx, column=col_map[h], value=row.get(h, ""))

            # Coluna B / PCF LINK: manter o texto curto, mas criar hyperlink real
            # apontando para o caminho completo do arquivo PCF.
            if h == "PCF LINK":
                caminho_pcf = row.get("Caminho", "")
                if aplicar_hyperlink_pcf(cell, caminho_pcf):
                    links_aplicados += 1

    aplicar_estilo_tabela(ws, headers)
    return links_aplicados


def atualizar_abas_pcf(wb):
    resumo = {}

    for aba, pasta in PASTAS_PCF.items():
        log_secao(f"Atualizando aba: {aba}")

        ws = wb[aba] if aba in wb.sheetnames else wb.create_sheet(aba)
        headers = ensure_headers(ws, HEADERS_PCF)

        arquivos = listar_arquivos_pcf(pasta, aba)

        if len(arquivos) == 0:
            log_aviso(f"0 PCFs encontradas em '{aba}'. Aba preservada.")
            resumo[aba] = {"arquivos": 0, "linhas": 0, "erros": ["0 arquivos encontrados; aba preservada."]}
            continue

        rows = []
        erros = []
        ignorados_modelo = 0

        for idx, caminho in enumerate(arquivos, start=1):
            nome = os.path.basename(caminho)
            try:
                log(f"Lendo {idx}/{len(arquivos)}: {nome}")
                row = extrair_dados_pcf(caminho)

                if row is None:
                    ignorados_modelo += 1
                    msg = f"Ignorada como não-PCF/modelo inválido: {caminho}"
                    if MOSTRAR_CADA_PCF_IGNORADA:
                        log_aviso(msg)
                    continue

                rows.append(row)

                if MOSTRAR_CADA_PCF_LIDA:
                    log_ok(
                        f"Lida: {row.get('PCF LINK', '')} | "
                        f"Doc: {row.get('Nº DOCUMENTO', '')} | "
                        f"Rev: {row.get('Revisão da PCF', '')} | "
                        f"Qtd comentários: {row.get('Qtd Comentarios', '')} | "
                        f"Status final: {row.get('STATUS FINAL', '')}"
                    )

            except Exception as e:
                erro = f"{caminho} :: {type(e).__name__}: {e}"
                erros.append(erro)
                if MOSTRAR_CADA_ERRO:
                    log_erro(erro)

        if len(rows) == 0:
            log_aviso(f"{len(arquivos)} arquivos encontrados, mas 0 PCFs lidas em '{aba}'. Aba preservada.")
            if erros:
                log("Primeiros erros:")
                for erro in erros[:20]:
                    print(" -", erro)

            resumo[aba] = {"arquivos": len(arquivos), "linhas": 0, "erros": erros or ["0 arquivos lidos; aba preservada."]}
            continue

        rows.sort(key=lambda x: (safe_str(x.get("Nº DOCUMENTO")), safe_str(x.get("Revisão da PCF")), safe_str(x.get("PCF LINK"))))
        links_aplicados = limpar_e_preencher_aba(ws, rows, headers)

        log_ok(f"Aba '{aba}' preenchida com {len(rows)} linhas.")
        log_ok(f"Links aplicados em '{aba}': {links_aplicados}")
        log(f"Ignorados como não-PCF/modelo inválido: {ignorados_modelo}")
        log(f"Erros de leitura: {len(erros)}")

        resumo[aba] = {
            "arquivos": len(arquivos),
            "linhas": len(rows),
            "links": links_aplicados,
            "erros": erros,
            "ignorados_modelo": ignorados_modelo,
        }

    return resumo

def build_map_from_source(ws, order_list):
    col_doc = find_col_by_header(ws, "Nº DOCUMENTO")
    col_pcf = find_col_by_header(ws, "N º PCF") or find_col_by_header(ws, "Nº PCF")
    col_titulo = find_col_by_header(ws, "TITULO") or find_col_by_header(ws, "TÍTULO")
    col_rev = find_col_by_header(ws, "Revisão da PCF")
    col_qtd = find_col_by_header(ws, "Qtd Comentarios")

    if not col_doc or not col_rev or not col_qtd:
        return {}

    data = {}
    for r in range(2, ws.max_row + 1):
        doc = safe_str(ws.cell(row=r, column=col_doc).value)
        if not doc:
            continue

        rev = norm_rev(ws.cell(row=r, column=col_rev).value)
        if rev not in order_list:
            continue

        try:
            qtd = int(ws.cell(row=r, column=col_qtd).value or 0)
        except Exception:
            qtd = 0

        pcf = safe_str(ws.cell(row=r, column=col_pcf).value) if col_pcf else ""
        titulo = safe_str(ws.cell(row=r, column=col_titulo).value) if col_titulo else ""

        data.setdefault(doc, {"doc": doc, "pcf": pcf, "titulo": titulo, "rev_map": {}})
        data[doc]["rev_map"][rev] = max(qtd, data[doc]["rev_map"].get(rev, 0))

    return data


def compute_evolution_rows(data_map, order_list):
    rows = []
    for _, info in data_map.items():
        rev_map = info["rev_map"]
        available = [rev for rev in order_list if rev in rev_map]
        if not available:
            continue

        de, para = available[0], available[-1]
        row = {
            "Nº DOCUMENTO": info.get("doc", ""),
            "Nº PCF": info.get("pcf", ""),
            "TITULO": info.get("titulo", ""),
            "De": de,
            "Para": para,
        }
        for rev in order_list:
            row[rev] = rev_map.get(rev, None)
        for i in range(1, len(order_list)):
            a, b = order_list[i - 1], order_list[i]
            row[f"Δ {b}-{a}"] = (rev_map[b] - rev_map[a]) if (a in rev_map and b in rev_map) else None
        row["Δ Total"] = (rev_map.get(para, 0) - rev_map.get(de, 0))
        rows.append(row)

    rows.sort(key=lambda x: (x.get("Δ Total", 0), x.get("Nº DOCUMENTO", "")), reverse=True)
    return rows


def build_headers(order_list):
    return ["Nº DOCUMENTO", "Nº PCF", "TITULO", "De", "Para"] + order_list[:] + [f"Δ {order_list[i]}-{order_list[i-1]}" for i in range(1, len(order_list))] + ["Δ Total"]


def write_evolution_sheet(ws, headers, rows):
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(h, None))
    aplicar_estilo_tabela(ws, headers)


def apply_iconsets_openpyxl(ws):
    if ws.max_row < 2:
        return
    rule = IconSetRule(icon_style="3Arrows", type="num", values=[0, 1e-9], showValue=True, reverse=False)
    for c in range(1, ws.max_column + 1):
        h = safe_str(ws.cell(row=1, column=c).value)
        if h.startswith("Δ"):
            col = ws.cell(row=1, column=c).column_letter
            ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", rule)


def recriar_evolucao(wb):
    delete_if_exists(wb, SHEET_OUT_RECEBIDAS)
    delete_if_exists(wb, SHEET_OUT_RESPONDIDAS)

    ws_rec_src = wb[SHEET_SRC_RECEBIDAS] if SHEET_SRC_RECEBIDAS in wb.sheetnames else wb.create_sheet(SHEET_SRC_RECEBIDAS)
    ws_res_src = wb[SHEET_SRC_RESPONDIDAS] if SHEET_SRC_RESPONDIDAS in wb.sheetnames else wb.create_sheet(SHEET_SRC_RESPONDIDAS)

    rows_rec = compute_evolution_rows(build_map_from_source(ws_rec_src, ORDER_RECEBIDAS), ORDER_RECEBIDAS)
    rows_res = compute_evolution_rows(build_map_from_source(ws_res_src, ORDER_RESPONDIDAS), ORDER_RESPONDIDAS)

    ws_rec = wb.create_sheet(SHEET_OUT_RECEBIDAS)
    ws_res = wb.create_sheet(SHEET_OUT_RESPONDIDAS)

    write_evolution_sheet(ws_rec, build_headers(ORDER_RECEBIDAS), rows_rec)
    write_evolution_sheet(ws_res, build_headers(ORDER_RESPONDIDAS), rows_res)

    apply_iconsets_openpyxl(ws_rec)
    apply_iconsets_openpyxl(ws_res)


def create_como_usar_sheet(wb, resumo):
    delete_if_exists(wb, SHEET_COMO_USAR)
    ws = wb.create_sheet(SHEET_COMO_USAR)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 120

    ws["A1"] = "Como usar - Timeline PCFs"
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    lines = ["Atualização executada pelo script local.", "", "Pastas varridas:"]
    for aba, pasta in PASTAS_PCF.items():
        info = resumo.get(aba, {})
        lines.append(f"• {aba}: {pasta}")
        lines.append(
            f"  Arquivos candidatos: {info.get('arquivos', 0)} | "
            f"Linhas gravadas: {info.get('linhas', 0)} | "
            f"Ignorados como não-PCF: {info.get('ignorados_modelo', 0)} | "
            f"Erros: {len(info.get('erros', []))}"
        )

    lines += [
        "",
        "Regra principal:",
        "• STATUS FINAL é preenchido com o último valor não vazio da coluna E da PCF, começando em E9.",
        "",
        "Trava antibanco:",
        "• Se nenhuma linha válida for gerada, o script aborta e restaura o backup.",
        "• A Timeline só é substituída depois de salvar e validar um arquivo temporário.",
    ]

    for i, line in enumerate(lines, start=3):
        ws.cell(row=i, column=1, value=line)
        ws.cell(row=i, column=1).font = Font(name="Arial", size=10)
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="top")


def contar_linhas_dados(wb, aba):
    if aba not in wb.sheetnames:
        return 0
    ws = wb[aba]
    return max(ws.max_row - 1, 0)


def salvar_com_trava(wb, resumo, backup_path):
    total_linhas = sum(info.get("linhas", 0) for info in resumo.values())

    if total_linhas == 0:
        wb.close()
        restaurar_backup(backup_path, ARQUIVO_XLSX)
        raise RuntimeError(
            "ABORTADO: nenhuma linha válida foi gerada. "
            "A Timeline original foi restaurada pelo backup e NÃO foi substituída por planilha vazia."
        )

    pasta_destino = os.path.dirname(ARQUIVO_XLSX)
    fd, temp_path = tempfile.mkstemp(prefix="Timeline_PCFs_TEMP_", suffix=".xlsx", dir=pasta_destino)
    os.close(fd)

    try:
        log(f"Salvando arquivo temporário: {temp_path}")
        wb.save(temp_path)
        wb.close()

        wb_check = load_workbook(temp_path, read_only=True, data_only=True)
        try:
            linhas_rec = contar_linhas_dados(wb_check, SHEET_SRC_RECEBIDAS)
            linhas_res = contar_linhas_dados(wb_check, SHEET_SRC_RESPONDIDAS)
        finally:
            wb_check.close()

        if linhas_rec + linhas_res == 0:
            restaurar_backup(backup_path, ARQUIVO_XLSX)
            raise RuntimeError(
                "ABORTADO: validação final encontrou 0 linhas nas abas principais. "
                "Backup restaurado."
            )

        log("Substituindo Timeline original pelo arquivo temporário validado...")
        shutil.move(temp_path, ARQUIVO_XLSX)
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass


def main():
    log_secao("Início da atualização da Timeline PCFs")
    log(f"Arquivo Timeline: {ARQUIVO_XLSX}")

    if not os.path.exists(ARQUIVO_XLSX):
        raise FileNotFoundError(f"Arquivo não encontrado: {ARQUIVO_XLSX}")

    for aba, pasta in PASTAS_PCF.items():
        if not os.path.exists(pasta):
            raise FileNotFoundError(f"Pasta não encontrada para '{aba}': {pasta}")

    backup_path = make_backup(ARQUIVO_XLSX)
    log_ok(f"Backup criado: {backup_path}")

    log("Abrindo Timeline...")
    wb = load_workbook(ARQUIVO_XLSX)
    resumo = atualizar_abas_pcf(wb)

    total_linhas = sum(info.get("linhas", 0) for info in resumo.values())
    if total_linhas == 0:
        wb.close()
        restaurar_backup(backup_path, ARQUIVO_XLSX)
        raise RuntimeError(
            "ABORTADO: 0 linhas válidas extraídas das PCFs. "
            "Nada foi salvo. A planilha foi restaurada pelo backup."
        )

    log_secao("Recriando abas de evolução")
    recriar_evolucao(wb)
    log_ok("Abas de evolução recriadas.")
    log("Recriando aba Como usar...")
    create_como_usar_sheet(wb, resumo)
    log_ok("Aba Como usar recriada.")
    log_secao("Salvando com trava de segurança")
    salvar_com_trava(wb, resumo, backup_path)
    log_ok("Arquivo salvo e validado com sucesso.")

    log_secao("Atualização concluída com segurança")
    for aba, info in resumo.items():
        log(f"{aba}: {info.get('linhas', 0)} linhas gravadas de {info.get('arquivos', 0)} arquivos candidatos.")
        log(f"Ignorados como não-PCF/modelo inválido: {info.get('ignorados_modelo', 0)}")
        if info.get("erros"):
            print(f"Avisos/erros em {aba}:")
            for erro in info["erros"][:30]:
                print(" -", erro)


if __name__ == "__main__":
    main()
