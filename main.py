import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import Browser, Page, Playwright, TimeoutError, sync_playwright


COLUNAS_PADRAO = [
    "Area",
    "No / Discipline",
    "P",
    "Status",
    "Doc Req / Data Type",
    "Responsible",
    "Contract Delivery Week No",
    "Contract Delivery Date",
    "Prelim Delivery",
    "First Delivery",
    "Input Updated",
    "Comments",
    "Yard Comments",
]

COLUNAS_DATA = [
    "Contract Delivery Date",
    "Prelim Delivery",
    "First Delivery",
    "Input Updated",
]

MAPA_COLUNAS = {
    "area": "Area",
    "discipline": "No / Discipline",
    "no": "No / Discipline",
    "no / discipline": "No / Discipline",
    "p": "P",
    "status": "Status",
    "doc req": "Doc Req / Data Type",
    "document requirement": "Doc Req / Data Type",
    "doc req / data type": "Doc Req / Data Type",
    "data type": "Doc Req / Data Type",
    "responsible": "Responsible",
    "contract delivery week no": "Contract Delivery Week No",
    "contract week": "Contract Delivery Week No",
    "contract delivery date": "Contract Delivery Date",
    "prelim delivery": "Prelim Delivery",
    "first delivery": "First Delivery",
    "input updated": "Input Updated",
    "comments": "Comments",
    "yard comments": "Yard Comments",
}


def configurar_logs() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%d/%m/%Y %H:%M:%S",
    )


def carregar_configuracoes() -> Dict[str, str]:
    load_dotenv()
    config = {
        "site_url": os.getenv("SITE_URL", "").strip(),
        "usuario": os.getenv("USUARIO", "").strip(),
        "senha": os.getenv("SENHA", "").strip(),
        "pasta_saida": os.getenv("PASTA_SAIDA", "").strip(),
        "headless": os.getenv("HEADLESS", "true").strip().lower() == "true",
    }

    faltantes = [chave for chave, valor in config.items() if chave != "headless" and not valor]
    if faltantes:
        raise ValueError(
            f"Configuração inválida: variáveis ausentes no .env -> {', '.join(faltantes)}"
        )

    pasta = Path(config["pasta_saida"])
    pasta.mkdir(parents=True, exist_ok=True)
    return config


def abrir_navegador(playwright: Playwright, headless: bool) -> Tuple[Browser, Page]:
    logging.info("Abrindo navegador (headless=%s)...", headless)
    browser = playwright.chromium.launch(headless=headless)
    context = browser.new_context(accept_downloads=True)
    page = context.new_page()
    return browser, page


def _normalizar_texto(texto: str) -> str:
    texto = re.sub(r"\s+", " ", texto.strip().lower())
    return texto


def _detectar_campo(page: Page, nomes: List[str], tipo: Optional[str] = None):
    for nome in nomes:
        seletores = [
            f'input[name*="{nome}" i]',
            f'input[id*="{nome}" i]',
            f'input[placeholder*="{nome}" i]',
            f'textbox[name*="{nome}" i]',
        ]
        for seletor in seletores:
            locator = page.locator(seletor).first
            if locator.count() > 0 and locator.is_visible():
                if tipo:
                    try:
                        locator.evaluate("(el, t) => el.type = t", tipo)
                    except Exception:
                        pass
                return locator
    return None


def _url_parece_logada(url: str) -> bool:
    url_lower = url.lower()
    return "login" not in url_lower and "signin" not in url_lower


def _aguardar_autenticacao_manual(page: Page) -> None:
    logging.warning(
        "Autenticação manual pode ser necessária (CAPTCHA/2FA). "
        "Finalize no navegador aberto e pressione ENTER no terminal para continuar."
    )
    input()
    page.wait_for_load_state("networkidle", timeout=60000)


def fazer_login(page: Page, usuario: str, senha: str, site_url: str, headless: bool) -> None:
    logging.info("Acessando URL inicial...")
    try:
        page.goto(site_url, wait_until="domcontentloaded", timeout=60000)
    except TimeoutError as exc:
        raise RuntimeError("Site fora do ar ou sem resposta no tempo esperado.") from exc

    logging.info("Tentando identificar tela de login...")
    campo_usuario = _detectar_campo(page, ["user", "email", "login", "username"])
    campo_senha = _detectar_campo(page, ["password", "senha"], tipo="password")

    if campo_usuario and campo_senha:
        logging.info("Preenchendo credenciais...")
        campo_usuario.fill(usuario)
        campo_senha.fill(senha)

        seletor_botao = [
            "button:has-text('Login')",
            "button:has-text('Sign in')",
            "button:has-text('Entrar')",
            "input[type='submit']",
            "button[type='submit']",
        ]

        clicou = False
        for seletor in seletor_botao:
            botao = page.locator(seletor).first
            if botao.count() > 0 and botao.is_visible():
                botao.click()
                clicou = True
                break

        if not clicou:
            raise RuntimeError("Falha de login: botão de envio não encontrado.")

        try:
            page.wait_for_load_state("networkidle", timeout=30000)
        except TimeoutError:
            logging.warning("Página demorou a estabilizar após login; continuando validação.")

    if not _url_parece_logada(page.url):
        if not headless:
            _aguardar_autenticacao_manual(page)
        if not _url_parece_logada(page.url):
            raise RuntimeError("Falha de login: sessão não autenticada.")


def acessar_pagina_projeto(page: Page, site_url: str) -> None:
    logging.info("Acessando página do projeto...")
    try:
        page.goto(site_url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_load_state("networkidle", timeout=45000)
    except TimeoutError as exc:
        raise RuntimeError("Página do projeto não carregada no tempo esperado.") from exc

    if "projectinformation" not in page.url.lower():
        raise RuntimeError("Página de Project Information não foi carregada corretamente.")


def localizar_relatorio(page: Page) -> None:
    logging.info("Localizando relatório de documentação (DocReq)...")
    seletores_relatorio = [
        "text=/DocReq/i",
        "text=/Document Requirement/i",
        "text=/Project Information/i",
        "text=/Documentation/i",
    ]
    for seletor in seletores_relatorio:
        alvo = page.locator(seletor).first
        if alvo.count() > 0:
            try:
                if alvo.is_visible():
                    alvo.click(timeout=3000)
                    page.wait_for_timeout(1200)
                    return
            except Exception:
                continue

    logging.info("Seção específica não encontrada por clique; seguindo para tentativa direta de extração.")


def _tentar_download_relatorio(page: Page, pasta_saida: Path) -> Optional[Path]:
    logging.info("Verificando possibilidade de exportação direta (Excel/CSV)...")
    seletores_exportar = [
        "button:has-text('Export')",
        "button:has-text('Excel')",
        "button:has-text('CSV')",
        "a:has-text('Export')",
        "a:has-text('Excel')",
        "a:has-text('CSV')",
        "[title*='Export' i]",
        "[aria-label*='Export' i]",
    ]

    for seletor in seletores_exportar:
        botao = page.locator(seletor).first
        if botao.count() == 0 or not botao.is_visible():
            continue

        try:
            with page.expect_download(timeout=15000) as download_info:
                botao.click()
            download = download_info.value
            nome_arquivo = download.suggested_filename
            caminho = pasta_saida / nome_arquivo
            download.save_as(str(caminho))
            logging.info("Download concluído: %s", caminho)
            return caminho
        except TimeoutError:
            logging.warning("Clique em exportação sem download detectado (seletor: %s).", seletor)
        except Exception as exc:
            logging.warning("Falha ao tentar exportar com %s: %s", seletor, exc)

    return None


def _ler_arquivo_exportado(caminho: Path) -> pd.DataFrame:
    sufixo = caminho.suffix.lower()
    if sufixo == ".csv":
        return pd.read_csv(caminho)
    if sufixo in {".xls", ".xlsx"}:
        return pd.read_excel(caminho)
    raise RuntimeError(f"Formato de exportação não suportado: {sufixo}")


def _extrair_tabelas_html(page: Page) -> List[pd.DataFrame]:
    frames = pd.read_html(page.content())
    return [frame for frame in frames if not frame.empty]


def _proxima_pagina(page: Page) -> bool:
    seletores_next = [
        "button:has-text('Next')",
        "a:has-text('Next')",
        "button[aria-label*='next' i]",
        "a[aria-label*='next' i]",
        "li.next:not(.disabled) a",
    ]
    for seletor in seletores_next:
        btn = page.locator(seletor).first
        if btn.count() == 0 or not btn.is_visible():
            continue
        classes = (btn.get_attribute("class") or "").lower()
        disabled = btn.get_attribute("disabled")
        aria_disabled = (btn.get_attribute("aria-disabled") or "").lower()
        if "disabled" in classes or disabled is not None or aria_disabled == "true":
            continue
        try:
            btn.click()
            page.wait_for_load_state("networkidle", timeout=20000)
            page.wait_for_timeout(1000)
            return True
        except Exception:
            continue
    return False


def extrair_dados(page: Page, pasta_saida: Path) -> pd.DataFrame:
    arquivo_exportado = _tentar_download_relatorio(page, pasta_saida)
    if arquivo_exportado:
        logging.info("Lendo dados do arquivo exportado...")
        df_exportado = _ler_arquivo_exportado(arquivo_exportado)
        if df_exportado.empty:
            raise RuntimeError("Relatório exportado está vazio.")
        return df_exportado

    logging.warning("Botão de exportação não encontrado ou sem download. Tentando extração da tabela HTML...")
    paginas_lidas = 0
    blocos: List[pd.DataFrame] = []

    while True:
        tabelas = _extrair_tabelas_html(page)
        if tabelas:
            melhor = max(tabelas, key=lambda frame: frame.shape[0] * frame.shape[1])
            if melhor.shape[0] > 0:
                blocos.append(melhor)
                paginas_lidas += 1
                logging.info("Tabela capturada na página %s com %s linhas.", paginas_lidas, melhor.shape[0])

        if not _proxima_pagina(page):
            break

    if not blocos:
        raise RuntimeError("Tabela não encontrada na página.")

    df = pd.concat(blocos, ignore_index=True).drop_duplicates()
    if df.empty:
        raise RuntimeError("Relatório vazio após extração.")
    return df


def _mapear_colunas(colunas_origem: List[str]) -> Dict[str, str]:
    mapeamento = {}
    for coluna in colunas_origem:
        normalizada = _normalizar_texto(coluna)
        melhor = None
        for chave, destino in MAPA_COLUNAS.items():
            if chave in normalizada or normalizada in chave:
                melhor = destino
                break
        if melhor:
            mapeamento[coluna] = melhor
    return mapeamento


def tratar_dados(df: pd.DataFrame) -> pd.DataFrame:
    logging.info("Tratando e padronizando dados...")
    df.columns = [str(c).strip() for c in df.columns]
    mapa = _mapear_colunas(list(df.columns))
    df = df.rename(columns=mapa)

    for coluna in COLUNAS_PADRAO:
        if coluna not in df.columns:
            df[coluna] = ""

    df_final = df[COLUNAS_PADRAO].copy()

    for coluna_data in COLUNAS_DATA:
        df_final[coluna_data] = pd.to_datetime(df_final[coluna_data], errors="coerce")

    return df_final


def _aplicar_formatacao_excel(ws, ultima_linha: int) -> None:
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:M{max(3, ultima_linha)}"

    larguras = {
        "A": 14,
        "B": 18,
        "C": 6,
        "D": 14,
        "E": 30,
        "F": 18,
        "G": 24,
        "H": 20,
        "I": 16,
        "J": 16,
        "K": 16,
        "L": 32,
        "M": 32,
    }
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    cabecalho_fill = PatternFill("solid", fgColor="1F4E78")
    cabecalho_font = Font(color="FFFFFF", bold=True)
    borda = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in ws[3]:
        cell.fill = cabecalho_fill
        cell.font = cabecalho_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda

    colunas_wrap = {"E", "L", "M"}
    colunas_data_idx = {8, 9, 10, 11}  # H, I, J, K
    for row in ws.iter_rows(min_row=4, max_row=ultima_linha, min_col=1, max_col=13):
        for cell in row:
            cell.border = borda
            if get_column_letter(cell.column) in colunas_wrap:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")
            if cell.column in colunas_data_idx and cell.value:
                cell.number_format = "dd/mm/yyyy"


def gerar_excel(df: pd.DataFrame, pasta_saida: Path) -> Path:
    logging.info("Gerando arquivo Excel final...")
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    arquivo_saida = pasta_saida / f"DocReq_Report_Atualizado_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Table 1"

    ws["A2"] = f"Document Requirement Report - Atualizado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:M2")

    for idx, coluna in enumerate(COLUNAS_PADRAO, start=1):
        ws.cell(row=3, column=idx, value=coluna)

    linha = 4
    for _, row in df.iterrows():
        for col_idx, coluna in enumerate(COLUNAS_PADRAO, start=1):
            valor = row[coluna]
            if pd.isna(valor):
                valor = None
            ws.cell(row=linha, column=col_idx, value=valor)
        linha += 1

    ultima_linha = max(3, linha - 1)
    _aplicar_formatacao_excel(ws, ultima_linha)

    try:
        wb.save(arquivo_saida)
    except Exception as exc:
        raise RuntimeError(f"Erro ao salvar o Excel: {exc}") from exc

    logging.info("Excel salvo com sucesso: %s", arquivo_saida)
    return arquivo_saida


def validar_colunas_finais(df: pd.DataFrame) -> None:
    colunas = list(df.columns)
    if colunas != COLUNAS_PADRAO:
        raise RuntimeError(
            "Colunas finais fora do padrão exigido.\n"
            f"Esperado: {COLUNAS_PADRAO}\n"
            f"Obtido: {colunas}"
        )


def main() -> None:
    configurar_logs()
    logging.info("Iniciando automação DocReq...")

    try:
        cfg = carregar_configuracoes()
        pasta_saida = Path(cfg["pasta_saida"])

        with sync_playwright() as playwright:
            browser, page = abrir_navegador(playwright, cfg["headless"])
            try:
                fazer_login(page, cfg["usuario"], cfg["senha"], cfg["site_url"], cfg["headless"])
                acessar_pagina_projeto(page, cfg["site_url"])
                localizar_relatorio(page)
                dados_brutos = extrair_dados(page, pasta_saida)
                dados_tratados = tratar_dados(dados_brutos)
                validar_colunas_finais(dados_tratados)
                gerar_excel(dados_tratados, pasta_saida)
            finally:
                browser.close()

        logging.info("Processo concluído com sucesso.")
    except Exception as exc:
        logging.error("Falha na execução: %s", exc)
        raise


if __name__ == "__main__":
    main()
