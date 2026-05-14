from pathlib import Path
import os
import time
import traceback
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.core.paginator import Paginator
from django.db.models import Avg, Count, Q, Sum
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.automacoes.models import TransmittalKM, PCFTimeline, DocumentoLD, ExecucaoAutomacao, KMFileIndex
from apps.automacoes.services import (
    atualizar_ld,
    grd_ghenova,
    timeline_pcfs,
    transmittal_km,
)
from apps.automacoes.services.ld_parser import extrair_tipo_documental
from apps.automacoes.services.ld_path_resolver import gerar_hyperlink_ld, resolver_caminho_ld
from apps.automacoes.services.status_normalizer import normalizar_status
from apps.automacoes.services.search_engine import buscar_global_enterprise
from apps.automacoes.services.search_analytics import obter_search_analytics
from apps.automacoes.services.km_index_jobs import executar_reindexacao_km_job
from apps.automacoes.services.ops_center_service import OperationsCenterService


KM_DOCUMENTOS_BASE = Path(
    r"\\virm-rgr022\FILESERVER\Projetos\05_HANDYMAX\09. Doc Control\15 - Documentos KM"
)

KM_EXTENSOES_PRIORITARIAS = {
    ".docx": 60,
    ".doc": 58,
    ".dwg": 54,
    ".xlsx": 48,
    ".xlsm": 46,
    ".xls": 44,
    ".pdf": 20,
}




def _model_has_field(model, nome):
    return any(field.name == nome for field in model._meta.get_fields())


def _latest_model_value(model, *campos):
    for campo in campos:
        if _model_has_field(model, campo):
            try:
                valor = (
                    model.objects.exclude(**{f"{campo}__isnull": True})
                    .order_by(f"-{campo}")
                    .values_list(campo, flat=True)
                    .first()
                )
                if valor:
                    return valor
            except Exception:
                continue
    return None


def _extrair_quantidade_processada(resultado):
    if not isinstance(resultado, dict):
        return 0

    chaves = (
        "quantidade_processada",
        "processados",
        "total_processado",
        "total",
        "importados",
        "criados",
        "atualizados",
    )

    for chave in chaves:
        valor = resultado.get(chave)
        if isinstance(valor, int):
            return valor

    detalhes = resultado.get("detalhes")
    if isinstance(detalhes, dict):
        for chave in chaves:
            valor = detalhes.get(chave)
            if isinstance(valor, int):
                return valor

    return 0


def _detalhes_execucao(resultado):
    if not isinstance(resultado, dict):
        return {}

    detalhes = resultado.get("detalhes")
    if isinstance(detalhes, dict):
        return detalhes

    return {
        chave: valor
        for chave, valor in resultado.items()
        if chave not in {"ok", "mensagem"}
        and isinstance(valor, (str, int, float, bool, list, dict, type(None)))
    }


def _formatar_duracao(segundos):
    segundos = float(segundos or 0)

    if segundos < 60:
        return f"{segundos:.1f}s" if segundos and segundos < 10 else f"{round(segundos)}s"

    minutos = int(segundos // 60)
    resto = int(segundos % 60)
    return f"{minutos}m {resto}s"


def _health_automacoes(nomes):
    health = {}

    for nome in nomes:
        qs = ExecucaoAutomacao.objects.filter(nome=nome)
        ultima = qs.order_by("-iniciado_em").first()
        total_finalizado = qs.exclude(status=ExecucaoAutomacao.STATUS_INICIADO).count()
        total_sucesso = qs.filter(status=ExecucaoAutomacao.STATUS_SUCESSO).count()
        total_erros = qs.filter(status=ExecucaoAutomacao.STATUS_ERRO).count()
        duracao_media = (
            qs.exclude(duracao_segundos=0)
            .aggregate(media=Avg("duracao_segundos"))
            .get("media")
            or 0
        )

        if not ultima:
            estado = "OCIOSO"
            classe = "auto-status-idle"
            icone = "bi-circle"
        elif ultima.status == ExecucaoAutomacao.STATUS_INICIADO:
            estado = "EXECUTANDO"
            classe = "auto-status-running"
            icone = "bi-arrow-repeat"
        elif ultima.status == ExecucaoAutomacao.STATUS_ERRO:
            estado = "ERRO"
            classe = "auto-status-error"
            icone = "bi-exclamation-triangle"
        else:
            estado = "ONLINE"
            classe = "auto-status-online"
            icone = "bi-check-circle"

        taxa_sucesso = round((total_sucesso / total_finalizado) * 100, 1) if total_finalizado else 0

        health[nome] = {
            "estado": estado,
            "classe": classe,
            "icone": icone,
            "ultima": ultima,
            "taxa_sucesso": taxa_sucesso,
            "total_erros": total_erros,
            "duracao_media": round(duracao_media, 2) if duracao_media else 0,
            "duracao_media_fmt": _formatar_duracao(duracao_media),
        }

    return health



@login_required
def painel(request):
    total_ld = DocumentoLD.objects.count()
    total_pcfs = PCFTimeline.objects.count()
    total_transmittals = TransmittalKM.objects.count()

    total_ld_com_pcf = 0
    if _model_has_field(DocumentoLD, "pcf"):
        total_ld_com_pcf = DocumentoLD.objects.exclude(pcf="").exclude(pcf__isnull=True).count()

    total_ld_sem_pcf = max(total_ld - total_ld_com_pcf, 0)

    total_pcfs_open = PCFTimeline.objects.filter(open_comments__gt=0).count()
    total_pcfs_not_released = PCFTimeline.objects.filter(status_final__iexact="NOT RELEASED").count()
    total_pcfs_released = PCFTimeline.objects.filter(status_final__iexact="RELEASED").count()

    total_transmittals_unicos = (
        TransmittalKM.objects.exclude(transmittal_numero="")
        .exclude(transmittal_numero__isnull=True)
        .values("transmittal_numero")
        .distinct()
        .count()
        if _model_has_field(TransmittalKM, "transmittal_numero")
        else 0
    )

    total_transmittals_sem_pdf = (
        TransmittalKM.objects.filter(Q(arquivo_pdf="") | Q(arquivo_pdf__isnull=True)).count()
        if _model_has_field(TransmittalKM, "arquivo_pdf")
        else 0
    )

    ultimos_pcfs = PCFTimeline.objects.order_by("-atualizado_em")[:5] if _model_has_field(PCFTimeline, "atualizado_em") else PCFTimeline.objects.all()[:5]
    ultimos_transmittals = TransmittalKM.objects.order_by("-criado_em")[:5] if _model_has_field(TransmittalKM, "criado_em") else TransmittalKM.objects.all()[:5]

    ultima_atualizacao = (
        _latest_model_value(PCFTimeline, "atualizado_em", "criado_em")
        or _latest_model_value(TransmittalKM, "criado_em", "atualizado_em")
        or _latest_model_value(DocumentoLD, "atualizado_em", "criado_em")
    )

    ultimas_execucoes = ExecucaoAutomacao.objects.select_related("usuario").order_by("-iniciado_em")[:8]
    ultima_execucao = ultimas_execucoes[0] if ultimas_execucoes else None
    ultima_falha = (
        ExecucaoAutomacao.objects.select_related("usuario")
        .filter(sucesso=False)
        .exclude(status=ExecucaoAutomacao.STATUS_INICIADO)
        .order_by("-iniciado_em")
        .first()
    )
    total_execucoes = ExecucaoAutomacao.objects.count()
    total_falhas = (
        ExecucaoAutomacao.objects.filter(sucesso=False)
        .exclude(status=ExecucaoAutomacao.STATUS_INICIADO)
        .count()
    )
    total_sucessos = ExecucaoAutomacao.objects.filter(status=ExecucaoAutomacao.STATUS_SUCESSO).count()
    total_finalizados = ExecucaoAutomacao.objects.exclude(status=ExecucaoAutomacao.STATUS_INICIADO).count()
    taxa_sucesso_global = round((total_sucessos / total_finalizados) * 100, 1) if total_finalizados else 0
    hoje = timezone.localdate()
    execucoes_hoje = ExecucaoAutomacao.objects.filter(iniciado_em__date=hoje).count()
    falhas_hoje = ExecucaoAutomacao.objects.filter(
        iniciado_em__date=hoje,
        status=ExecucaoAutomacao.STATUS_ERRO,
    ).count()
    duracao_media_global = (
        ExecucaoAutomacao.objects.exclude(duracao_segundos=0)
        .aggregate(media=Avg("duracao_segundos"))
        .get("media")
        or 0
    )
    duracao_media_global = round(duracao_media_global, 2) if duracao_media_global else 0

    total_km_index = KMFileIndex.objects.filter(ativo=True).count()
    total_km_docs_index = KMFileIndex.objects.filter(ativo=True, eh_transmittal_letter=False).count()
    total_km_transmittals_index = KMFileIndex.objects.filter(ativo=True, eh_transmittal_letter=True).count()
    ultima_indexacao_km = (
        KMFileIndex.objects.filter(ativo=True)
        .order_by("-indexado_em")
        .values_list("indexado_em", flat=True)
        .first()
    )

    automacoes = [
        {
            "nome": "Atualização LD",
            "subtitulo": "Lista de Documentos",
            "icone": "bi-file-earmark-spreadsheet",
            "badge": "Crítica",
            "badge_class": "auto-badge-warning",
            "descricao": "Sincroniza dados documentais, revisões, PCFs, GRDs, links de rede e medição.",
            "form_url": "automacoes:atualizar_ld",
            "botao": "Executar Atualização LD",
            "botao_class": "btn-primary",
            "dashboard_url": "automacoes:dashboard_ld",
            "registros_url": "automacoes:lista_ld",
            "metricas": [
                {"label": "Linhas LD", "valor": total_ld},
                {"label": "Com PCF", "valor": total_ld_com_pcf},
                {"label": "Sem PCF", "valor": total_ld_sem_pcf},
            ],
        },
        {
            "nome": "Timeline PCFs",
            "subtitulo": "Comentários e revisões",
            "icone": "bi-bar-chart-line",
            "badge": "Integrada",
            "badge_class": "auto-badge-success",
            "descricao": "Gera e atualiza a timeline das PCFs recebidas e respondidas.",
            "form_url": "automacoes:timeline_pcfs",
            "botao": "Gerar Timeline PCFs",
            "botao_class": "btn-success",
            "dashboard_url": "automacoes:dashboard_pcfs",
            "registros_url": "automacoes:pcfs_timeline",
            "metricas": [
                {"label": "PCFs", "valor": total_pcfs},
                {"label": "Open", "valor": total_pcfs_open},
                {"label": "Not Released", "valor": total_pcfs_not_released},
            ],
        },
        {
            "nome": "Transmittal KM",
            "subtitulo": "Parser PDF",
            "icone": "bi-box-seam",
            "badge": "Parser PDF",
            "badge_class": "auto-badge-info",
            "descricao": "Lê PDFs KM e consolida transmittals para acompanhamento documental.",
            "form_url": "automacoes:transmittal_km",
            "botao": "Consolidar Transmittals KM",
            "botao_class": "btn-info",
            "dashboard_url": "automacoes:dashboard_transmittals",
            "registros_url": "automacoes:transmittals_km",
            "metricas": [
                {"label": "Registros", "valor": total_transmittals},
                {"label": "Transmittals", "valor": total_transmittals_unicos},
                {"label": "Sem PDF", "valor": total_transmittals_sem_pdf},
            ],
        },
        {
            "nome": "Índice KM",
            "subtitulo": "Arquivos e documentos KM",
            "icone": "bi-hdd-network",
            "badge": "Indexação",
            "badge_class": "auto-badge-info",
            "descricao": "Varre a pasta Documentos KM, indexa arquivos técnicos e acelera a abertura direta dos documentos.",
            "form_url": "automacoes:indexar_km",
            "botao": "Atualizar Índice KM",
            "botao_class": "btn-primary",
            "dashboard_url": "automacoes:transmittals_km",
            "registros_url": "automacoes:transmittals_km",
            "metricas": [
                {"label": "Arquivos", "valor": total_km_index},
                {"label": "Docs técnicos", "valor": total_km_docs_index},
                {"label": "Letters", "valor": total_km_transmittals_index},
            ],
        },
        {
            "nome": "GRD GHENOVA",
            "subtitulo": "Consolidação GRDs 7K e 14K",
            "icone": "bi-diagram-3",
            "badge": "Engenharia",
            "badge_class": "auto-badge-neutral",
            "descricao": "Processa PDFs de GRD e gera planilhas consolidadas por empreendimento.",
            "form_url": "automacoes:grd_ghenova",
            "botao": "Consolidar GRDs GHENOVA",
            "botao_class": "btn-secondary",
            "dashboard_url": "",
            "registros_url": "",
            "metricas": [
                {"label": "Fonte", "valor": "PDF"},
                {"label": "Escopo", "valor": "7K/14K"},
                {"label": "Status", "valor": "Ativo"},
            ],
        },
    ]

    health_map = _health_automacoes([rotina["nome"] for rotina in automacoes])
    for rotina in automacoes:
        rotina["health"] = health_map.get(rotina["nome"], {})

    return render(
        request,
        "automacoes/painel.html",
        {
            "total_ld": total_ld,
            "total_pcfs": total_pcfs,
            "total_pcfs_open": total_pcfs_open,
            "total_pcfs_released": total_pcfs_released,
            "total_pcfs_not_released": total_pcfs_not_released,
            "total_transmittals": total_transmittals,
            "total_transmittals_unicos": total_transmittals_unicos,
            "ultima_atualizacao": ultima_atualizacao,
            "automacoes": automacoes,
            "ultimos_pcfs": ultimos_pcfs,
            "ultimos_transmittals": ultimos_transmittals,
            "ultimas_execucoes": ultimas_execucoes,
            "ultima_execucao": ultima_execucao,
            "ultima_falha": ultima_falha,
            "total_execucoes": total_execucoes,
            "total_falhas": total_falhas,
            "total_sucessos": total_sucessos,
            "taxa_sucesso_global": taxa_sucesso_global,
            "execucoes_hoje": execucoes_hoje,
            "falhas_hoje": falhas_hoje,
            "duracao_media_global": duracao_media_global,
            "total_km_index": total_km_index,
            "total_km_docs_index": total_km_docs_index,
            "total_km_transmittals_index": total_km_transmittals_index,
            "ultima_indexacao_km": ultima_indexacao_km,
        },
    )


def _executar_automacao(request, executor, nome):
    if request.method != "POST":
        messages.error(request, f"Método inválido para executar {nome}.")
        return redirect("automacoes:painel")

    inicio = time.monotonic()
    log = ExecucaoAutomacao.objects.create(
        nome=nome,
        usuario=request.user if request.user.is_authenticated else None,
        status=ExecucaoAutomacao.STATUS_INICIADO,
        mensagem="Execução iniciada.",
    )

    try:
        resultado = executor()
        ok = bool(resultado.get("ok")) if isinstance(resultado, dict) else False
        mensagem = (
            resultado.get("mensagem")
            if isinstance(resultado, dict)
            else f"{nome} executado."
        )

        log.status = ExecucaoAutomacao.STATUS_SUCESSO if ok else ExecucaoAutomacao.STATUS_ERRO
        log.sucesso = ok
        log.mensagem = mensagem or (
            f"{nome} executado com sucesso." if ok else f"Falha ao executar {nome}."
        )
        log.quantidade_processada = _extrair_quantidade_processada(resultado)
        log.detalhes = _detalhes_execucao(resultado)

        if ok:
            messages.success(request, log.mensagem)
        else:
            messages.error(request, log.mensagem)

    except Exception as exc:
        log.status = ExecucaoAutomacao.STATUS_ERRO
        log.sucesso = False
        log.mensagem = f"Erro ao executar {nome}: {exc}"
        log.detalhes = {"erro": str(exc)}
        messages.error(request, log.mensagem)

    finally:
        log.finalizado_em = timezone.now()
        log.duracao_segundos = round(time.monotonic() - inicio, 3)
        log.save(
            update_fields=[
                "status",
                "sucesso",
                "mensagem",
                "detalhes",
                "quantidade_processada",
                "duracao_segundos",
                "finalizado_em",
            ]
        )

    return redirect("automacoes:painel")


@login_required
def logs_automacoes(request):
    busca = request.GET.get("q", "").strip()
    tipo_doc = request.GET.get("tipo_doc", "").strip().upper()

    tipos_documentais = [
        "AC", "AF", "AP", "AR", "AV", "BM", "BS", "CA", "CC", "CE", "CF", "CG",
        "CI", "CL", "CM", "CO", "CP", "CQ", "CR", "CT", "CV", "DB", "DC", "DE",
        "DF", "DI", "DL", "DO", "DR", "DT", "DU", "EE", "EC", "EM", "ES", "ET",
        "EQ", "FD", "GE", "GI", "ID", "IT", "IS", "LA", "LC", "LD", "LE", "LI",
        "LM", "LP", "LO", "LV", "LT", "MA", "MC", "MD", "MG", "MI", "ML", "MM",
        "MO", "NA", "NC", "NF", "NP", "NQ", "NT", "OA", "OC", "OG", "OS", "PC",
        "PE", "PG", "PI", "PJ", "PL", "PM", "PO", "PP", "PQ", "PR", "PT", "QT",
        "RA", "RC", "RD", "RE", "RH", "RL", "RM", "RV", "SC", "SM", "SP", "TF",
        "TI", "TP", "TR",
    ]

    status = request.GET.get("status", "").strip()
    nome = request.GET.get("nome", "").strip()

    logs = ExecucaoAutomacao.objects.select_related("usuario").order_by("-iniciado_em")

    if busca:
        logs = logs.filter(
            Q(nome__icontains=busca)
            | Q(mensagem__icontains=busca)
            | Q(usuario__username__icontains=busca)
            | Q(usuario__first_name__icontains=busca)
            | Q(usuario__last_name__icontains=busca)
        )

    if status:
        logs = logs.filter(status=status)

    if nome:
        logs = logs.filter(nome__iexact=nome)

    nomes = (
        ExecucaoAutomacao.objects.exclude(nome="")
        .values_list("nome", flat=True)
        .distinct()
        .order_by("nome")
    )

    total = logs.count()
    total_sucesso = logs.filter(sucesso=True).count()
    total_erros = logs.filter(status=ExecucaoAutomacao.STATUS_ERRO).count()
    total_iniciados = logs.filter(status=ExecucaoAutomacao.STATUS_INICIADO).count()

    duracao_media = logs.exclude(duracao_segundos=0).aggregate(
        media=Avg("duracao_segundos")
    ).get("media")
    duracao_media = round(duracao_media, 2) if duracao_media else 0

    total_finalizados = logs.exclude(status=ExecucaoAutomacao.STATUS_INICIADO).count()
    taxa_sucesso = round((total_sucesso / total_finalizados) * 100, 1) if total_finalizados else 0
    hoje = timezone.localdate()
    falhas_hoje = logs.filter(iniciado_em__date=hoje, status=ExecucaoAutomacao.STATUS_ERRO).count()
    execucoes_hoje = logs.filter(iniciado_em__date=hoje).count()

    por_automacao = list(
        logs.values("nome")
        .annotate(total=Count("id"))
        .order_by("-total", "nome")[:10]
    )
    automacao_chart_labels = [(item.get("nome") or "Sem nome") for item in por_automacao]
    automacao_chart_values = [item.get("total") or 0 for item in por_automacao]

    erros_por_automacao = list(
        logs.filter(status=ExecucaoAutomacao.STATUS_ERRO)
        .values("nome")
        .annotate(total=Count("id"))
        .order_by("-total", "nome")[:10]
    )
    erros_chart_labels = [(item.get("nome") or "Sem nome") for item in erros_por_automacao]
    erros_chart_values = [item.get("total") or 0 for item in erros_por_automacao]

    paginator = Paginator(logs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "automacoes/logs_automacoes.html",
        {
            "logs": page_obj,
            "page_obj": page_obj,
            "busca": busca,
            "status": status,
            "nome": nome,
            "nomes": nomes,
            "total": total,
            "total_sucesso": total_sucesso,
            "total_erros": total_erros,
            "total_iniciados": total_iniciados,
            "duracao_media": duracao_media,
            "taxa_sucesso": taxa_sucesso,
            "falhas_hoje": falhas_hoje,
            "execucoes_hoje": execucoes_hoje,
            "automacao_chart_labels": automacao_chart_labels,
            "automacao_chart_values": automacao_chart_values,
            "erros_chart_labels": erros_chart_labels,
            "erros_chart_values": erros_chart_values,
            "status_choices": ExecucaoAutomacao.STATUS_CHOICES,
        },
    )


@login_required
def executar_atualizar_ld(request):
    return _executar_automacao(
        request,
        atualizar_ld.executar,
        "Atualização LD",
    )


@login_required
def timeline_pcfs_view(request):
    return _executar_automacao(
        request,
        timeline_pcfs.executar,
        "Timeline PCFs",
    )


@login_required
def executar_transmittal_km(request):
    return _executar_automacao(
        request,
        transmittal_km.executar,
        "Transmittal KM",
    )


@login_required
def executar_grd_ghenova(request):
    return _executar_automacao(
        request,
        grd_ghenova.executar,
        "GRD GHENOVA",
    )


def _km_documento_extraido_do_nome(path):
    """
    Extrai um identificador documental provável do nome do arquivo KM.
    Mantém formato legível quando possível; a busca usa também campos normalizados.
    """
    stem = Path(path).stem
    stem = re.sub(r"[_]+", "-", stem)
    m = re.search(r"(\d{2}-\d{4}-\d{2}-\d{3,4}-\d{2,4}-\d{2}(?:-[A-Z0-9]+)?)", stem, re.IGNORECASE)
    if m:
        return m.group(1).upper()

    m = re.search(r"(\d{3,4}-\d{2,4}-\d{2}(?:-[A-Z0-9]+)?)", stem, re.IGNORECASE)
    if m:
        return m.group(1).upper()

    return ""


def _km_indexar_banco():
    """
    Varre a árvore KM e grava um índice persistente no banco.
    Isso evita varredura de rede a cada abertura de documento.
    """
    inicio = time.monotonic()

    if not KM_DOCUMENTOS_BASE.exists():
        return {
            "ok": False,
            "mensagem": f"Pasta KM não encontrada: {KM_DOCUMENTOS_BASE}",
            "quantidade_processada": 0,
            "detalhes": {"base": str(KM_DOCUMENTOS_BASE)},
        }

    KMFileIndex.objects.update(ativo=False)

    total = 0
    criados = 0
    atualizados = 0
    erros = 0
    por_extensao = {}

    for arquivo in KM_DOCUMENTOS_BASE.rglob("*"):
        try:
            if not arquivo.is_file():
                continue

            stat = arquivo.stat()
            extensao = arquivo.suffix.lower()
            por_extensao[extensao or "sem_extensao"] = por_extensao.get(extensao or "sem_extensao", 0) + 1

            defaults = {
                "nome_arquivo": arquivo.name,
                "pasta": str(arquivo.parent),
                "extensao": extensao,
                "tamanho_bytes": int(stat.st_size or 0),
                "modificado_em": timezone.datetime.fromtimestamp(
                    stat.st_mtime,
                    tz=timezone.get_current_timezone(),
                ),
                "nome_normalizado": _km_normalizar(arquivo.name),
                "stem_normalizado": _km_normalizar(arquivo.stem),
                "documento_extraido": _km_documento_extraido_do_nome(arquivo),
                "eh_transmittal_letter": _km_eh_transmittal_letter(arquivo),
                "ativo": True,
            }

            _, created = KMFileIndex.objects.update_or_create(
                caminho_completo=str(arquivo),
                defaults=defaults,
            )

            total += 1
            if created:
                criados += 1
            else:
                atualizados += 1

        except Exception:
            erros += 1
            continue

    _km_limpar_cache()

    removidos = KMFileIndex.objects.filter(ativo=False).count()
    status = "sucesso" if erros == 0 else "sucesso_parcial"

    return {
        "ok": True,
        "status": status,
        "mensagem": (
            f"Índice KM atualizado: {total} arquivos ativos, "
            f"{criados} novos, {atualizados} atualizados, {removidos} inativos."
        ),
        "quantidade_processada": total,
        "detalhes": {
            "base": str(KM_DOCUMENTOS_BASE),
            "arquivos_ativos": total,
            "criados": criados,
            "atualizados": atualizados,
            "inativos": removidos,
            "erros": erros,
            "por_extensao": por_extensao,
            "duracao_segundos": round(time.monotonic() - inicio, 3),
        },
    }


@login_required
def executar_indice_km(request):
    return _executar_automacao(
        request,
        executar_reindexacao_km_job,
        "Índice KM",
    )



_KM_INDEX_CACHE = None


def _km_limpar_cache():
    global _KM_INDEX_CACHE
    _KM_INDEX_CACHE = None


def _km_normalizar(valor):
    return "".join(ch for ch in str(valor or "").upper() if ch.isalnum())


def _km_eh_transmittal_letter(path):
    texto = str(path).replace("/", "\\").lower()
    nome = path.name.lower()
    return (
        "\\0 transmittal letters\\transmittal letters\\" in texto
        or "transmittal letters\\transmittal letters\\" in texto
        or nome.startswith("t-")
        or "transmittal" in nome
    )


def _km_indexar_documentos():
    """
    Indexa arquivos da árvore KM uma vez por processo do Django.

    Importante:
    - Mantém os Transmittal Letters no índice apenas como fallback.
    - Prioriza documentos técnicos em subpastas como 1.4 ETS.
    """
    global _KM_INDEX_CACHE

    if _KM_INDEX_CACHE is not None:
        return _KM_INDEX_CACHE

    itens = []

    if not KM_DOCUMENTOS_BASE.exists():
        _KM_INDEX_CACHE = itens
        return itens

    try:
        for arquivo in KM_DOCUMENTOS_BASE.rglob("*"):
            try:
                if not arquivo.is_file():
                    continue
            except OSError:
                continue

            nome_norm = _km_normalizar(arquivo.name)
            stem_norm = _km_normalizar(arquivo.stem)
            suffix = arquivo.suffix.lower()

            if not nome_norm:
                continue

            itens.append({
                "path": arquivo,
                "nome_norm": nome_norm,
                "stem_norm": stem_norm,
                "suffix": suffix,
                "is_transmittal_letter": _km_eh_transmittal_letter(arquivo),
            })
    except Exception:
        itens = []

    _KM_INDEX_CACHE = itens
    return itens


def _km_score_documento(documento, item):
    doc_norm = _km_normalizar(documento)

    if not doc_norm:
        return 0

    nome_norm = item["nome_norm"]
    stem_norm = item["stem_norm"]
    path = item["path"]
    suffix = item["suffix"]

    score = 0

    if stem_norm == doc_norm:
        score = 100
    elif nome_norm == doc_norm:
        score = 98
    elif doc_norm in stem_norm:
        score = 88
    elif doc_norm in nome_norm:
        score = 84
    elif stem_norm in doc_norm and len(stem_norm) >= 8:
        score = 60

    if not score:
        return 0

    # Prioriza documentos reais de engenharia sobre PDF do transmittal.
    score += KM_EXTENSOES_PRIORITARIAS.get(suffix, 5)

    if item["is_transmittal_letter"]:
        score -= 120
    else:
        score += 35

    # Subpastas técnicas costumam ter arquivos reais; raiz de letters tende a ser só carta.
    texto_path = str(path).replace("/", "\\").lower()
    if "\\0 transmittal letters\\" in texto_path and "\\transmittal letters\\" not in texto_path:
        score += 8

    # Nome típico do PDF de carta começa com T-; nunca deve vencer um docx/dwg real.
    if path.name.upper().startswith("T-"):
        score -= 80

    return score


def _km_score_documento_indexado(documento, item):
    doc_norm = _km_normalizar(documento)

    if not doc_norm:
        return 0

    nome_norm = item.nome_normalizado or _km_normalizar(item.nome_arquivo)
    stem_norm = item.stem_normalizado or _km_normalizar(Path(item.nome_arquivo).stem)
    suffix = (item.extensao or "").lower()

    score = 0

    if stem_norm == doc_norm:
        score = 100
    elif nome_norm == doc_norm:
        score = 98
    elif doc_norm in stem_norm:
        score = 88
    elif doc_norm in nome_norm:
        score = 84
    elif stem_norm in doc_norm and len(stem_norm) >= 8:
        score = 60

    documento_extraido_norm = _km_normalizar(item.documento_extraido)
    if documento_extraido_norm:
        if documento_extraido_norm == doc_norm:
            score = max(score, 105)
        elif doc_norm in documento_extraido_norm or documento_extraido_norm in doc_norm:
            score = max(score, 86)

    if not score:
        return 0

    score += KM_EXTENSOES_PRIORITARIAS.get(suffix, 5)

    if item.eh_transmittal_letter:
        score -= 120
    else:
        score += 35

    caminho_lower = (item.caminho_completo or "").replace("/", "\\").lower()
    if "\\0 transmittal letters\\" in caminho_lower and "\\transmittal letters\\" not in caminho_lower:
        score += 8

    if item.nome_arquivo.upper().startswith("T-"):
        score -= 80

    return score


def _km_buscar_documento_indexado(documento, permitir_transmittal=False):
    doc_norm = _km_normalizar(documento)

    if not doc_norm:
        return None

    qs = KMFileIndex.objects.filter(ativo=True)

    # Primeiro tenta reduzir no banco; se o código vier abreviado, ainda há fallback abaixo.
    candidatos_qs = qs.filter(
        Q(nome_normalizado__icontains=doc_norm)
        | Q(stem_normalizado__icontains=doc_norm)
        | Q(documento_extraido__icontains=str(documento or "").strip())
    )[:500]

    candidatos = []

    for item in candidatos_qs:
        score = _km_score_documento_indexado(documento, item)
        if score <= 0:
            continue

        if item.eh_transmittal_letter and not permitir_transmittal:
            continue

        candidatos.append((score, Path(item.caminho_completo)))

    if not candidatos:
        # Fallback amplo, ainda baseado no banco, para códigos extraídos/formatos inesperados.
        for item in qs.order_by("-indexado_em")[:20000]:
            score = _km_score_documento_indexado(documento, item)
            if score <= 0:
                continue

            if item.eh_transmittal_letter and not permitir_transmittal:
                continue

            candidatos.append((score, Path(item.caminho_completo)))

    if not candidatos and permitir_transmittal:
        for item in qs.filter(eh_transmittal_letter=True).order_by("-indexado_em")[:5000]:
            score = _km_score_documento_indexado(documento, item)
            if score > 0:
                candidatos.append((score, Path(item.caminho_completo)))

    if not candidatos:
        return None

    candidatos.sort(key=lambda par: (par[0], -len(str(par[1]))), reverse=True)
    return candidatos[0][1]


def _km_buscar_documento(documento, permitir_transmittal=False):
    """
    Busca o documento real no índice KM persistido.
    Se o índice ainda não existir, usa a varredura antiga como fallback.
    """
    arquivo = _km_buscar_documento_indexado(documento, permitir_transmittal=permitir_transmittal)
    if arquivo:
        return arquivo

    candidatos = []

    for item in _km_indexar_documentos():
        score = _km_score_documento(documento, item)

        if score <= 0:
            continue

        if item["is_transmittal_letter"] and not permitir_transmittal:
            continue

        candidatos.append((score, item["path"]))

    if not candidatos and permitir_transmittal:
        for item in _km_indexar_documentos():
            score = _km_score_documento(documento, item)
            if score > 0:
                candidatos.append((score, item["path"]))

    if not candidatos:
        return None

    candidatos.sort(key=lambda par: (par[0], -len(str(par[1]))), reverse=True)
    return candidatos[0][1]
def _km_buscar_documento_com_debug(documento):
    candidatos = []

    for item in _km_indexar_documentos():
        score = _km_score_documento(documento, item)
        if score > 0:
            candidatos.append((score, item["path"], item["is_transmittal_letter"]))

    candidatos.sort(key=lambda par: (par[0], -len(str(par[1]))), reverse=True)
    return candidatos



def _tr_texto(valor):
    return str(valor or "").strip()


def _tr_documento_normalizado(valor):
    """
    Normalização leve para exibição/busca textual.
    Mantém hífens porque muitos documentos KM/LD usam hífen como parte do código.
    """
    texto = _tr_texto(valor).upper()
    texto = texto.replace("\\", "/").split("/")[-1]
    texto = texto.split(".", 1)[0]
    texto = texto.replace("_", "-")
    texto = " ".join(texto.split())
    return texto.strip()


def _tr_documento_compacto(valor):
    """
    Normalização forte para comparação entre códigos vindos de fontes diferentes.

    Exemplos:
    - 108-505-02
    - 108_505_02
    - 108 505 02
    - DOC-108-505-02-REV0

    viram uma string comparável sem símbolos.
    """
    texto = _tr_documento_normalizado(valor)
    return "".join(ch for ch in texto if ch.isalnum())


def _tr_tokens_documento(valor):
    texto = _tr_documento_normalizado(valor)
    tokens = [t for t in re.split(r"[^A-Z0-9]+", texto) if t]
    return tokens


def _tr_score_match_documento(doc_busca, item_ld):
    """
    Pontua possíveis correspondências entre documento KM e registro LD.
    Evita depender apenas de iexact/icontains, porque os códigos KM podem vir
    resumidos e a LD pode conter prefixos/sufixos/revisões.
    """
    busca_norm = _tr_documento_normalizado(doc_busca)
    busca_compacta = _tr_documento_compacto(doc_busca)
    tokens = _tr_tokens_documento(doc_busca)

    campos = [
        getattr(item_ld, "documento", ""),
        getattr(item_ld, "titulo", ""),
        getattr(item_ld, "caminho_documento", ""),
        getattr(item_ld, "caminho_grd", ""),
        getattr(item_ld, "caminho_pcf", ""),
        getattr(item_ld, "caminho_resposta", ""),
        getattr(item_ld, "caminho_grd_resposta", ""),
    ]

    melhor = 0

    for valor in campos:
        texto = _tr_texto(valor)
        if not texto:
            continue

        texto_norm = _tr_documento_normalizado(texto)
        texto_compacto = _tr_documento_compacto(texto)

        if texto_norm == busca_norm:
            melhor = max(melhor, 100)

        if busca_norm and busca_norm in texto_norm:
            melhor = max(melhor, 85)

        if busca_compacta and busca_compacta == texto_compacto:
            melhor = max(melhor, 95)

        if busca_compacta and busca_compacta in texto_compacto:
            melhor = max(melhor, 80)

        if texto_compacto and texto_compacto in busca_compacta:
            melhor = max(melhor, 70)

        if tokens and all(token in texto_norm for token in tokens):
            melhor = max(melhor, 65)

        if tokens and all(token in texto_compacto for token in tokens):
            melhor = max(melhor, 60)

    return melhor


def _tr_buscar_ld_por_documento(numero_documento):
    """
    Localiza o documento KM dentro da Lista LD para permitir abertura direta
    do arquivo/pasta real na rede.

    A busca é propositalmente tolerante:
    - tenta match exato;
    - tenta contains no campo documento;
    - tenta comparar código compacto sem símbolos;
    - tenta procurar também em título e caminhos da LD.
    """
    doc = _tr_texto(numero_documento)

    if not doc:
        return None

    # 1) caminho rápido: exato no campo documento
    candidatos = DocumentoLD.objects.filter(documento__iexact=doc).order_by("-id")
    if candidatos.exists():
        return candidatos.first()

    doc_norm = _tr_documento_normalizado(doc)
    doc_compacto = _tr_documento_compacto(doc)

    if not doc_norm and not doc_compacto:
        return None

    # 2) contains tradicional
    candidatos = DocumentoLD.objects.filter(documento__icontains=doc_norm).order_by("-id")
    if candidatos.exists():
        return candidatos.first()

    # 3) busca em campos textuais/caminhos, útil quando a LD tem prefixos/sufixos
    busca_q = (
        Q(documento__icontains=doc_norm)
        | Q(titulo__icontains=doc_norm)
        | Q(caminho_documento__icontains=doc_norm)
        | Q(caminho_grd__icontains=doc_norm)
        | Q(caminho_pcf__icontains=doc_norm)
        | Q(caminho_resposta__icontains=doc_norm)
        | Q(caminho_grd_resposta__icontains=doc_norm)
    )
    candidatos_textuais = list(DocumentoLD.objects.filter(busca_q).order_by("-id")[:200])
    if candidatos_textuais:
        candidatos_textuais.sort(
            key=lambda item: _tr_score_match_documento(doc, item),
            reverse=True,
        )
        if _tr_score_match_documento(doc, candidatos_textuais[0]) >= 50:
            return candidatos_textuais[0]

    # 4) fallback amplo com pontuação. Limite para não pesar demais.
    melhor_item = None
    melhor_score = 0

    for item in DocumentoLD.objects.exclude(documento="").order_by("-id")[:10000]:
        score = _tr_score_match_documento(doc, item)
        if score > melhor_score:
            melhor_score = score
            melhor_item = item

        if melhor_score >= 95:
            break

    if melhor_item and melhor_score >= 60:
        return melhor_item

    return None


def _tr_caminho_documento_ld(item_ld):
    if not item_ld:
        return ""

    for campo in [
        "caminho_documento",
        "caminho_grd",
        "caminho_pcf",
        "caminho_resposta",
        "caminho_grd_resposta",
    ]:
        valor = getattr(item_ld, campo, "")
        if _tr_texto(valor):
            return valor

    return ""


def _tr_montar_central_transmittals(registros):
    grupos = {}

    for item in registros:
        numero = _tr_texto(item.transmittal_numero) or "Sem número"

        if numero not in grupos:
            grupos[numero] = {
                "numero": numero,
                "pdf_id": None,
                "pdf_path": "",
                "data_envio": "",
                "emissao": "",
                "proposito": "",
                "pastas": set(),
                "status": set(),
                "docs": [],
                "total_docs": 0,
            }

        grupo = grupos[numero]

        if item.arquivo_pdf and not grupo["pdf_id"]:
            grupo["pdf_id"] = item.id
            grupo["pdf_path"] = item.arquivo_pdf

        if item.data_envio and not grupo["data_envio"]:
            grupo["data_envio"] = item.data_envio

        if item.emissao and not grupo["emissao"]:
            grupo["emissao"] = item.emissao

        if item.proposito_emissao and not grupo["proposito"]:
            grupo["proposito"] = item.proposito_emissao

        if item.pasta:
            grupo["pastas"].add(item.pasta)

        if item.status_parse:
            grupo["status"].add(item.status_parse)

        item.km_arquivo = _km_buscar_documento(item.documento, permitir_transmittal=False)
        item.ld_vinculado = None

        grupo["docs"].append(item)

    transmittals = []

    for numero, grupo in grupos.items():
        grupo["docs"] = sorted(
            grupo["docs"],
            key=lambda doc: (
                _tr_texto(doc.pasta).lower(),
                _tr_texto(doc.documento).lower(),
                _tr_texto(doc.titulo).lower(),
            ),
        )
        grupo["total_docs"] = len(grupo["docs"])
        grupo["pastas_lista"] = sorted(grupo["pastas"])
        grupo["status_lista"] = sorted(grupo["status"])

        if any(str(s).upper() == "FALHA" for s in grupo["status_lista"]):
            grupo["status_badge"] = "FALHA"
            grupo["status_class"] = "bg-danger"
        elif any(str(s).upper() == "PARCIAL" for s in grupo["status_lista"]):
            grupo["status_badge"] = "PARCIAL"
            grupo["status_class"] = "bg-warning text-dark"
        elif grupo["status_lista"]:
            grupo["status_badge"] = "OK"
            grupo["status_class"] = "bg-success"
        else:
            grupo["status_badge"] = "—"
            grupo["status_class"] = "bg-secondary"

        transmittals.append(grupo)

    return sorted(
        transmittals,
        key=lambda grupo: (
            0 if grupo["numero"] != "Sem número" else 1,
            grupo["numero"].lower(),
        ),
    )


@login_required
def listar_transmittals_km(request):
    busca = request.GET.get("q", "").strip()
    pasta = request.GET.get("pasta", "").strip()
    emissao = request.GET.get("emissao", "").strip()
    transmittal = request.GET.get("transmittal", "").strip()

    registros = TransmittalKM.objects.all().order_by(
        "transmittal_numero",
        "pasta",
        "documento",
    )

    if busca:
        registros = registros.filter(
            Q(documento__icontains=busca)
            | Q(titulo__icontains=busca)
            | Q(transmittal_numero__icontains=busca)
            | Q(pasta__icontains=busca)
            | Q(emissao__icontains=busca)
            | Q(proposito_emissao__icontains=busca)
        )

    if pasta:
        registros = registros.filter(pasta__iexact=pasta)

    if emissao:
        registros = registros.filter(emissao__iexact=emissao)

    if transmittal:
        registros = registros.filter(transmittal_numero__iexact=transmittal)

    registros_lista = list(registros)
    transmittals_agrupados = _tr_montar_central_transmittals(registros_lista)

    total_documentos = len(registros_lista)
    total_transmittals = len(transmittals_agrupados)
    total_com_pdf = sum(1 for grupo in transmittals_agrupados if grupo.get("pdf_id"))
    total_sem_pdf = max(total_transmittals - total_com_pdf, 0)

    return render(
        request,
        "automacoes/transmittals_km.html",
        {
            "registros": registros_lista,
            "transmittals": transmittals_agrupados,
            "busca": busca,
            "pasta": pasta,
            "emissao": emissao,
            "transmittal": transmittal,
            "total_documentos": total_documentos,
            "total_transmittals": total_transmittals,
            "total_com_pdf": total_com_pdf,
            "total_sem_pdf": total_sem_pdf,
        },
    )


@login_required
def abrir_pdf_transmittal_km(request, pk):
    registro = TransmittalKM.objects.get(pk=pk)

    caminho_pdf = registro.arquivo_pdf

    if not caminho_pdf:
        raise Http404("PDF não localizado.")

    arquivo = Path(caminho_pdf)

    if not arquivo.exists():
        raise Http404(
            f"Arquivo não encontrado: {arquivo}"
        )

    return FileResponse(
        open(arquivo, "rb"),
        content_type="application/pdf",
    )


@login_required
def abrir_documento_transmittal_km(request, pk):
    registro = TransmittalKM.objects.get(pk=pk)

    arquivo = _km_buscar_documento(registro.documento, permitir_transmittal=False)

    # Fallback controlado: só abre o PDF do transmittal se nenhum documento técnico for localizado.
    if not arquivo and registro.arquivo_pdf:
        pdf = Path(registro.arquivo_pdf)
        if pdf.exists():
            arquivo = pdf

    if not arquivo:
        candidatos = _km_buscar_documento_com_debug(registro.documento)
        caminhos_testados = "\n".join(
            f"{score} | {'TRANS' if is_trans else 'DOC'} | {path}"
            for score, path, is_trans in candidatos[:20]
        )
        raise Http404(
            f"Documento KM não encontrado: {registro.documento}\n\n"
            f"Base KM: {KM_DOCUMENTOS_BASE}\n\n"
            f"Candidatos:\n{caminhos_testados}"
        )

    if os.name == "nt":
        os.startfile(str(arquivo))
        return HttpResponse(
            f"Arquivo aberto: {arquivo}",
            content_type="text/plain; charset=utf-8",
        )

    return FileResponse(
        open(arquivo, "rb"),
        as_attachment=False,
        filename=arquivo.name,
    )


@login_required
def abrir_pasta_documento_transmittal_km(request, pk):
    registro = TransmittalKM.objects.get(pk=pk)

    arquivo = _km_buscar_documento(registro.documento, permitir_transmittal=False)

    if not arquivo:
        candidatos = _km_buscar_documento_com_debug(registro.documento)
        caminhos_testados = "\n".join(
            f"{score} | {'TRANS' if is_trans else 'DOC'} | {path}"
            for score, path, is_trans in candidatos[:20]
        )
        raise Http404(
            f"Pasta KM não encontrada para: {registro.documento}\n\n"
            f"Base KM: {KM_DOCUMENTOS_BASE}\n\n"
            f"Candidatos:\n{caminhos_testados}"
        )

    pasta = arquivo.parent

    if os.name == "nt":
        os.startfile(str(pasta))
        return HttpResponse(
            f"Pasta aberta: {pasta}",
            content_type="text/plain; charset=utf-8",
        )

    return HttpResponse(
        f"Pasta localizada: {pasta}",
        content_type="text/plain; charset=utf-8",
    )


def _filtrar_pcfs_timeline(request):
    busca = request.GET.get("q", "").strip()
    tipo = request.GET.get("tipo", "").strip()
    status = request.GET.get("status", "").strip()
    somente_open = request.GET.get("somente_open", "").strip()

    registros = PCFTimeline.objects.all().order_by(
        "numero_documento",
        "revisao_pcf",
        "pcf_link",
    )

    if busca:
        registros = (
            registros.filter(numero_documento__icontains=busca)
            | registros.filter(numero_pcf__icontains=busca)
            | registros.filter(pcf_link__icontains=busca)
            | registros.filter(titulo__icontains=busca)
        )

    if tipo:
        registros = registros.filter(tipo=tipo)

    if status:
        registros = registros.filter(status_final__iexact=status)

    if somente_open:
        registros = registros.filter(open_comments__gt=0)

    return registros

@login_required
def listar_pcfs_timeline(request):
    busca = request.GET.get("q", "").strip()
    tipo = request.GET.get("tipo", "").strip()
    status = request.GET.get("status", "").strip()
    somente_open = request.GET.get("somente_open", "").strip()

    registros = _filtrar_pcfs_timeline(request)

    tipos = (
        PCFTimeline.objects.exclude(tipo="")
        .values_list("tipo", flat=True)
        .distinct()
        .order_by("tipo")
    )

    total = registros.count()
    total_open = registros.filter(open_comments__gt=0).count()
    total_sem_status = registros.filter(status_final="").count()

    paginator = Paginator(registros, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "automacoes/pcfs_timeline.html",
        {
            "registros": page_obj,
            "page_obj": page_obj,
            "busca": busca,
            "tipo": tipo,
            "status": status,
            "somente_open": somente_open,
            "tipos": tipos,
            "total": total,
            "total_open": total_open,
            "total_sem_status": total_sem_status,
        },
    )

@login_required
def abrir_arquivo_pcf(request, pk):
    registro = PCFTimeline.objects.get(pk=pk)

    caminho = registro.caminho

    if not caminho:
        raise Http404("Arquivo PCF não localizado.")

    arquivo = Path(caminho)

    if not arquivo.exists():
        raise Http404(
            f"Arquivo não encontrado: {arquivo}"
        )

    return FileResponse(
        open(arquivo, "rb"),
        as_attachment=False,
        filename=arquivo.name,
    )


@login_required
def exportar_pcfs_timeline_excel(request):
    registros = _filtrar_pcfs_timeline(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Timeline PCFs"

    headers = [
        "Tipo",
        "PCF Link",
        "Nº PCF",
        "Nº Documento",
        "Título",
        "Revisão",
        "Data Recebimento",
        "Open Comments",
        "Qtd Comentários",
        "Status Final",
        "Caminho",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for item in registros:
        ws.append([
            item.tipo,
            item.pcf_link,
            item.numero_pcf,
            item.numero_documento,
            item.titulo,
            item.revisao_pcf,
            item.data_recebimento,
            item.open_comments,
            item.qtd_comentarios,
            item.status_final,
            item.caminho,
        ])

    larguras = {
        "A": 25,
        "B": 35,
        "C": 30,
        "D": 30,
        "E": 60,
        "F": 12,
        "G": 18,
        "H": 16,
        "I": 18,
        "J": 25,
        "K": 90,
    }

    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="timeline_pcfs_filtrada.xlsx"'
    wb.save(response)
    return response



@login_required
def dashboard_pcfs(request):
    registros = PCFTimeline.objects.all()

    total = registros.count()
    total_open = registros.filter(open_comments__gt=0).count()
    total_sem_status = registros.filter(Q(status_final__isnull=True) | Q(status_final="")).count()
    total_not_released = registros.filter(status_final__icontains="NOT RELEASED").count()
    total_released = registros.filter(status_final__icontains="RELEASED").exclude(
        status_final__icontains="NOT RELEASED"
    ).count()

    total_comentarios_abertos = (
        registros.aggregate(total=Sum("open_comments")).get("total") or 0
    )

    por_tipo = list(
        registros.values("tipo")
        .annotate(total=Count("id"), open_total=Sum("open_comments"))
        .order_by("-total", "tipo")
    )

    status_agregado = {}
    for status, total_status in registros.values_list("status_final").annotate(total=Count("id")):
        status_norm = normalizar_status(status)
        status_agregado[status_norm] = status_agregado.get(status_norm, 0) + (total_status or 0)

    por_status = [
        {"status_final": status, "total": total_status}
        for status, total_status in sorted(
            status_agregado.items(),
            key=lambda item: (-item[1], item[0]),
        )[:10]
    ]

    top_pendencias = (
        registros.filter(open_comments__gt=0)
        .order_by("-open_comments", "numero_documento")[:15]
    )

    recentes = registros.order_by("-atualizado_em")[:10]

    status_chart_labels = [item.get("status_final") or "SEM STATUS" for item in por_status]
    status_chart_values = [item.get("total") or 0 for item in por_status]

    tipo_chart_labels = [(item.get("tipo") or "Sem tipo") for item in por_tipo]
    tipo_chart_values = [item.get("total") or 0 for item in por_tipo]

    tipo_open_labels = [(item.get("tipo") or "Sem tipo") for item in por_tipo]
    tipo_open_values = [item.get("open_total") or 0 for item in por_tipo]

    critical_rate = round((total_not_released / total) * 100, 1) if total else 0

    return render(
        request,
        "automacoes/dashboard_pcfs.html",
        {
            "total": total,
            "total_open": total_open,
            "total_sem_status": total_sem_status,
            "total_not_released": total_not_released,
            "total_released": total_released,
            "total_comentarios_abertos": total_comentarios_abertos,
            "critical_rate": critical_rate,
            "por_tipo": por_tipo,
            "por_status": por_status,
            "top_pendencias": top_pendencias,
            "recentes": recentes,
            "status_chart_labels": status_chart_labels,
            "status_chart_values": status_chart_values,
            "tipo_chart_labels": tipo_chart_labels,
            "tipo_chart_values": tipo_chart_values,
            "tipo_open_labels": tipo_open_labels,
            "tipo_open_values": tipo_open_values,
        },
    )



def _ld_has_field(nome):
    return any(field.name == nome for field in DocumentoLD._meta.get_fields())


def _ld_texto(valor):
    return str(valor or "").strip()


def _ld_bool(valor):
    return _ld_texto(valor).lower() in {"1", "true", "on", "sim", "yes"}



def _ld_valores_distintos(campo, extras=None):
    valores = []
    vistos = set()

    if _ld_has_field(campo):
        for valor in (
            DocumentoLD.objects.exclude(**{campo: ""})
            .exclude(**{f"{campo}__isnull": True})
            .values_list(campo, flat=True)
            .distinct()
            .order_by(campo)
        ):
            texto = _ld_texto(valor)
            chave = texto.lower()
            if texto and chave not in vistos:
                vistos.add(chave)
                valores.append(texto)

    for valor in extras or []:
        texto = _ld_texto(valor)
        chave = texto.lower()
        if texto and chave not in vistos:
            vistos.add(chave)
            valores.append(texto)

    return valores


def _ld_normalizar_origem(valor):
    texto = _ld_texto(valor).lower()
    texto = " ".join(texto.replace("_", " ").replace("-", " ").split())

    if not texto:
        return ""

    if "marenova" in texto:
        return "ld marenova"

    if texto in {"ld", "lista ld", "aba ld"}:
        return "ld"

    return texto


def _ld_filtrar_origem(queryset, origem):
    """
    Filtra a origem sem zerar a lista quando a importação antiga veio com
    origem_aba vazia, "LD", "Lista LD" ou variações de escrita.

    Regra operacional:
    - LD Marenova: registros cuja origem contém "Marenova".
    - LD: registros LD explícitos + registros sem origem + registros que não são Marenova.
    - Outras origens: busca flexível por texto.
    """
    origem = _ld_texto(origem)

    if not origem or not _ld_has_field("origem_aba"):
        return queryset

    origem_norm = _ld_normalizar_origem(origem)

    if origem_norm == "ld marenova":
        return queryset.filter(origem_aba__icontains="Marenova")

    if origem_norm == "ld":
        return queryset.filter(
            Q(origem_aba__isnull=True)
            | Q(origem_aba="")
            | Q(origem_aba__iexact="LD")
            | Q(origem_aba__iexact="Lista LD")
            | Q(origem_aba__icontains="LD")
        ).exclude(origem_aba__icontains="Marenova")

    return queryset.filter(origem_aba__icontains=origem)


def _ld_revisao_peso(revisao):
    texto = _ld_texto(revisao).upper()

    if not texto:
        return -1

    if texto.isdigit():
        return int(texto)

    peso = 0
    for char in texto:
        if "A" <= char <= "Z":
            peso = peso * 26 + (ord(char) - ord("A") + 1)

    return 1000 + peso


def _ld_filtrar_ultimas_revisoes(queryset):
    dados = list(queryset.values_list("pk", "documento", "revisao"))

    ultimos = {}

    for pk, documento, revisao in dados:
        doc = _ld_texto(documento).upper()
        peso = _ld_revisao_peso(revisao)

        if doc not in ultimos or peso > ultimos[doc][0]:
            ultimos[doc] = (peso, pk)

    ids = [pk for _, pk in ultimos.values()]

    return queryset.filter(pk__in=ids)


def _ld_filtrar_queryset(request):
    busca = _ld_texto(request.GET.get("q"))
    origem = _ld_texto(request.GET.get("origem"))
    disciplina = _ld_texto(request.GET.get("disciplina"))
    status_doc = _ld_texto(request.GET.get("status_doc"))
    status_grd = _ld_texto(request.GET.get("status_grd"))
    status_pcf = _ld_texto(request.GET.get("status_pcf"))

    com_pcf = _ld_bool(request.GET.get("com_pcf"))
    sem_pcf = _ld_bool(request.GET.get("sem_pcf"))
    com_resposta = _ld_bool(request.GET.get("com_resposta"))
    ultimas_revisoes = _ld_bool(request.GET.get("ultimas_revisoes"))

    filtro_rapido = _ld_texto(request.GET.get("filtro"))

    if filtro_rapido == "recebidos":
        status_doc = "Recebido"
    elif filtro_rapido == "aprovados":
        status_doc = "Aprovado"
    elif filtro_rapido == "grd_emitido":
        status_grd = "Emitido"
    elif filtro_rapido == "com_pcf":
        com_pcf = True
    elif filtro_rapido == "sem_pcf":
        sem_pcf = True
    elif filtro_rapido == "com_resposta":
        com_resposta = True
    elif filtro_rapido == "not_released":
        status_pcf = "NOT RELEASED"
    elif filtro_rapido == "ultimas_revisoes":
        ultimas_revisoes = True

    registros = DocumentoLD.objects.all().order_by("documento", "revisao")

    registros = _ld_filtrar_origem(registros, origem)

    if busca:
        registros = registros.filter(
            Q(documento__icontains=busca)
            | Q(titulo__icontains=busca)
            | Q(disciplina__icontains=busca)
            | Q(grd__icontains=busca)
            | Q(pcf__icontains=busca)
            | Q(pcf_resposta__icontains=busca)
            | Q(grd_resposta__icontains=busca)
        )

    if disciplina:
        registros = registros.filter(disciplina__icontains=disciplina)

    if status_doc:
        registros = registros.filter(status_documento__iexact=status_doc)

    if status_grd:
        registros = registros.filter(status_grd__iexact=status_grd)

    if status_pcf:
        status_pcf_norm = status_pcf.strip().upper()
        if status_pcf_norm in ["RELEASED", "NOT RELEASED"]:
            registros = registros.filter(status_final_pcf__iexact=status_pcf)
        else:
            registros = registros.filter(status_final_pcf__icontains=status_pcf)

    if com_pcf and not sem_pcf:
        registros = registros.exclude(pcf__isnull=True).exclude(pcf="")

    if sem_pcf and not com_pcf:
        registros = registros.filter(Q(pcf__isnull=True) | Q(pcf=""))

    if com_resposta:
        registros = registros.exclude(pcf_resposta__isnull=True).exclude(pcf_resposta="")

    if ultimas_revisoes:
        registros = _ld_filtrar_ultimas_revisoes(registros)

    filtros = {
        "busca": busca,
        "origem": origem,
        "disciplina": disciplina,
        "status_doc": status_doc,
        "status_grd": status_grd,
        "status_pcf": status_pcf,
        "com_pcf": com_pcf,
        "sem_pcf": sem_pcf,
        "com_resposta": com_resposta,
        "ultimas_revisoes": ultimas_revisoes,
        "filtro_rapido": filtro_rapido,
    }

    return registros, filtros


def _ld_kpis(registros):
    total = registros.count()

    return {
        "total": total,
        "total_exclusivos": registros.order_by().values("documento").distinct().count(),
        "total_recebidos": registros.filter(status_documento__iexact="Recebido").count(),
        "total_aprovados": registros.filter(status_documento__iexact="Aprovado").count(),
        "total_emitidos": registros.filter(status_grd__iexact="Emitido").count(),
        "total_com_pcf": registros.exclude(pcf__isnull=True).exclude(pcf="").count(),
        "total_sem_pcf": registros.filter(Q(pcf__isnull=True) | Q(pcf="")).count(),
        "total_com_resposta": registros.exclude(pcf_resposta__isnull=True).exclude(pcf_resposta="").count(),
    }


def _ld_resolver_caminho(caminho_salvo):
    return resolver_caminho_ld(caminho_salvo)


def _ld_hyperlink(caminho):
    return gerar_hyperlink_ld(caminho)




def _ld_querystring(request, updates=None, clears=None):
    """
    Monta querystring preservando os filtros atuais, removendo paginação
    e evitando parâmetros duplicados.

    Usado pelos chips rápidos da Lista LD.
    """
    query = request.GET.copy()
    query.pop("page", None)

    for key in clears or []:
        query.pop(key, None)

    for key, value in (updates or {}).items():
        query.pop(key, None)
        if value not in (None, "", False):
            query[key] = str(value)

    encoded = query.urlencode()
    return f"?{encoded}" if encoded else ""


def _ld_chip(label, query, active=False):
    return {
        "label": label,
        "query": query,
        "active": active,
    }


def _ld_montar_chips(request, filtros, status_documentos, status_grds):
    status_doc = _ld_texto(filtros.get("status_doc"))
    status_grd = _ld_texto(filtros.get("status_grd"))
    status_pcf = _ld_texto(filtros.get("status_pcf"))

    com_pcf = bool(filtros.get("com_pcf"))
    sem_pcf = bool(filtros.get("sem_pcf"))
    com_resposta = bool(filtros.get("com_resposta"))
    ultimas_revisoes = bool(filtros.get("ultimas_revisoes"))

    status_documento_chips = [
        _ld_chip(
            "Todos",
            _ld_querystring(
                request,
                clears=["status_doc", "filtro"],
            ),
            active=not status_doc,
        )
    ]

    for valor in status_documentos:
        status_documento_chips.append(
            _ld_chip(
                valor,
                _ld_querystring(
                    request,
                    updates={"status_doc": valor},
                    clears=["status_doc", "filtro"],
                ),
                active=status_doc.casefold() == _ld_texto(valor).casefold(),
            )
        )

    status_grd_chips = [
        _ld_chip(
            "Todos",
            _ld_querystring(
                request,
                clears=["status_grd", "filtro"],
            ),
            active=not status_grd,
        )
    ]

    for valor in status_grds:
        status_grd_chips.append(
            _ld_chip(
                valor,
                _ld_querystring(
                    request,
                    updates={"status_grd": valor},
                    clears=["status_grd", "filtro"],
                ),
                active=status_grd.casefold() == _ld_texto(valor).casefold(),
            )
        )

    operacional_chips = [
        _ld_chip(
            "Todos",
            "",
            active=not any(
                key != "page" and _ld_texto(value)
                for key, value in request.GET.items()
            ),
        ),
        _ld_chip(
            "Recebidos",
            _ld_querystring(
                request,
                updates={"status_doc": "Recebido"},
                clears=["status_doc", "filtro"],
            ),
            active=status_doc.casefold() == "recebido",
        ),
        _ld_chip(
            "Aprovados",
            _ld_querystring(
                request,
                updates={"status_doc": "Aprovado"},
                clears=["status_doc", "filtro"],
            ),
            active=status_doc.casefold() == "aprovado",
        ),
        _ld_chip(
            "GRD emitido",
            _ld_querystring(
                request,
                updates={"status_grd": "Emitido"},
                clears=["status_grd", "filtro"],
            ),
            active=status_grd.casefold() == "emitido",
        ),
        _ld_chip(
            "Com PCF",
            _ld_querystring(
                request,
                updates={"com_pcf": "1"},
                clears=["sem_pcf", "filtro"],
            ),
            active=com_pcf and not sem_pcf,
        ),
        _ld_chip(
            "Sem PCF",
            _ld_querystring(
                request,
                updates={"sem_pcf": "1"},
                clears=["com_pcf", "filtro"],
            ),
            active=sem_pcf and not com_pcf,
        ),
        _ld_chip(
            "Com resposta",
            _ld_querystring(
                request,
                updates={"com_resposta": "1"},
                clears=["filtro"],
            ),
            active=com_resposta,
        ),
        _ld_chip(
            "Not Released",
            _ld_querystring(
                request,
                updates={"status_pcf": "NOT RELEASED"},
                clears=["status_pcf", "filtro"],
            ),
            active=status_pcf.casefold() == "not released",
        ),
        _ld_chip(
            "Últimas revisões",
            _ld_querystring(
                request,
                updates={"ultimas_revisoes": "1"},
                clears=["filtro"],
            ),
            active=ultimas_revisoes,
        ),
    ]

    return {
        "status_documento_chips": status_documento_chips,
        "status_grd_chips": status_grd_chips,
        "operacional_chips": operacional_chips,
    }


@login_required
def listar_ld(request):
    tipo_doc = _ld_texto(request.GET.get("tipo_doc")).upper()

    registros, filtros = _ld_filtrar_queryset(request)

    tipos_encontrados = {
        extrair_tipo_documental(documento)
        for documento in registros.values_list("documento", flat=True)
    }
    tipos_documentais = sorted(tipo for tipo in tipos_encontrados if tipo)

    if tipo_doc:
        ids_tipo_doc = [
            pk
            for pk, documento in registros.values_list("pk", "documento")
            if extrair_tipo_documental(documento) == tipo_doc
        ]
        registros = registros.filter(pk__in=ids_tipo_doc)

    kpis = _ld_kpis(registros)

    disciplinas = _ld_valores_distintos("disciplina")
    origens = _ld_valores_distintos("origem_aba", extras=["LD", "LD Marenova"])
    # Mantém as duas origens operacionais sempre disponíveis, mesmo quando a
    # importação antiga gravou origem_aba em branco.
    origens_norm = []
    for origem_item in ["LD", "LD Marenova", *origens]:
        if origem_item not in origens_norm:
            origens_norm.append(origem_item)
    origens = origens_norm
    status_documentos = _ld_valores_distintos("status_documento")
    status_grds = _ld_valores_distintos("status_grd")
    chips_ld = _ld_montar_chips(
        request,
        filtros,
        status_documentos,
        status_grds,
    )

    paginator = Paginator(registros, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()

    return render(
        request,
        "automacoes/lista_ld.html",
        {
            "registros": page_obj,
            "page_obj": page_obj,
            "query_string": query_string,
            "tipos_documentais": tipos_documentais,
            "tipo_doc": tipo_doc,

            "origens": origens,
            "disciplinas": disciplinas,
            "status_documentos": status_documentos,
            "status_grds": status_grds,

            **chips_ld,
            **filtros,
            **kpis,
        },
    )



@login_required
def exportar_ld_excel(request):
    registros, _ = _ld_filtrar_queryset(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista LD Filtrada"

    headers = [
        "Origem",
        "Documento",
        "Revisão",
        "Disciplina",
        "Título",
        "Status Documento",
        "Status GRD",
        "Status PCF",
        "GRD",
        "Data GRD",
        "PCF",
        "Data PCF",
        "Resposta PCF",
        "Data Resposta",
        "GRD Resposta",
        "Caminho Documento",
        "Caminho GRD",
        "Caminho PCF",
        "Caminho Resposta",
        "Caminho GRD Resposta",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for item in registros:
        row = [
            getattr(item, "origem_aba", ""),
            item.documento,
            item.revisao,
            item.disciplina,
            item.titulo,
            item.status_documento,
            item.status_grd,
            item.status_final_pcf,
            item.grd,
            item.data_grd,
            item.pcf,
            item.data_pcf,
            item.pcf_resposta,
            item.data_resposta,
            item.grd_resposta,
            item.caminho_documento,
            item.caminho_grd,
            item.caminho_pcf,
            item.caminho_resposta,
            item.caminho_grd_resposta,
        ]

        ws.append(row)

        current_row = ws.max_row

        caminho_cols = {
            16: item.caminho_documento,
            17: item.caminho_grd,
            18: item.caminho_pcf,
            19: item.caminho_resposta,
            20: item.caminho_grd_resposta,
        }

        for col_idx, caminho in caminho_cols.items():
            if caminho:
                cell = ws.cell(row=current_row, column=col_idx)
                cell.hyperlink = _ld_hyperlink(caminho)
                cell.style = "Hyperlink"

    widths = {
        "A": 16,
        "B": 34,
        "C": 10,
        "D": 28,
        "E": 60,
        "F": 20,
        "G": 18,
        "H": 20,
        "I": 18,
        "J": 14,
        "K": 36,
        "L": 14,
        "M": 36,
        "N": 14,
        "O": 18,
        "P": 80,
        "Q": 80,
        "R": 80,
        "S": 80,
        "T": 80,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="lista_ld_filtrada.xlsx"'
    wb.save(response)

    return response



@login_required
def dashboard_ld(request):
    registros = DocumentoLD.objects.all()

    kpis = _ld_kpis(registros)

    total_not_released = registros.filter(
        status_final_pcf__iexact="NOT RELEASED"
    ).count()

    total_released = registros.filter(
        status_final_pcf__iexact="RELEASED"
    ).count()

    total_sem_status_doc = registros.filter(
        Q(status_documento__isnull=True) | Q(status_documento="")
    ).count()

    total_sem_grd = registros.filter(
        Q(status_grd__isnull=True) | Q(status_grd="")
    ).count()

    total_sem_resposta = registros.filter(
        Q(pcf_resposta__isnull=True) | Q(pcf_resposta="")
    ).exclude(
        Q(pcf__isnull=True) | Q(pcf="")
    ).count()

    por_disciplina = list(
        registros.values("disciplina")
        .annotate(total=Count("id"))
        .order_by("-total")[:12]
    )

    por_origem = list(
        registros.values("origem_aba")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    por_status_documento = list(
        registros.values("status_documento")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    por_status_grd = list(
        registros.values("status_grd")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    top_disciplinas_pendentes = list(
        registros.filter(Q(pcf__isnull=True) | Q(pcf=""))
        .values("disciplina")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    top_not_released = registros.filter(
        status_final_pcf__iexact="NOT RELEASED"
    ).order_by("documento", "revisao")[:12]

    recentes = registros.order_by("-id")[:12]

    disciplina_labels = [
        item.get("disciplina") or "Sem disciplina"
        for item in por_disciplina
    ]
    disciplina_values = [
        item.get("total") or 0
        for item in por_disciplina
    ]

    origem_labels = [
        item.get("origem_aba") or "Sem origem"
        for item in por_origem
    ]
    origem_values = [
        item.get("total") or 0
        for item in por_origem
    ]

    status_doc_labels = [
        item.get("status_documento") or "Sem status"
        for item in por_status_documento
    ]
    status_doc_values = [
        item.get("total") or 0
        for item in por_status_documento
    ]

    status_grd_labels = [
        item.get("status_grd") or "Sem status"
        for item in por_status_grd
    ]
    status_grd_values = [
        item.get("total") or 0
        for item in por_status_grd
    ]

    pcf_labels = ["Com PCF", "Sem PCF", "Com resposta", "Sem resposta"]
    pcf_values = [
        kpis["total_com_pcf"],
        kpis["total_sem_pcf"],
        kpis["total_com_resposta"],
        total_sem_resposta,
    ]

    taxa_pcf = 0
    if kpis["total"]:
        taxa_pcf = round((kpis["total_com_pcf"] / kpis["total"]) * 100, 1)

    taxa_grd = 0
    if kpis["total"]:
        taxa_grd = round((kpis["total_emitidos"] / kpis["total"]) * 100, 1)

    taxa_aprovacao = 0
    if kpis["total"]:
        taxa_aprovacao = round((kpis["total_aprovados"] / kpis["total"]) * 100, 1)

    return render(
        request,
        "automacoes/dashboard_ld.html",
        {
            **kpis,
            "total_not_released": total_not_released,
            "total_released": total_released,
            "total_sem_status_doc": total_sem_status_doc,
            "total_sem_grd": total_sem_grd,
            "total_sem_resposta": total_sem_resposta,
            "taxa_pcf": taxa_pcf,
            "taxa_grd": taxa_grd,
            "taxa_aprovacao": taxa_aprovacao,
            "por_disciplina": por_disciplina,
            "por_origem": por_origem,
            "por_status_documento": por_status_documento,
            "por_status_grd": por_status_grd,
            "top_disciplinas_pendentes": top_disciplinas_pendentes,
            "top_not_released": top_not_released,
            "recentes": recentes,
            "disciplina_labels": disciplina_labels,
            "disciplina_values": disciplina_values,
            "origem_labels": origem_labels,
            "origem_values": origem_values,
            "status_doc_labels": status_doc_labels,
            "status_doc_values": status_doc_values,
            "status_grd_labels": status_grd_labels,
            "status_grd_values": status_grd_values,
            "pcf_labels": pcf_labels,
            "pcf_values": pcf_values,
        },
    )


@login_required
def dashboard_transmittals(request):
    registros = TransmittalKM.objects.all()

    total_registros = registros.count()
    total_transmittals = registros.values("transmittal_numero").distinct().count()
    total_ok = registros.filter(status_parse__icontains="OK").count()
    total_pdf = registros.exclude(arquivo_pdf="").count()
    total_sem_pdf = registros.filter(Q(arquivo_pdf__isnull=True) | Q(arquivo_pdf="")).count()

    media_docs_transmittal = round(total_registros / total_transmittals, 1) if total_transmittals else 0

    por_pasta = list(
        registros.values("pasta")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    por_emissao = list(
        registros.values("emissao")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    top_transmittals = list(
        registros.values("transmittal_numero")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    ultimos = registros.order_by("-criado_em")[:15]

    pastas_labels = [item.get("pasta") or "Sem pasta" for item in por_pasta]
    pastas_values = [item.get("total") or 0 for item in por_pasta]

    emissoes_labels = [item.get("emissao") or "Sem emissão" for item in por_emissao]
    emissoes_values = [item.get("total") or 0 for item in por_emissao]

    transmittal_labels = [item.get("transmittal_numero") or "Sem número" for item in top_transmittals]
    transmittal_values = [item.get("total") or 0 for item in top_transmittals]

    return render(
        request,
        "automacoes/dashboard_transmittals.html",
        {
            "total_registros": total_registros,
            "total_transmittals": total_transmittals,
            "total_ok": total_ok,
            "total_pdf": total_pdf,
            "total_sem_pdf": total_sem_pdf,
            "media_docs_transmittal": media_docs_transmittal,
            "por_pasta": por_pasta,
            "por_emissao": por_emissao,
            "top_transmittals": top_transmittals,
            "ultimos": ultimos,
            "pastas_labels": pastas_labels,
            "pastas_values": pastas_values,
            "emissoes_labels": emissoes_labels,
            "emissoes_values": emissoes_values,
            "transmittal_labels": transmittal_labels,
            "transmittal_values": transmittal_values,
        },
    )



@login_required
def abrir_arquivo_ld(request, pk, tipo):
    registro = DocumentoLD.objects.get(pk=pk)

    mapa = {
        "documento": registro.caminho_documento,
        "grd": registro.caminho_grd,
        "pcf": registro.caminho_pcf,
        "resposta": registro.caminho_resposta,
        "grd-resposta": registro.caminho_grd_resposta,
    }

    caminho_salvo = mapa.get(tipo)

    if not caminho_salvo:
        raise Http404("Caminho não localizado para este item da LD.")

    arquivo, candidatos = _ld_resolver_caminho(caminho_salvo)

    if not arquivo:
        html = "<h3>Arquivo ou pasta não encontrado.</h3>"
        html += f"<p><strong>Caminho salvo no banco:</strong> {_ld_texto(caminho_salvo)}</p>"
        html += "<p><strong>Caminhos testados:</strong></p><ul>"

        for candidato in candidatos:
            html += f"<li>{candidato}</li>"

        html += "</ul>"
        html += "<p>Verifique se a pasta da rede está acessível e se o caminho salvo continua válido.</p>"

        return HttpResponse(html, status=404)

    try:
        if arquivo.is_dir():
            os.startfile(str(arquivo))
            html = "<h3>Pasta aberta no Windows Explorer.</h3>"
            html += f"<p><strong>Caminho:</strong> {arquivo}</p>"
            html += "<p>Você pode fechar esta aba.</p>"
            return HttpResponse(html)

        return FileResponse(
            open(arquivo, "rb"),
            as_attachment=False,
            filename=arquivo.name,
        )

    except Exception as exc:
        html = "<h3>Arquivo ou pasta localizado, mas não foi possível abrir automaticamente.</h3>"
        html += f"<p><strong>Caminho:</strong> {arquivo}</p>"
        html += f"<p><strong>Erro:</strong> {exc}</p>"
        return HttpResponse(html, status=500)


# ============================================================
# BUSCA GLOBAL GED ENTERPRISE
# ============================================================

def _bg_texto(valor):
    return str(valor or "").strip()


def _bg_url(path):
    return path


def _bg_score(q, *valores):
    q_norm = _km_normalizar(q)
    if not q_norm:
        return 0

    melhor = 0
    for valor in valores:
        texto = _bg_texto(valor)
        if not texto:
            continue

        texto_norm = _km_normalizar(texto)

        if texto_norm == q_norm:
            melhor = max(melhor, 100)
        elif q_norm in texto_norm:
            melhor = max(melhor, 85)
        elif texto_norm in q_norm and len(texto_norm) >= 6:
            melhor = max(melhor, 70)
        elif q.lower() in texto.lower():
            melhor = max(melhor, 60)

    return melhor


def _bg_limite(qs, limite=20):
    return list(qs[:limite])


@login_required
def abrir_km_index(request, pk):
    item = KMFileIndex.objects.get(pk=pk, ativo=True)
    arquivo = Path(item.caminho_completo)

    if not arquivo.exists():
        raise Http404(f"Arquivo KM não encontrado: {arquivo}")

    if os.name == "nt":
        os.startfile(str(arquivo))
        return HttpResponse(
            f"Arquivo aberto: {arquivo}",
            content_type="text/plain; charset=utf-8",
        )

    return FileResponse(
        open(arquivo, "rb"),
        as_attachment=False,
        filename=arquivo.name,
    )


@login_required
def abrir_pasta_km_index(request, pk):
    item = KMFileIndex.objects.get(pk=pk, ativo=True)
    arquivo = Path(item.caminho_completo)
    pasta = arquivo.parent if arquivo.suffix else arquivo

    if not pasta.exists():
        raise Http404(f"Pasta KM não encontrada: {pasta}")

    if os.name == "nt":
        os.startfile(str(pasta))
        return HttpResponse(
            f"Pasta aberta: {pasta}",
            content_type="text/plain; charset=utf-8",
        )

    return HttpResponse(
        f"Pasta localizada: {pasta}",
        content_type="text/plain; charset=utf-8",
    )




@login_required
def dashboard_search(request):
    dias = request.GET.get("dias", 30)
    contexto = obter_search_analytics(dias=dias)

    return render(
        request,
        "automacoes/dashboard_search.html",
        contexto,
    )

@login_required
def busca_global_ged(request):
    q = (request.GET.get("q") or request.GET.get("busca") or "").strip()
    tipo = (request.GET.get("tipo") or "todos").strip().lower()

    contexto = buscar_global_enterprise(
        q,
        tipo=tipo,
        usuario=request.user,
        origem="web",
        auditar=bool(q),
    )

    return render(
        request,
        "automacoes/busca_global.html",
        contexto,
    )


@login_required
def api_busca_global_ged(request):
    q = _bg_texto(request.GET.get("q") or request.GET.get("busca"))
    if len(q) < 2:
        return JsonResponse({"results": []})

    q_norm = _km_normalizar(q)
    results = []

    for item in KMFileIndex.objects.filter(ativo=True).filter(
        Q(nome_arquivo__icontains=q)
        | Q(documento_extraido__icontains=q)
        | Q(nome_normalizado__icontains=q_norm)
        | Q(stem_normalizado__icontains=q_norm)
    ).order_by("eh_transmittal_letter", "nome_arquivo")[:8]:
        results.append({
            "type": "KM",
            "title": item.nome_arquivo,
            "subtitle": item.documento_extraido or item.pasta,
            "url": f"/automacoes/km-index/{item.id}/abrir/",
        })

    for item in DocumentoLD.objects.filter(
        Q(documento__icontains=q) | Q(titulo__icontains=q)
    ).order_by("documento")[:5]:
        results.append({
            "type": "LD",
            "title": item.documento,
            "subtitle": item.titulo[:120] if item.titulo else "",
            "url": f"/automacoes/ld/?q={q}",
        })

    for item in TransmittalKM.objects.filter(
        Q(documento__icontains=q)
        | Q(titulo__icontains=q)
        | Q(transmittal_numero__icontains=q)
    ).order_by("transmittal_numero")[:5]:
        results.append({
            "type": "Transmittal",
            "title": item.documento or item.transmittal_numero,
            "subtitle": item.titulo[:120] if item.titulo else item.transmittal_numero,
            "url": f"/automacoes/transmittals-km/?q={q}",
        })

    return JsonResponse({"results": results[:15]})



# ============================================================
# UNIFIED OPERATIONS CENTER
# ============================================================

@login_required
def ops_center(request):
    contexto = {
        "ops": OperationsCenterService.build_dashboard(),
    }

    return render(
        request,
        "automacoes/ops_center.html",
        contexto,
    )



@login_required
def ops_center_live_partial(request):
    return render(
        request,
        "automacoes/partials/_ops_live_operations.html",
        LiveOperationsService.build_payload(),
    )

