# -*- coding: utf-8 -*-
"""
Consolidador de GRDs PDF -> Excel
Processa automaticamente:
- GASEIRO 14K
- GASEIRO 7K

Extrai da tabela dos PDFs GRD as colunas:
Nº / Número do Documento / Título do Documento / Revisão / Finalidade da Emissão / Data / Obs.

Gera:
- LD recebidos GHENOVA 14K.xlsx
- LD recebidos GHENOVA 7K.xlsx
"""

import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

fitz = None
pd = None
Font = None
PatternFill = None
Border = None
Side = None
Alignment = None
get_column_letter = None


def install_if_missing(pkg, import_name=None):
    try:
        __import__(import_name if import_name else pkg)
    except ImportError:
        print(f"📦 Instalando dependência: {pkg} ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg])


# ==========================================================
# CONFIGURAÇÕES
# ==========================================================
BASE_DIRS = [
    {
        "empreendimento": "14K",
        "base_dir": Path(
            r"\\virm-rgr022\FILESERVER\Departamento\DENG - Engenharia\98 - ESTUDOS\TRANSPETRO GASEIROS\Gaseiro 14k\(001) DOCUMENTOS DE ENGENHARIA"
        ),
        "output_filename": "LD recebidos GHENOVA 14K.xlsx",
    },
    {
        "empreendimento": "7K",
        "base_dir": Path(
            r"\\virm-rgr022\FILESERVER\Departamento\DENG - Engenharia\98 - ESTUDOS\TRANSPETRO GASEIROS\Gaseiro 7k\(001) DOCUMENTOS DE ENGENHARIA"
        ),
        "output_filename": "LD recebidos GHENOVA 7K.xlsx",
    },
]

GRD_PDF_NAME_RE = re.compile(r"ERG005-0000-GRD-\d+.*\.pdf$", re.IGNORECASE)
DOC_CODE_RE = re.compile(r"^[A-Z0-9]+(?:-[A-Z0-9]+)+$", re.IGNORECASE)
REV_RE = re.compile(r"^[A-Z0-9]{1,4}$", re.IGNORECASE)
DATE_RE = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
GRD_ID_RE = re.compile(r"\bERG005-0000-GRD-\d{4}(?:-[A-Z0-9]+)?\b", re.IGNORECASE)

FINALIDADES_VALIDAS = [
    "Para Comentários",
    "Para Aprovação",
    "Para Conhecimento",
    "Para Construção",
    "Para Fabricação",
    "Para Informação",
    "As Built",
    "For Comment",
    "For Approval",
    "For Information",
    "For Construction",
    "For Review",
]


# ==========================================================
# UTILITÁRIOS
# ==========================================================
def normalizar_linhas(texto: str) -> list[str]:
    texto = texto.replace("\u00a0", " ")
    linhas = []
    for linha in texto.splitlines():
        linha = re.sub(r"[ \t]+", " ", linha).strip()
        if linha:
            linhas.append(linha)
    return linhas


def extract_text(pdf_path: Path) -> str:
    partes = []
    with fitz.open(pdf_path) as doc:
        for page in doc:
            partes.append(page.get_text("text") or "")
    return "\n".join(partes)


def encontrar_data(linhas: list[str]) -> str:
    for i, linha in enumerate(linhas):
        m = DATE_RE.search(linha)
        if m:
            return m.group(0)

        if linha.lower() == "data" and i + 1 < len(linhas):
            m2 = DATE_RE.search(linhas[i + 1])
            if m2:
                return m2.group(0)
    return ""


def encontrar_grd_id(linhas: list[str], pdf_path: Path) -> str:
    for linha in linhas:
        m = GRD_ID_RE.search(linha)
        if m:
            return m.group(0).strip()
    return pdf_path.parent.name


def eh_finalidade(texto: str) -> bool:
    texto_norm = texto.strip().lower()
    for item in FINALIDADES_VALIDAS:
        if texto_norm == item.lower():
            return True
    return False


def parece_inicio_rodape(texto: str) -> bool:
    texto_low = texto.lower()
    termos_rodape = [
        "projeto de conceitual",
        "guia de remessa de documentos",
        "erg005-0000-grd-",
        "erg005",
    ]
    return any(t in texto_low for t in termos_rodape)


def localizar_inicio_dados(linhas: list[str]) -> int:
    for i in range(len(linhas) - 5):
        bloco = " | ".join(linhas[i : i + 6]).lower()
        if (
            "nº" in bloco
            and "número do documento" in bloco
            and "título do documento" in bloco
            and "revisão" in bloco
            and "finalidade da emissão" in bloco
        ):
            return i + 6

    for i, linha in enumerate(linhas):
        if linha.lower().startswith("obs"):
            return i + 1

    return -1


def limpar_numero_documento(doc: str) -> str:
    return doc.strip().rstrip("-").strip()


# ==========================================================
# PARSER PRINCIPAL
# ==========================================================
def parse_grd(pdf_path: Path, empreendimento: str) -> list[dict]:
    texto = extract_text(pdf_path)
    if not texto.strip():
        return []

    linhas = normalizar_linhas(texto)
    if not linhas:
        return []

    data_emissao = encontrar_data(linhas)
    grd_id = encontrar_grd_id(linhas, pdf_path)
    inicio = localizar_inicio_dados(linhas)

    if inicio < 0 or inicio >= len(linhas):
        print("⚠️ Header da tabela não encontrado.")
        return []

    registros = []
    i = inicio

    while i < len(linhas):
        linha = linhas[i]

        if parece_inicio_rodape(linha):
            break

        if not linha.isdigit():
            i += 1
            continue

        numero_item = linha.strip()

        if i + 1 >= len(linhas):
            break

        numero_documento = limpar_numero_documento(linhas[i + 1].strip())

        if not DOC_CODE_RE.match(numero_documento):
            i += 1
            continue

        j = i + 2
        titulo_partes = []
        revisao = ""
        finalidade = ""

        while j < len(linhas):
            atual = linhas[j].strip()

            if REV_RE.match(atual) and (j + 1) < len(linhas) and eh_finalidade(linhas[j + 1]):
                revisao = atual
                finalidade = linhas[j + 1].strip()
                j += 2
                break

            if parece_inicio_rodape(atual):
                break

            titulo_partes.append(atual)
            j += 1

        titulo_documento = " ".join(titulo_partes).strip()

        # aceita item mesmo que venha truncado no fim do PDF,
        # desde que tenha Nº + Documento + Título
        if not titulo_documento:
            i += 1
            continue

        obs_partes = []
        while j < len(linhas):
            atual = linhas[j].strip()

            if parece_inicio_rodape(atual):
                break

            if atual.isdigit() and (j + 1) < len(linhas):
                prox_doc = limpar_numero_documento(linhas[j + 1].strip())
                if DOC_CODE_RE.match(prox_doc):
                    break

            obs_partes.append(atual)
            j += 1

        obs = " ".join(obs_partes).strip()

        registros.append(
            {
                "Empreendimento": empreendimento,
                "GRD": grd_id,
                "Subpasta": pdf_path.parent.name,
                "Pasta": str(pdf_path.parent),
                "Arquivo": pdf_path.name,
                "Data": data_emissao,
                "Nº": int(numero_item),
                "Número do Documento": numero_documento,
                "Título do Documento": titulo_documento,
                "Revisão": revisao,
                "Finalidade da Emissão": finalidade,
                "Obs.": obs,
            }
        )

        if j == i:
            i += 1
        else:
            i = j

    return registros


# ==========================================================
# BUSCA DE PDFs
# ==========================================================
def find_grd_pdfs(base_dir: Path) -> list[Path]:
    pdfs = []
    for pdf in base_dir.rglob("*.pdf"):
        if GRD_PDF_NAME_RE.search(pdf.name):
            pdfs.append(pdf)
    return sorted(pdfs)


# ==========================================================
# FORMATAÇÃO EXCEL
# ==========================================================
def aplicar_formatacao_profissional(ws, total_rows_dados: int, total_cols: int, titulo: str):
    azul_escuro = "1F4E78"
    azul = "2F75B5"
    azul_claro = "D9EAF7"
    faixa = "F7FBFF"
    branco = "FFFFFF"
    cinza_borda = "B7C9D6"
    cinza_topo = "EAF2F8"

    thin = Side(style="thin", color=cinza_borda)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.insert_rows(1, amount=2)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1)
    c.value = titulo
    c.font = Font(name="Arial", size=12, bold=True, color=branco)
    c.fill = PatternFill("solid", fgColor=azul_escuro)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Subtítulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c2 = ws.cell(row=2, column=1)
    c2.value = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    c2.font = Font(name="Arial", size=10, italic=True, color="1F1F1F")
    c2.fill = PatternFill("solid", fgColor=azul_claro)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    header_row = 3
    data_start_row = 4
    data_end_row = total_rows_dados + 3

    # Cabeçalho
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(name="Arial", size=10, bold=True, color=branco)
        cell.fill = PatternFill("solid", fgColor=azul)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Dados
    for row in range(data_start_row, data_end_row + 1):
        fill_color = faixa if (row - data_start_row) % 2 == 0 else branco
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border

    # Centralização de colunas específicas
    headers = {ws.cell(row=header_row, column=col).value: col for col in range(1, total_cols + 1)}
    for nome_coluna in ["Empreendimento", "Data", "Nº", "Revisão"]:
        if nome_coluna in headers:
            col_idx = headers[nome_coluna]
            for row in range(data_start_row, data_end_row + 1):
                ws.cell(row=row, column=col_idx).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

    # Auto filtro normal
    last_col_letter = get_column_letter(total_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{data_end_row}"

    # Congelar painel
    ws.freeze_panes = "A4"

    # Altura das linhas
    for row in range(data_start_row, data_end_row + 1):
        ws.row_dimensions[row].height = 20

    # Ajuste de largura sem erro de MergedCell
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, data_end_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            valor = "" if cell.value is None else str(cell.value)
            if len(valor) > max_len:
                max_len = len(valor)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

    # Formato da data
    if "Data" in headers:
        col_idx = headers["Data"]
        for row in range(data_start_row, data_end_row + 1):
            ws.cell(row=row, column=col_idx).number_format = "DD/MM/YYYY"

    # Destaque leve no topo da tabela
    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.border = Border(
            left=thin,
            right=thin,
            top=Side(style="medium", color=azul_escuro),
            bottom=thin,
        )

    # Faixa visual nas 2 primeiras linhas
    for row in [1, 2]:
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            if row == 2:
                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=Side(style="medium", color=cinza_topo),
                )


def salvar_excel(df, output_path: Path, empreendimento: str):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="LD")

        ws = writer.book["LD"]
        aplicar_formatacao_profissional(
            ws=ws,
            total_rows_dados=len(df),
            total_cols=len(df.columns),
            titulo=f"LD RECEBIDOS GHENOVA - {empreendimento}",
        )


# ==========================================================
# PROCESSAMENTO DE CADA EMPREENDIMENTO
# ==========================================================
def processar_base(base_dir: Path, empreendimento: str, output_filename: str):
    if not base_dir.exists():
        print(f"❌ Pasta base não encontrada: {base_dir}")
        return

    output_xlsx = base_dir / output_filename

    pdfs = find_grd_pdfs(base_dir)
    print(f"\n{'=' * 80}")
    print(f"📁 Empreendimento: {empreendimento}")
    print(f"📂 Pasta base: {base_dir}")
    print(f"🔎 GRDs encontrados: {len(pdfs)}")

    if not pdfs:
        print("⚠️ Nenhum PDF GRD encontrado.")
        return

    todos_registros = []

    for idx, pdf_path in enumerate(pdfs, start=1):
        rel = pdf_path.relative_to(base_dir)
        print(f"\n[{idx}/{len(pdfs)}] Processando: {rel}")

        try:
            registros = parse_grd(pdf_path, empreendimento)
        except Exception as e:
            print(f"❌ Erro ao processar {pdf_path.name}: {e}")
            continue

        if registros:
            print(f"✅ Linhas extraídas: {len(registros)}")
            todos_registros.extend(registros)
        else:
            print("⚠️ Nenhuma linha de tabela reconhecida nesse PDF.")
            try:
                texto_debug = extract_text(pdf_path)
                linhas_debug = normalizar_linhas(texto_debug)
                print("🧪 DEBUG - primeiras 30 linhas extraídas:")
                for n, linha in enumerate(linhas_debug[:30], start=1):
                    print(f"{n:02d}: {linha}")
            except Exception as dbg_e:
                print(f"⚠️ Falha no debug: {dbg_e}")

    if not todos_registros:
        print(f"\n❌ Nenhuma linha foi extraída no empreendimento {empreendimento}.")
        return

    df = pd.DataFrame(
        todos_registros,
        columns=[
            "Empreendimento",
            "GRD",
            "Subpasta",
            "Pasta",
            "Arquivo",
            "Data",
            "Nº",
            "Número do Documento",
            "Título do Documento",
            "Revisão",
            "Finalidade da Emissão",
            "Obs.",
        ],
    )

    df.sort_values(by=["GRD", "Nº"], inplace=True, kind="stable")
    salvar_excel(df, output_xlsx, empreendimento)

    print(f"\n✅ Concluído com sucesso - {empreendimento}")
    print(f"💾 Excel salvo em: {output_xlsx}")
    print(f"📄 Total de linhas consolidadas: {len(df)}")


def main():
    for item in BASE_DIRS:
        processar_base(
            base_dir=item["base_dir"],
            empreendimento=item["empreendimento"],
            output_filename=item["output_filename"],
        )

    print(f"\n{'=' * 80}")
    print("✅ Processamento finalizado para todos os empreendimentos.")


def executar():
    global fitz, pd, Font, PatternFill, Border, Side, Alignment, get_column_letter

    try:
        install_if_missing("pymupdf", "fitz")
        install_if_missing("pandas")
        install_if_missing("openpyxl")

        import fitz as fitz_module
        import pandas as pandas_module
        from openpyxl.styles import Alignment as AlignmentType
        from openpyxl.styles import Border as BorderType
        from openpyxl.styles import Font as FontType
        from openpyxl.styles import PatternFill as PatternFillType
        from openpyxl.styles import Side as SideType
        from openpyxl.utils import get_column_letter as get_column_letter_fn

        fitz = fitz_module
        pd = pandas_module
        Font = FontType
        PatternFill = PatternFillType
        Border = BorderType
        Side = SideType
        Alignment = AlignmentType
        get_column_letter = get_column_letter_fn

        main()

        return {
            "ok": True,
            "mensagem": "GRD GHENOVA executado com sucesso.",
            "detalhes": {
                "bases": [str(item.get("base_dir")) for item in BASE_DIRS],
                "outputs": [item.get("output_filename") for item in BASE_DIRS],
            },
        }
    except Exception as e:
        return {
            "ok": False,
            "mensagem": f"Erro na GRD GHENOVA: {e}",
            "detalhes": {"erro": str(e), "tipo": e.__class__.__name__},
        }
