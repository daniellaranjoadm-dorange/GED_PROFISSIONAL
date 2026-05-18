"""
Importador real da Lista de Documentos Kongsberg/KM.

Entrada esperada:
- Planilha XLSX com aba "LD_KM"
- Cabeçalho normalmente na linha 2
- Coluna "Number" como chave mestre do DocumentoKM

Este service é propositalmente defensivo:
- ignora fórmulas quebradas como #NAME?
- usa introspecção de campos para evitar quebrar migrations antigas
- expõe aliases compatíveis: importar_ld_kongsberg e importar_lista_kongsberg
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook

from apps.automacoes.models import DocumentoKM, DocumentoLD, TransmittalKM


VALORES_INVALIDOS = {
    "#NAME?",
    "#VALUE!",
    "#REF!",
    "#DIV/0!",
    "#N/A",
    "#NULL!",
    "#NUM!",
}


COLUNAS_MAPEADAS = {
    "phase": "phase",
    "toc": "toc",
    "number": "numero_km",
    "title": "titulo",
    "discipline": "disciplina",
    "contractual delivery": "contractual_delivery",
    "preliminay delivery": "preliminary_delivery",
    "preliminary delivery": "preliminary_delivery",
    "agreed delivery": "agreed_delivery",
    "first delivery": "first_delivery",
    "released for": "released_for",
    "status": "status_km",
    "core share document": "core_share_document",
    "core share folder": "core_share_folder",
    "transmittal number": "transmittal_numero",
    "data recebimento km": "data_recebimento_km",
    "numero documento tp": "documento_tp",
    "número documento tp": "documento_tp",
}


def _texto(valor: Any) -> str:
    if valor is None:
        return ""

    texto = str(valor).strip()

    if not texto:
        return ""

    if texto.upper() in VALORES_INVALIDOS:
        return ""

    if texto.startswith("="):
        return ""

    return texto


def _normalizar_chave(valor: Any) -> str:
    texto = _texto(valor).lower()
    texto = texto.replace("\n", " ").replace("\r", " ")
    texto = " ".join(texto.split())
    return texto


def _compactar_documento(valor: Any) -> str:
    return "".join(ch for ch in _texto(valor).upper() if ch.isalnum())


def _model_has_field(model, nome: str) -> bool:
    return any(field.name == nome for field in model._meta.get_fields())


def _detectar_aba(workbook):
    if "LD_KM" in workbook.sheetnames:
        return workbook["LD_KM"]

    for nome in workbook.sheetnames:
        if nome.strip().lower() in {"ld km", "ld_km", "document list", "documentos km"}:
            return workbook[nome]

    return workbook.active


def _detectar_cabecalho(sheet) -> tuple[int, dict[str, int]]:
    """
    Procura uma linha de cabeçalho contendo pelo menos Number e Title.
    Retorna: (linha_cabecalho, mapa_coluna_normalizada_para_indice_1_based)
    """
    melhor_linha = 1
    melhor_mapa = {}

    for row_idx in range(1, min(sheet.max_row, 15) + 1):
        mapa = {}

        for col_idx in range(1, sheet.max_column + 1):
            chave = _normalizar_chave(sheet.cell(row=row_idx, column=col_idx).value)
            if chave:
                mapa[chave] = col_idx

        score = 0
        if "number" in mapa:
            score += 3
        if "title" in mapa:
            score += 2
        if "discipline" in mapa:
            score += 1

        if score > len(melhor_mapa):
            melhor_linha = row_idx
            melhor_mapa = mapa

        if "number" in mapa and "title" in mapa:
            return row_idx, mapa

    return melhor_linha, melhor_mapa


def _valor_linha(sheet, row_idx: int, colunas: dict[str, int], nome_coluna: str) -> str:
    col_idx = colunas.get(nome_coluna)
    if not col_idx:
        return ""
    return _texto(sheet.cell(row=row_idx, column=col_idx).value)


def _montar_defaults(sheet, row_idx: int, colunas: dict[str, int], origem_planilha: str) -> dict:
    defaults = {}

    for coluna_origem, campo_model in COLUNAS_MAPEADAS.items():
        if campo_model == "numero_km":
            continue

        if not _model_has_field(DocumentoKM, campo_model):
            continue

        valor = _valor_linha(sheet, row_idx, colunas, coluna_origem)
        defaults[campo_model] = valor

    if _model_has_field(DocumentoKM, "origem_planilha"):
        defaults["origem_planilha"] = origem_planilha

    if _model_has_field(DocumentoKM, "linha_origem"):
        defaults["linha_origem"] = row_idx

    return defaults


def importar_lista_kongsberg(arquivo, usuario=None, origem_planilha: str | None = None) -> dict:
    """
    Importa a LD Kongsberg para DocumentoKM.

    Aceita:
    - caminho string/path
    - UploadedFile do Django
    - file-like object
    """
    origem = origem_planilha or getattr(arquivo, "name", "") or str(arquivo)

    wb = load_workbook(arquivo, data_only=True, read_only=True)
    sheet = _detectar_aba(wb)
    header_row, colunas = _detectar_cabecalho(sheet)

    if "number" not in colunas:
        return {
            "ok": False,
            "mensagem": "Coluna obrigatória 'Number' não encontrada na LD Kongsberg.",
            "aba": sheet.title,
            "linha_cabecalho": header_row,
            "processados": 0,
            "criados": 0,
            "atualizados": 0,
            "ignorados": 0,
        }

    processados = 0
    criados = 0
    atualizados = 0
    ignorados = 0
    erros = []

    with transaction.atomic():
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            numero_km = _valor_linha(sheet, row_idx, colunas, "number")

            if not numero_km:
                ignorados += 1
                continue

            defaults = _montar_defaults(sheet, row_idx, colunas, origem)

            try:
                _, created = DocumentoKM.objects.update_or_create(
                    numero_km=numero_km,
                    defaults=defaults,
                )
                processados += 1

                if created:
                    criados += 1
                else:
                    atualizados += 1

            except Exception as exc:
                erros.append(
                    {
                        "linha": row_idx,
                        "numero_km": numero_km,
                        "erro": str(exc),
                    }
                )

    return {
        "ok": not erros,
        "mensagem": (
            f"LD Kongsberg importada: {processados} processados, "
            f"{criados} criados, {atualizados} atualizados, {ignorados} ignorados."
        ),
        "aba": sheet.title,
        "linha_cabecalho": header_row,
        "processados": processados,
        "criados": criados,
        "atualizados": atualizados,
        "ignorados": ignorados,
        "erros": erros[:20],
        "quantidade_processada": processados,
    }


def importar_ld_kongsberg(*args, **kwargs) -> dict:
    """Alias de compatibilidade usado por views existentes."""
    return importar_lista_kongsberg(*args, **kwargs)


def _buscar_transmittal_para_km(numero_km: str):
    compacto = _compactar_documento(numero_km)
    if not compacto:
        return None

    candidatos = TransmittalKM.objects.exclude(documento="").order_by("-id")[:5000]

    for item in candidatos:
        if _compactar_documento(item.documento) == compacto:
            return item

    for item in candidatos:
        doc_compacto = _compactar_documento(item.documento)
        if compacto and (compacto in doc_compacto or doc_compacto in compacto):
            return item

    return None


def _buscar_ld_para_km(numero_km: str):
    compacto = _compactar_documento(numero_km)
    if not compacto:
        return None, 0

    campos = [
        "numero_documento_km",
        "documento",
        "titulo",
        "caminho_documento",
        "caminho_grd",
        "caminho_pcf",
        "caminho_resposta",
        "caminho_grd_resposta",
    ]

    query = Q()
    for campo in campos:
        if _model_has_field(DocumentoLD, campo):
            query |= Q(**{f"{campo}__icontains": numero_km})

    candidatos = DocumentoLD.objects.filter(query).order_by("-id")[:500] if query else []

    melhor = None
    melhor_score = 0

    for item in candidatos:
        for campo in campos:
            if not _model_has_field(DocumentoLD, campo):
                continue

            valor = getattr(item, campo, "")
            valor_compacto = _compactar_documento(valor)

            score = 0
            if valor_compacto == compacto:
                score = 100
            elif compacto and compacto in valor_compacto:
                score = 85
            elif valor_compacto and valor_compacto in compacto:
                score = 70

            if score > melhor_score:
                melhor = item
                melhor_score = score

    return melhor, melhor_score


def executar_cruzamento_ld_km(limite: int | None = None) -> dict:
    """
    Cruza DocumentoKM importado com TransmittalKM e DocumentoLD.

    Atualiza:
    - status_recebimento/transmittal/data_recebimento
    - documento_ld/documento_tp/status_vinculo_ld/score_vinculo_ld
    """
    qs = DocumentoKM.objects.all().order_by("numero_km")
    if limite:
        qs = qs[: int(limite)]

    processados = 0
    recebidos = 0
    pendentes_recebimento = 0
    vinculados = 0
    sem_vinculo = 0

    for doc_km in qs:
        update_fields = []

        transmittal = _buscar_transmittal_para_km(doc_km.numero_km)

        if transmittal:
            if _model_has_field(DocumentoKM, "status_recebimento"):
                doc_km.status_recebimento = DocumentoKM.STATUS_RECEBIMENTO_RECEBIDO
                update_fields.append("status_recebimento")

            if _model_has_field(DocumentoKM, "transmittal_numero"):
                doc_km.transmittal_numero = _texto(transmittal.transmittal_numero)
                update_fields.append("transmittal_numero")

            if _model_has_field(DocumentoKM, "data_recebimento_km"):
                doc_km.data_recebimento_km = _texto(transmittal.data_envio)
                update_fields.append("data_recebimento_km")

            recebidos += 1
        else:
            if _model_has_field(DocumentoKM, "status_recebimento"):
                doc_km.status_recebimento = DocumentoKM.STATUS_RECEBIMENTO_PENDENTE
                update_fields.append("status_recebimento")
            pendentes_recebimento += 1

        item_ld, score = _buscar_ld_para_km(doc_km.numero_km)

        if item_ld and score >= 70:
            if _model_has_field(DocumentoKM, "documento_ld"):
                doc_km.documento_ld = item_ld
                update_fields.append("documento_ld")

            if _model_has_field(DocumentoKM, "documento_tp"):
                doc_km.documento_tp = _texto(getattr(item_ld, "documento", ""))
                update_fields.append("documento_tp")

            if _model_has_field(DocumentoKM, "status_vinculo_ld"):
                doc_km.status_vinculo_ld = DocumentoKM.STATUS_VINCULO_LD_AUTO
                update_fields.append("status_vinculo_ld")

            if _model_has_field(DocumentoKM, "score_vinculo_ld"):
                doc_km.score_vinculo_ld = score
                update_fields.append("score_vinculo_ld")

            if _model_has_field(DocumentoLD, "numero_documento_km"):
                item_ld.numero_documento_km = doc_km.numero_km
                item_ld.save(update_fields=["numero_documento_km"])

            vinculados += 1
        else:
            if _model_has_field(DocumentoKM, "status_vinculo_ld"):
                doc_km.status_vinculo_ld = DocumentoKM.STATUS_VINCULO_LD_SEM_MATCH
                update_fields.append("status_vinculo_ld")

            if _model_has_field(DocumentoKM, "score_vinculo_ld"):
                doc_km.score_vinculo_ld = score
                update_fields.append("score_vinculo_ld")

            sem_vinculo += 1

        if update_fields:
            update_fields = sorted(set(update_fields + ["atualizado_em"])) if _model_has_field(DocumentoKM, "atualizado_em") else sorted(set(update_fields))
            doc_km.save(update_fields=update_fields)

        processados += 1

    return {
        "ok": True,
        "mensagem": (
            f"Cruzamento KM concluído: {processados} processados, "
            f"{recebidos} recebidos, {vinculados} vinculados à LD."
        ),
        "processados": processados,
        "recebidos": recebidos,
        "pendentes_recebimento": pendentes_recebimento,
        "vinculados_ld": vinculados,
        "sem_vinculo_ld": sem_vinculo,
        "quantidade_processada": processados,
    }


# Alias adicional para scheduler/jobs futuros.
executar_cruzamento_km_ld = executar_cruzamento_ld_km
