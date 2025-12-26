from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.storage import default_storage
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.contrib.auth.models import Group
from datetime import date, datetime
from io import BytesIO
import csv
import json
import os
import re
from difflib import unified_diff

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from apps.contas.permissions import has_perm
from apps.contas.decorators import allow_admin

from .models import (
    Projeto,
    Documento,
    ArquivoDocumento,
    DocumentoVersao,
    ResponsavelDisciplina,
    WorkflowEtapa,
    DocumentoWorkflowStatus,
    LogAuditoria,
    ProjetoFinanceiro,
    DocumentoAprovacao,   # ★ AGORA IMPORTADO ★
    registrar_log,
)

from .utils_email import enviar_email

# =================================================================
# CONFIG GLOBAL
# =================================================================

from decimal import Decimal  # <-- garante precisão nos valores em dólar/reais

VALOR_MEDICAO_USD = Decimal("979.00")  # ainda usado na medição genérica
TAXA_CAMBIO_REAIS = Decimal("5.7642")  # taxa fixa que você passou

# ============================
# 🎯 CONTRATO TP25 – BASICO
# ============================
VALOR_TP25_BASICO_TOTAL_USD = Decimal("563901.21")
VALOR_TP25_BASICO_APROVADO_USD = Decimal("451120.97")
VALOR_TP25_BASICO_ASBUILT_USD = Decimal("112780.24")


ETAPAS_WORKFLOW = [
    "Revisão Interna – Disciplina",
    "Aprovação Técnica – Coordenador",
    "Envio ao Cliente",
    "Aprovação do Cliente",
    "Emissão Final",
]


# =================================================================
# FUNÇÕES AUXILIARES
# =================================================================

def normalizar_revisao(rev_bruto):
    """Normaliza revisões e valida formatos aceitos."""
    if rev_bruto is None:
        return ""

    rev = str(rev_bruto).strip().upper()

    if rev == "":
        return ""
    if rev == "0":
        return "0"
    if rev.isdigit():
        return None
    if any(ch.isdigit() for ch in rev):
        return None

    if rev.startswith("R") and rev[1:].isalpha():
        return rev[1:]

    return rev if rev.isalpha() else None


def highlight_text(texto, termo):
    """Destaca termo com <mark> (case-insensitive)."""
    if not texto or not termo:
        return texto or ""

    regex = re.compile(re.escape(termo), re.IGNORECASE)

    def repl(match):
        return (
            "<mark style='background:#ffeb3b; padding:2px; border-radius:4px;'>"
            f"{match.group(0)}</mark>"
        )

    return regex.sub(repl, texto)


def usuario_em_grupos(user, grupos):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=grupos).exists()


def pode_avancar_etapa(user, documento: Documento):
    """Permissão de avanço baseada em grupos + etapa atual."""
    if user.is_superuser:
        return True

    doc_control = usuario_em_grupos(
        user, ["Doc-Control", "DOC CONTROL", "DOC_CONTROL"]
    )
    etapa = documento.etapa_atual or ETAPAS_WORKFLOW[0]

    if etapa == "Revisão Interna – Disciplina":
        return doc_control or user.groups.filter(name__startswith="REVISORES_").exists()

    if etapa == "Aprovação Técnica – Coordenador":
        return doc_control or user.groups.filter(name__startswith="COORD_").exists()

    if etapa == "Envio ao Cliente":
        return doc_control

    if etapa == "Aprovação do Cliente":
        return doc_control or usuario_em_grupos(user, ["CLIENTE"])

    if etapa == "Emissão Final":
        return doc_control or user.groups.filter(name__startswith="COORD_").exists()

    return False


def pode_retornar_etapa(user, documento: Documento):
    if user.is_superuser:
        return True

    doc_control = usuario_em_grupos(
        user, ["Doc-Control", "DOC_CONTROL", "DOC CONTROL"]
    )
    coord = user.groups.filter(name__startswith="COORD_").exists()
    return doc_control or coord


def proxima_etapa(documento: Documento):
    etapa = documento.etapa_atual or ETAPAS_WORKFLOW[0]
    try:
        idx = ETAPAS_WORKFLOW.index(etapa)
    except ValueError:
        return None
    if idx + 1 < len(ETAPAS_WORKFLOW):
        return ETAPAS_WORKFLOW[idx + 1]
    return None


def etapa_anterior(documento: Documento):
    etapa = documento.etapa_atual or ETAPAS_WORKFLOW[0]
    try:
        idx = ETAPAS_WORKFLOW.index(etapa)
    except ValueError:
        return None
    if idx > 0:
        return ETAPAS_WORKFLOW[idx - 1]
    return None


def registrar_workflow(documento: Documento, etapa: str, status: str, request, observacao: str = ""):
    """
    Registra movimentação do workflow no log sem depender de DocumentoAprovacao.
    Mantém histórico enterprise para auditoria e dashboard.
    """

    etapa_obj = WorkflowEtapa.objects.filter(nome=etapa).first()
    if not etapa_obj:
        return  # Segurança — não quebra se etapa não existir

    usuario = request.user if request.user.is_authenticated else None

    descricao = (
        f"[WORKFLOW] Documento {documento.codigo} → {etapa} | "
        f"Status: {status} | {observacao or ''}"
    )

    # 🔥 Tudo passa pelo log enterprise centralizado
    registrar_log(
        usuario,
        documento,
        f"Workflow: {status}",
        descricao
    )



def notificar_evento_documento(documento: Documento, tipo_evento: str):
    """Centraliza notificações por e-mail (simples)."""
    assunto = ""
    mensagem = ""

    if tipo_evento == "envio_revisao":
        assunto = f"Documento em Revisão: {documento.codigo}"
        mensagem = (
            f"O documento {documento.codigo} (Rev {documento.revisao}) "
            "foi enviado para revisão."
        )
    elif tipo_evento == "aprovacao":
        assunto = f"Documento Aprovado: {documento.codigo}"
        mensagem = (
            f"O documento {documento.codigo} (Rev {documento.revisao}) "
            "foi aprovado."
        )
    elif tipo_evento == "emissao":
        assunto = f"Documento Emitido: {documento.codigo}"
        mensagem = (
            f"O documento {documento.codigo} (Rev {documento.revisao}) "
            f"foi emitido em {documento.data_emissao_grdt}."
        )
    elif tipo_evento == "cancelamento":
        assunto = f"Documento Cancelado: {documento.codigo}"
        mensagem = (
            f"O documento {documento.codigo} (Rev {documento.revisao}) "
            "foi cancelado."
        )
    else:
        return

    try:
        enviar_email(
            assunto=assunto,
            mensagem=mensagem,
            destinatarios=["seu.email@empresa.com"],
        )
    except Exception:
        # Nunca quebrar o fluxo por causa de e-mail
        pass


def mover_para_lixeira(documento: Documento, request, etapa: str = "Exclusão", motivo: str = ""):
    """Soft delete."""
    if not documento.ativo:
        return False

    documento.ativo = False
    documento.deletado_em = timezone.now()
    documento.deletado_por = request.user.username
    documento.motivo_exclusao = motivo or ""
    documento.save(update_fields=["ativo", "deletado_em", "deletado_por", "motivo_exclusao"])

    registrar_workflow(documento, etapa, "Excluído", request, motivo)
    return True


def restaurar_da_lixeira(documento: Documento, request):
    documento.ativo = True
    documento.deletado_em = None
    documento.deletado_por = None
    documento.motivo_exclusao = ""
    documento.save(update_fields=["ativo", "deletado_em", "deletado_por", "motivo_exclusao"])

    registrar_workflow(documento, "Restauração", "Restaurado", request)

def montar_descricao_log(usuario, documento, acao_verbosa: str):
    """
    Gera uma descrição padrão e detalhada para o LogAuditoria,
    no formato enterprise que combinamos.
    """
    from django.utils import timezone

    user = usuario.username if usuario and usuario.is_authenticated else "Sistema"
    cod = documento.codigo or "Sem código"
    rev = (documento.revisao or "").strip() or "-"
    projeto_nome = (
        documento.projeto.nome
        if getattr(documento, "projeto", None)
        else "Projeto TP25 - NAVIOS HANDY CLASSE 80"
    )
    status_doc = documento.status_documento or "Sem status doc"
    status_emis = documento.status_emissao or "Sem status emissão"
    agora = timezone.localtime(timezone.now())

    return (
        f"Usuário {user} {acao_verbosa} o documento {cod} (Rev {rev}) "
        f"no {projeto_nome} às {agora:%d/%m/%Y %H:%M} – "
        f"status: {status_doc} / {status_emis}"
    )


# =================================================================
# DASHBOARD SIMPLES
# =================================================================

@login_required
def dashboard(request):
    docs = Documento.objects.filter(ativo=True)

    total_docs = docs.count()
    total_aprovados = docs.filter(status_documento="Aprovado").count()
    total_em_revisao = docs.filter(status_documento="Em Revisão").count()
    total_emitidos = docs.filter(status_emissao="Emitido").count()
    total_nao_recebidos = docs.filter(status_emissao="Não Recebido").count()

    valor_emitidos_usd = total_emitidos * VALOR_MEDICAO_USD
    valor_nao_rec_usd = total_nao_recebidos * VALOR_MEDICAO_USD
    valor_total_usd = valor_emitidos_usd + valor_nao_rec_usd

    return render(
        request,
        "documentos/dashboard.html",
        {
            "total_docs": total_docs,
            "total_aprovados": total_aprovados,
            "total_em_revisao": total_em_revisao,
            "total_emitidos": total_emitidos,
            "total_nao_recebidos": total_nao_recebidos,
            "valor_total_usd": f"{valor_total_usd:,.2f}",
        },
    )


# =================================================================
# DASHBOARD ENTERPRISE
# =================================================================

@login_required
@has_perm("sistema.dashboard_enterprise")
def dashboard_enterprise(request):
    docs = Documento.objects.filter(ativo=True).select_related("projeto")

    filtros = {
        "projeto": request.GET.get("projeto") or "",
        "disciplina": request.GET.get("disciplina") or "",
        "tipo_doc": request.GET.get("tipo_doc") or "",
        "status_documento": request.GET.get("status_ldp") or "",
        "status_emissao": request.GET.get("status_emissao") or "",
        "dt_ini": request.GET.get("dt_ini") or "",
        "dt_fim": request.GET.get("dt_fim") or "",
    }

    if filtros["projeto"]:
        docs = docs.filter(projeto__nome__icontains=filtros["projeto"])

    if filtros["disciplina"]:
        docs = docs.filter(disciplina=filtros["disciplina"])

    if filtros["tipo_doc"]:
        docs = docs.filter(tipo_doc=filtros["tipo_doc"])

    if filtros["status_documento"]:
        docs = docs.filter(status_documento=filtros["status_documento"])

    if filtros["status_emissao"]:
        docs = docs.filter(status_emissao=filtros["status_emissao"])

    if filtros["dt_ini"]:
        dt_ini = datetime.strptime(filtros["dt_ini"], "%Y-%m-%d").date()
        docs = docs.filter(data_emissao_grdt__gte=dt_ini)

    if filtros["dt_fim"]:
        dt_fim = datetime.strptime(filtros["dt_fim"], "%Y-%m-%d").date()
        docs = docs.filter(data_emissao_grdt__lte=dt_fim)

    total_docs = docs.count()
    total_emitidos = docs.filter(status_emissao="Emitido").count()
    total_nao_recebidos = docs.filter(status_emissao="Não Recebido").count()
    total_em_revisao = docs.filter(status_documento="Em Revisão").count()
    total_aprovados = docs.filter(status_documento="Aprovado").count()

    valor_emitidos_usd = total_emitidos * VALOR_MEDICAO_USD
    valor_nao_rec_usd = total_nao_recebidos * VALOR_MEDICAO_USD
    valor_total_usd = valor_emitidos_usd + valor_nao_rec_usd

    valor_emitidos_brl = valor_emitidos_usd * TAXA_CAMBIO_REAIS
    valor_nao_rec_brl = valor_nao_rec_usd * TAXA_CAMBIO_REAIS
    valor_total_brl = valor_emitidos_brl + valor_nao_rec_brl

    por_disc = docs.values("disciplina").annotate(qtd=Count("id"))
    disc_labels = [d["disciplina"] or "Sem Disciplina" for d in por_disc]
    disc_data = [d["qtd"] for d in por_disc]

    por_status = docs.values("status_documento").annotate(qtd=Count("id"))
    status_labels = [d["status_documento"] or "Sem Status" for d in por_status]
    status_data = [d["qtd"] for d in por_status]

    med_raw = docs.values("tipo_doc").annotate(
        total=Count("id"),
        emitidos=Count("id", filter=Q(status_emissao="Emitido")),
        nr=Count("id", filter=Q(status_emissao="Não Recebido")),
    )

    medicao_linhas = []
    total_usd = 0
    total_brl = 0

    for m in med_raw:
        tipo = m["tipo_doc"] or "Sem Tipo"
        emit = m["emitidos"]
        nr = m["nr"]

        v_emit_usd = emit * VALOR_MEDICAO_USD
        v_nr_usd = nr * VALOR_MEDICAO_USD
        v_emit_brl = v_emit_usd * TAXA_CAMBIO_REAIS
        v_nr_brl = v_nr_usd * TAXA_CAMBIO_REAIS

        total_usd += v_emit_usd + v_nr_usd
        total_brl += v_emit_brl + v_nr_brl

        medicao_linhas.append(
            {
                "tipo_doc": tipo,
                "total": m["total"],
                "emitidos": emit,
                "nao_recebidos": nr,
                "valor_emitidos_usd": f"{v_emit_usd:,.2f}",
                "valor_nr_usd": f"{v_nr_usd:,.2f}",
                "valor_emitidos_brl": f"{v_emit_brl:,.2f}",
                "valor_nr_brl": f"{v_nr_brl:,.2f}",
            }
        )

    medicao_totais = {
        "total_docs": sum(m["total"] for m in med_raw),
        "total_usd": f"{total_usd:,.2f}",
        "total_brl": f"{total_brl:,.2f}",
    }

    base_qs = Documento.objects.filter(ativo=True).select_related("projeto")

    return render(
        request,
        "documentos/dashboard_enterprise.html",
        {
            "filtros": filtros,
            "total_docs": total_docs,
            "total_emitidos": total_emitidos,
            "total_nao_recebidos": total_nao_recebidos,
            "total_em_revisao": total_em_revisao,
            "total_aprovados": total_aprovados,
            "valor_emitidos_usd": f"{valor_emitidos_usd:,.2f}",
            "valor_emitidos_brl": f"{valor_emitidos_brl:,.2f}",
            "valor_nao_rec_usd": f"{valor_nao_rec_usd:,.2f}",
            "valor_nao_rec_brl": f"{valor_nao_rec_brl:,.2f}",
            "valor_total_usd": f"{valor_total_usd:,.2f}",
            "valor_total_brl": f"{valor_total_brl:,.2f}",
            "disc_labels": json.dumps(disc_labels),
            "disc_data": json.dumps(disc_data),
            "status_labels": json.dumps(status_labels),
            "status_data": json.dumps(status_data),
            "medicao_linhas": medicao_linhas,
            "medicao_totais": medicao_totais,
            "lista_projetos": base_qs.values_list("projeto__nome", flat=True).distinct(),
            "lista_disciplinas": base_qs.values_list("disciplina", flat=True).distinct(),
            "lista_tipos": base_qs.values_list("tipo_doc", flat=True).distinct(),
            "lista_status_ldp": base_qs.values_list("status_documento", flat=True).distinct(),
            "lista_status_emissao": base_qs.values_list("status_emissao", flat=True).distinct(),
        },
    )


# =================================================================
# PAINEL DE WORKFLOW
# =================================================================

@login_required
def painel_workflow(request):
    docs_base = Documento.objects.filter(ativo=True).select_related("projeto")

    etapa_filtro = request.GET.get("etapa") or ""
    disciplina_filtro = request.GET.get("disciplina") or ""
    projeto_filtro = request.GET.get("projeto") or ""
    status_ldp_filtro = request.GET.get("status_ldp") or ""
    status_emissao_filtro = request.GET.get("status_emissao") or ""

    docs = docs_base

    if etapa_filtro and etapa_filtro != "Sem etapa":
        docs = docs.filter(etapa_atual=etapa_filtro)
    elif etapa_filtro == "Sem etapa":
        docs = docs.filter(Q(etapa_atual__isnull=True) | Q(etapa_atual=""))

    if disciplina_filtro:
        docs = docs.filter(disciplina=disciplina_filtro)

    if projeto_filtro:
        docs = docs.filter(projeto__nome=projeto_filtro)

    if status_ldp_filtro:
        docs = docs.filter(status_documento=status_ldp_filtro)

    if status_emissao_filtro:
        docs = docs.filter(status_emissao=status_emissao_filtro)

    order = request.GET.get("order") or ""
    direction = request.GET.get("direction") or "asc"

    valid_fields = {
        "codigo": "codigo",
        "titulo": "titulo",
        "disciplina": "disciplina",
        "projeto": "projeto__nome",
        "etapa_atual": "etapa_atual",
        "status_ldp": "status_documento",
        "status_emissao": "status_emissao",
        "revisao": "revisao",
    }

    if order in valid_fields:
        field = valid_fields[order]
        if direction == "desc":
            field = "-" + field
        docs = docs.order_by(field)
    else:
        docs = docs.order_by("etapa_atual", "codigo")

    disciplinas = (
        docs_base.exclude(disciplina__isnull=True)
        .exclude(disciplina__exact="")
        .values_list("disciplina", flat=True)
        .distinct()
        .order_by("disciplina")
    )

    projetos = (
        docs_base.filter(projeto__isnull=False)
        .values_list("projeto__nome", flat=True)
        .distinct()
        .order_by("projeto__nome")
    )

    status_ldp_lista = (
        docs_base.exclude(status_documento__isnull=True)
        .exclude(status_documento__exact="")
        .values_list("status_documento", flat=True)
        .distinct()
        .order_by("status_documento")
    )

    status_emissao_lista = (
        docs_base.exclude(status_emissao__isnull=True)
        .exclude(status_emissao__exact="")
        .values_list("status_emissao", flat=True)
        .distinct()
        .order_by("status_emissao")
    )

    per_page = request.GET.get("per_page") or 25
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 25

    total_filtrados = docs.count()

    paginator = Paginator(docs, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    documentos = page_obj

    por_etapa_qs = docs_base.values("etapa_atual").annotate(qtd=Count("id"))
    mapa_etapas = {row["etapa_atual"] or "Sem etapa": row["qtd"] for row in por_etapa_qs}

    etapas_kpi = []
    for nome in ETAPAS_WORKFLOW:
        url = (
            f"?etapa={nome}"
            f"&disciplina={disciplina_filtro}"
            f"&projeto={projeto_filtro}"
            f"&status_ldp={status_ldp_filtro}"
            f"&status_emissao={status_emissao_filtro}"
            f"&order={order}"
            f"&direction={direction}"
            f"&per_page={per_page}"
        )
        etapas_kpi.append(
            {
                "nome": nome,
                "qtd": mapa_etapas.get(nome, 0),
                "url": url,
            }
        )

    url_sem_etapa = (
        f"?etapa=Sem%20etapa"
        f"&disciplina={disciplina_filtro}"
        f"&projeto={projeto_filtro}"
        f"&status_ldp={status_ldp_filtro}"
        f"&status_emissao={status_emissao_filtro}"
        f"&order={order}"
        f"&direction={direction}"
        f"&per_page={per_page}"
    )

    etapas_kpi.append(
        {
            "nome": "Sem etapa definida",
            "qtd": mapa_etapas.get("Sem etapa", mapa_etapas.get(None, 0)),
            "url": url_sem_etapa,
        }
    )

    ultimos_movimentos = (
        DocumentoAprovacao.objects
        .select_related("documento", "etapa", "usuario")
        .order_by("-data")[:20]
    )

    return render(
        request,
        "documentos/painel_workflow.html",
        {
            "documentos": documentos,
            "page_obj": page_obj,
            "per_page": per_page,
            "total_docs": total_filtrados,
            "order": order,
            "direction": direction,
            "etapas_kpi": etapas_kpi,
            "etapas_workflow": ETAPAS_WORKFLOW,
            "etapa_filtro": etapa_filtro,
            "disciplinas": disciplinas,
            "disciplina_filtro": disciplina_filtro,
            "projetos": projetos,
            "projeto_filtro": projeto_filtro,
            "status_ldp_lista": status_ldp_lista,
            "status_ldp_filtro": status_ldp_filtro,
            "status_emissao_lista": status_emissao_lista,
            "status_emissao_filtro": status_emissao_filtro,
            "ultimos_movimentos": ultimos_movimentos,
        },
    )


# =================================================================
# EXPORTAR EXCEL DO PAINEL
# =================================================================

@login_required
def painel_workflow_exportar_excel(request):
    etapa = request.GET.get("etapa") or ""
    disciplina = request.GET.get("disciplina") or ""
    projeto = request.GET.get("projeto") or ""
    status_ldp = request.GET.get("status_ldp") or ""
    status_emissao = request.GET.get("status_emissao") or ""

    docs = Documento.objects.filter(ativo=True).select_related("projeto")

    if etapa and etapa != "Sem etapa":
        docs = docs.filter(etapa_atual=etapa)
    elif etapa == "Sem etapa":
        docs = docs.filter(Q(etapa_atual__isnull=True) | Q(etapa_atual=""))

    if disciplina:
        docs = docs.filter(disciplina=disciplina)
    if projeto:
        docs = docs.filter(projeto__nome=projeto)
    if status_ldp:
        docs = docs.filter(status_documento=status_ldp)
    if status_emissao:
        docs = docs.filter(status_emissao=status_emissao)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Workflow GED"

    headers = [
        "Código",
        "Revisão",
        "Título",
        "Disciplina",
        "Projeto",
        "Etapa Atual",
        "Status Documento",
        "Status Emissão",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="0052A2", end_color="0052A2", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in ws[1]:
        col.fill = header_fill
        col.font = Font(color="FFFFFF", bold=True)
        col.border = border
        col.alignment = Alignment(horizontal="center")

    for d in docs:
        ws.append(
            [
                d.codigo or "",
                d.revisao or "",
                d.titulo or "",
                d.disciplina or "",
                d.projeto.nome if d.projeto else "",
                d.etapa_atual or "",
                d.status_documento or "",
                d.status_emissao or "",
            ]
        )

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"workflow_{date.today().isoformat()}.xlsx"

    return HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =================================================================
# LISTAR DOCUMENTOS
# =================================================================

@login_required
def listar_documentos(request):
    docs = Documento.objects.filter(ativo=True).select_related("projeto").order_by("-criado_em")

    projeto = request.GET.get("projeto") or ""
    disciplina = request.GET.get("disciplina") or ""
    status_ldp = request.GET.get("status_ldp") or ""
    status_emissao = request.GET.get("status_emissao") or ""
    busca = request.GET.get("busca") or ""

    if projeto:
        docs = docs.filter(projeto__nome__icontains=projeto)
    if disciplina:
        docs = docs.filter(disciplina=disciplina)
    if status_ldp:
        docs = docs.filter(status_documento=status_ldp)
    if status_emissao:
        docs = docs.filter(status_emissao=status_emissao)
    if busca:
        docs = docs.filter(Q(titulo__icontains=busca) | Q(codigo__icontains=busca))

    base_qs = Documento.objects.filter(ativo=True).select_related("projeto")

    filtros_ativos = {
        "projeto": projeto,
        "disciplina": disciplina,
        "status_ldp": status_ldp,
        "status_emissao": status_emissao,
        "busca": busca,
    }

    ha_filtros = any(filtros_ativos.values())

    return render(
        request,
        "documentos/listar.html",
        {
            "documentos": docs,
            "projetos": base_qs.values_list("projeto__nome", flat=True).distinct(),
            "disciplinas": base_qs.values_list("disciplina", flat=True).distinct(),
            "status_ldp_list": base_qs.values_list("status_documento", flat=True).distinct(),
            "status_emissao_list": base_qs.values_list("status_emissao", flat=True).distinct(),
            "filtros_ativos": filtros_ativos,
            "ha_filtros": ha_filtros,
        },
    )
# ==============================
# Revisões permitidas no sistema
# ==============================
REVISOES_VALIDAS = ["0"] + [chr(i) for i in range(ord("A"), ord("Z")+1)]

# =================================================================
# DETALHE DO DOCUMENTO
# =================================================================

@login_required
def detalhes_documento(request, documento_id):
    documento = get_object_or_404(
        Documento.objects.select_related("projeto"), id=documento_id
    )

    anexos = list(documento.arquivos.all())
    for a in anexos:
        try:
            a.url_absoluta = request.build_absolute_uri(a.arquivo.url)
        except Exception:
            a.url_absoluta = a.arquivo.url

    versoes_qs = documento.versoes.all().order_by("-criado_em")
    versao_atual = versoes_qs.first()

    rev_atual = str(documento.revisao or "").strip().upper()
    if rev_atual not in REVISOES_VALIDAS:
        rev_atual = "0"

    try:
        idx = REVISOES_VALIDAS.index(rev_atual)
        proxima_revisao = REVISOES_VALIDAS[idx + 1] if idx + 1 < len(REVISOES_VALIDAS) else rev_atual
    except ValueError:
        proxima_revisao = "A"

    try:
        workflow_status = documento.workflow_status
    except DocumentoWorkflowStatus.DoesNotExist:
        workflow_status = None

    pode = pode_avancar_etapa(request.user, documento)
    prox = proxima_etapa(documento)
    pode_ret = pode_retornar_etapa(request.user, documento)
    etapa_prev = etapa_anterior(documento)

    historico = (
        documento.aprovacoes.select_related("etapa", "usuario").order_by("-data")
    )

    return render(
        request,
        "documentos/detalhes.html",
        {
            "documento": documento,
            "historico_workflow": historico,
            "anexos": anexos,
            "versoes": versoes_qs,
            "versao_atual": versao_atual,
            "proxima_revisao": proxima_revisao,
            "workflow_status": workflow_status,
            "pode_avancar_etapa": pode,
            "proxima_etapa": prox,
            "pode_retornar_etapa": pode_ret,
            "etapa_anterior": etapa_prev,
        },
    )

# =================================================================
# HISTÓRICO POR CÓDIGO
# =================================================================

@login_required
def historico(request, codigo):
    documentos = Documento.objects.filter(codigo=codigo).order_by("-revisao")
    versoes = (
        DocumentoVersao.objects.filter(documento__codigo=codigo)
        .select_related("documento", "criado_por")
        .order_by("-criado_em")
    )
    return render(
        request,
        "documentos/historico.html",
        {"documentos": documentos, "versoes": versoes, "codigo": codigo},
    )


# =================================================================
# AVANÇAR / RETORNAR ETAPA (workflow textual)
# =================================================================

@login_required
def enviar_proxima_etapa(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    if not pode_avancar_etapa(request.user, documento):
        messages.error(request, "Você não tem permissão para avançar esta etapa.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    prox = proxima_etapa(documento)
    if not prox:
        messages.warning(
            request, "Este documento já está na última etapa do workflow."
        )
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    documento.etapa_atual = prox
    documento.save(update_fields=["etapa_atual"])

    registrar_workflow(
        documento,
        prox,
        "Avançado",
        request,
        observacao=f"Documento avançado para: {prox}",
    )

    messages.success(request, f"Documento avançado para a próxima etapa: {prox}")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)


@login_required
def retornar_etapa(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    if not pode_retornar_etapa(request.user, documento):
        messages.error(request, "Você não tem permissão para retornar esta etapa.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    etapa_prev = etapa_anterior(documento)
    if not etapa_prev:
        messages.warning(request, "Este documento já está na primeira etapa.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    documento.etapa_atual = etapa_prev
    documento.save(update_fields=["etapa_atual"])

    registrar_workflow(
        documento,
        etapa_prev,
        "Retornado",
        request,
        observacao=f"Etapa retornada para: {etapa_prev}",
    )

    messages.success(request, f"Etapa retornada para: {etapa_prev}")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)


# =================================================================
# NOVA VERSÃO (upload simples de arquivo)
# =================================================================

@login_required
def nova_versao(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    rev_atual = str(documento.revisao or "").strip().upper()
    if rev_atual not in REVISOES_VALIDAS:
        rev_atual = "0"
    try:
        idx = REVISOES_VALIDAS.index(rev_atual)
        proxima_revisao = (
            REVISOES_VALIDAS[idx + 1] if idx + 1 < len(REVISOES_VALIDAS) else rev_atual
        )
    except ValueError:
        proxima_revisao = "A"

    if request.method == "POST":
        numero_revisao = request.POST.get("numero_revisao", "").strip().upper()
        observacao = request.POST.get("observacao", "").strip()
        arquivo = request.FILES.get("arquivo")

        if not arquivo:
            messages.error(request, "Selecione um arquivo.")
            return redirect("documentos:detalhes_documento", documento_id=documento.id)

        if not numero_revisao:
            numero_revisao = proxima_revisao

        if numero_revisao not in REVISOES_VALIDAS:
            messages.error(request, "Revisão inválida.")
            return redirect("documentos:detalhes_documento", documento_id=documento.id)

        DocumentoVersao.objects.create(
            documento=documento,
            numero_revisao=numero_revisao,
            arquivo=arquivo,
            observacao=observacao,
            criado_por=request.user,
            status_revisao="REVISAO",
        )

        documento.revisao = numero_revisao
        documento.save(update_fields=["revisao"])

        registrar_workflow(
            documento,
            etapa="Nova Versão",
            status="Criada",
            request=request,
            observacao=f"Versão {numero_revisao} adicionada.",
        )

        messages.success(request, "Nova versão adicionada com sucesso!")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    return redirect("documentos:detalhes_documento", documento_id=documento.id)


# =================================================================
# UPLOAD DOCUMENTO (CRIAR)
# =================================================================

@login_required
@has_perm("documento.criar")
def upload_documento(request):
    if request.method == "POST":
        revisao = normalizar_revisao(request.POST.get("revisao"))
        if revisao is None:
            messages.error(request, "Revisão inválida!")
            return redirect("documentos:upload_documento")

        projeto_id = request.POST.get("projeto")
        projeto = None
        if projeto_id:
            try:
                projeto = Projeto.objects.get(id=projeto_id)
            except Projeto.DoesNotExist:
                messages.error(request, "Projeto inválido.")
                return redirect("documentos:upload_documento")
        else:
            messages.error(request, "Selecione um projeto.")
            return redirect("documentos:upload_documento")

        doc = Documento.objects.create(
            projeto=projeto,
            titulo=request.POST.get("titulo"),
            codigo=request.POST.get("codigo"),
            revisao=revisao,
            disciplina=request.POST.get("disciplina"),
            tipo_doc=request.POST.get("tipo_doc"),
            fase=request.POST.get("fase") or "",
        )

        arquivos = request.FILES.getlist("arquivos")
        for arq in arquivos:
            ArquivoDocumento.objects.create(
                documento=doc,
                arquivo=arq,
                nome_original=arq.name,
                tipo=arq.name.split(".")[-1].lower(),
            )

        registrar_workflow(doc, "Criação", "Criado", request)

        # =============== 🔥 AUDITORIA ENTERPRISE 🔥 ===============
        descricao = montar_descricao_log(
            request.user,
            doc,
            "criou"   # ação no texto do log
        )
        registrar_log(
            request.user,
            doc,
            "Criação de Documento",
            descricao
        )
        # ==========================================================

        messages.success(request, "Documento criado com sucesso!")
        return redirect("documentos:listar_documentos")

    projetos = Projeto.objects.filter(ativo=True).order_by("nome")
    return render(request, "documentos/upload.html", {"projetos": projetos})



# =================================================================
# EDITAR DOCUMENTO
# =================================================================

@login_required
@has_perm("documento.editar")
def editar_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    if request.method == "POST":
        revisao = normalizar_revisao(request.POST.get("revisao"))
        if revisao is None:
            messages.error(request, "Revisão inválida.")
            return redirect("documentos:editar_documento", documento_id=documento.id)

        projeto_id = request.POST.get("projeto")
        if projeto_id:
            try:
                projeto = Projeto.objects.get(id=projeto_id)
                documento.projeto = projeto
            except Projeto.DoesNotExist:
                messages.error(request, "Projeto inválido.")
                return redirect("documentos:editar_documento", documento_id=documento.id)

        documento.fase = request.POST.get("fase") or ""
        documento.tipo_doc = request.POST.get("tipo_doc") or ""
        documento.codigo = request.POST.get("codigo") or ""
        documento.disciplina = request.POST.get("disciplina") or ""
        documento.titulo = request.POST.get("titulo") or ""
        documento.status_documento = request.POST.get("status_documento") or ""
        documento.status_emissao = request.POST.get("status_emissao") or ""
        documento.grdt_cliente = request.POST.get("grdt_cliente") or ""
        documento.resposta_cliente = request.POST.get("resposta_cliente") or ""
        documento.ged_interna = request.POST.get("ged_interna") or ""
        documento.revisao = revisao

        data_emissao_grdt = request.POST.get("data_emissao_grdt") or ""
        if data_emissao_grdt:
            try:
                documento.data_emissao_grdt = datetime.strptime(
                    data_emissao_grdt, "%Y-%m-%d"
                ).date()
            except ValueError:
                messages.error(request, "Data de emissão GRDT inválida.")
                return redirect("documentos:editar_documento", documento_id=documento.id)
        else:
            documento.data_emissao_grdt = None

        documento.save()

        messages.success(request, "Documento atualizado.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    projetos = Projeto.objects.filter(ativo=True).order_by("nome")
    return render(
        request,
        "documentos/editar.html",
        {
            "documento": documento,
            "projetos": projetos,
            "REVISOES_VALIDAS": REVISOES_VALIDAS,
        },
    )


# =================================================================
# NOVA REVISÃO (WORKFLOW)
# =================================================================

@login_required
@has_perm("documento.revisar")
def nova_revisao(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    idx = (
        REVISOES_VALIDAS.index(documento.revisao)
        if documento.revisao in REVISOES_VALIDAS
        else -1
    )
    nova_rev = (
        REVISOES_VALIDAS[idx + 1]
        if idx + 1 < len(REVISOES_VALIDAS)
        else "A"
    )

    if request.method == "POST":
        arquivo = request.FILES.get("arquivo")
        observacao = request.POST.get("observacao", "")

        if not arquivo:
            messages.error(request, "Envie o arquivo da nova revisão.")
            return redirect("documentos:detalhes_documento", documento_id=documento.id)

        DocumentoVersao.objects.create(
            documento=documento,
            numero_revisao=nova_rev,
            arquivo=arquivo,
            criado_por=request.user,
            observacao=observacao,
            status_revisao="REVISAO",
        )

        documento.revisao = nova_rev
        documento.status_documento = "Em Revisão"
        documento.save(update_fields=["revisao", "status_documento"])

        registrar_workflow(documento, "Nova Revisão", "Criado", request)
        notificar_evento_documento(documento, "envio_revisao")

        messages.success(request, f"Revisão {nova_rev} criada com sucesso!")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    return render(
        request,
        "documentos/nova_revisao.html",
        {"documento": documento, "proxima_revisao": nova_rev},
    )


# =================================================================
# UPLOAD DE ANEXOS
# =================================================================

@login_required
def adicionar_arquivos(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    if request.method == "POST":
        arquivos = request.FILES.getlist("arquivos")

        for arq in arquivos:
            ArquivoDocumento.objects.create(
                documento=documento,
                arquivo=arq,
                nome_original=arq.name,
                tipo=arq.name.split(".")[-1].lower(),
            )

        registrar_workflow(
            documento,
            "Upload de anexos",
            "Arquivos adicionados",
            request,
            observacao=f"{len(arquivos)} arquivos enviados",
        )

        messages.success(request, "Arquivos enviados com sucesso!")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    return render(
        request,
        "documentos/adicionar_arquivos.html",
        {"documento": documento},
    )


# =================================================================
# EXCLUIR ARQUIVO INDIVIDUAL
# =================================================================

@login_required
def excluir_arquivo(request, arquivo_id):
    arq = get_object_or_404(ArquivoDocumento, id=arquivo_id)
    documento_id = arq.documento.id

    try:
        default_storage.delete(arq.arquivo.name)
    except Exception:
        pass

    arq.delete()

    messages.success(request, "Arquivo excluído com sucesso.")
    return redirect("documentos:detalhes_documento", documento_id=documento_id)


# =================================================================
# WORKFLOW DE EMISSÃO / APROVAÇÃO
# =================================================================

@login_required
def enviar_para_revisao(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    documento.status_documento = "Em Revisão"
    documento.save(update_fields=["status_documento"])

    registrar_workflow(documento, "Revisão Interna – Disciplina", "Enviado", request)
    notificar_evento_documento(documento, "envio_revisao")

    messages.success(request, "Documento enviado para revisão.")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)


@login_required
def aprovar_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    documento.status_documento = "Aprovado"
    documento.save(update_fields=["status_documento"])

    registrar_workflow(documento, "Aprovação Técnica – Coordenador", "Aprovado", request)
    notificar_evento_documento(documento, "aprovacao")

    messages.success(request, "Documento aprovado.")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)


@login_required
def emitir_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    documento.status_emissao = "Emitido"
    documento.data_emissao_grdt = date.today()
    documento.save(update_fields=["status_emissao", "data_emissao_grdt"])

    registrar_workflow(documento, "Emissão Final", "Emitido", request)
    notificar_evento_documento(documento, "emissao")

    messages.success(request, "Documento emitido.")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)


@login_required
def cancelar_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    documento.status_documento = "Cancelado"
    documento.status_emissao = "Cancelado"
    documento.save(update_fields=["status_documento", "status_emissao"])

    registrar_workflow(documento, "Emissão Final", "Cancelado", request)
    notificar_evento_documento(documento, "cancelamento")

    messages.success(request, "Documento cancelado.")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)


# =================================================================
# GERAR DIFF ENTRE DUAS REVISÕES
# =================================================================

@login_required
def gerar_diff(request, documento_id, revA, revB):
    documento = get_object_or_404(Documento, id=documento_id)

    versao_a = get_object_or_404(
        DocumentoVersao, documento=documento, numero_revisao=revA
    )
    versao_b = get_object_or_404(
        DocumentoVersao, documento=documento, numero_revisao=revB
    )

    def ler_arquivo_texto(file_field):
        try:
            path = file_field.path
        except Exception:
            return None
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read().splitlines()
        except Exception:
            return None

    linhas_a = ler_arquivo_texto(versao_a.arquivo)
    linhas_b = ler_arquivo_texto(versao_b.arquivo)

    diff_text = []
    if linhas_a is not None and linhas_b is not None:
        diff_text = list(
            unified_diff(
                linhas_a,
                linhas_b,
                fromfile=f"Rev {revA}",
                tofile=f"Rev {revB}",
                lineterm="",
            )
        )

    return render(
        request,
        "documentos/diff.html",
        {
            "documento": documento,
            "versao_a": versao_a,
            "versao_b": versao_b,
            "diff": diff_text,
        },
    )


# =================================================================
# IMPORTAÇÃO LDP (SIMPLIFICADA)
# =================================================================

@login_required
def importar_ldp(request):
    if request.method == "POST":
        arquivo = request.FILES.get("arquivo")
        if not arquivo:
            messages.error(request, "Selecione um arquivo CSV ou XLSX.")
            return redirect("documentos:importar_ldp")

        nome = arquivo.name.lower()
        total_linhas = 0

        # Aqui está uma implementação segura e genérica.
        # Você pode adaptar o mapeamento de colunas depois.
        try:
            if nome.endswith(".csv"):
                decoded = arquivo.read().decode("utf-8", errors="ignore").splitlines()
                reader = csv.reader(decoded, delimiter=";")
                for row in reader:
                    total_linhas += 1
                    # TODO: mapear colunas → Documento
            elif nome.endswith(".xlsx"):
                wb = openpyxl.load_workbook(arquivo)
                ws = wb.active
                for _ in ws.iter_rows(min_row=2):  # pula cabeçalho
                    total_linhas += 1
                    # TODO: mapear colunas → Documento
            else:
                messages.error(request, "Formato não suportado. Use CSV ou XLSX.")
                return redirect("documentos:importar_ldp")
        except Exception as e:
            messages.error(request, f"Erro ao processar arquivo: {e}")
            return redirect("documentos:importar_ldp")

        messages.success(
            request,
            f"Arquivo processado com sucesso ({total_linhas} linhas lidas). Ajuste a lógica de importação conforme necessário.",
        )
        return redirect("documentos:listar_documentos")

    return render(request, "documentos/importar_ldp.html")


# =================================================================
# MEDIÇÃO / EXPORTAÇÃO MEDIÇÃO
# =================================================================

def _calcular_medicao_queryset(docs_qs):
    med_raw = docs_qs.values("tipo_doc").annotate(
        total=Count("id"),
        emitidos=Count("id", filter=Q(status_emissao="Emitido")),
        nr=Count("id", filter=Q(status_emissao="Não Recebido")),
    )

    linhas = []
    total_usd = 0
    total_brl = 0

    for m in med_raw:
        tipo = m["tipo_doc"] or "Sem Tipo"
        emit = m["emitidos"]
        nr = m["nr"]

        v_emit_usd = emit * VALOR_MEDICAO_USD
        v_nr_usd = nr * VALOR_MEDICAO_USD
        v_emit_brl = v_emit_usd * TAXA_CAMBIO_REAIS
        v_nr_brl = v_nr_usd * TAXA_CAMBIO_REAIS

        total_usd += v_emit_usd + v_nr_usd
        total_brl += v_emit_brl + v_nr_brl

        linhas.append(
            {
                "tipo_doc": tipo,
                "total": m["total"],
                "emitidos": emit,
                "nao_recebidos": nr,
                "valor_emitidos_usd": f"{v_emit_usd:,.2f}",
                "valor_nr_usd": f"{v_nr_usd:,.2f}",
                "valor_emitidos_brl": f"{v_emit_brl:,.2f}",
                "valor_nr_brl": f"{v_nr_brl:,.2f}",
            }
        )

    totais = {
        "total_docs": sum(m["total"] for m in med_raw),
        "total_usd": f"{total_usd:,.2f}",
        "total_brl": f"{total_brl:,.2f}",
    }

    return linhas, totais


@login_required
def medicao(request):
    docs = Documento.objects.filter(ativo=True)

    projeto = request.GET.get("projeto") or ""
    if projeto:
        docs = docs.filter(projeto__nome__icontains=projeto)

    linhas, totais = _calcular_medicao_queryset(docs)

    base_qs = Documento.objects.filter(ativo=True).select_related("projeto")

    return render(
        request,
        "documentos/medicao.html",
        {
            "linhas": linhas,
            "totais": totais,
            "projetos": base_qs.values_list("projeto__nome", flat=True).distinct(),
            "projeto_selecionado": projeto,
        },
    )


@login_required
def exportar_medicao_excel(request):
    docs = Documento.objects.filter(ativo=True)

    projeto = request.GET.get("projeto") or ""
    if projeto:
        docs = docs.filter(projeto__nome__icontains=projeto)

    linhas, totais = _calcular_medicao_queryset(docs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medição GED"

    headers = [
        "Tipo de Documento",
        "Total",
        "Emitidos",
        "Não Recebidos",
        "Valor Emitidos (USD)",
        "Valor Não Recebidos (USD)",
        "Valor Emitidos (BRL)",
        "Valor Não Recebidos (BRL)",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="0052A2", end_color="0052A2", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in ws[1]:
        col.fill = header_fill
        col.font = Font(color="FFFFFF", bold=True)
        col.border = border
        col.alignment = Alignment(horizontal="center")

    for linha in linhas:
        ws.append(
            [
                linha["tipo_doc"],
                linha["total"],
                linha["emitidos"],
                linha["nao_recebidos"],
                linha["valor_emitidos_usd"],
                linha["valor_nr_usd"],
                linha["valor_emitidos_brl"],
                linha["valor_nr_brl"],
            ]
        )

    row_total = [
        "TOTAL",
        totais["total_docs"],
        "",
        "",
        totais["total_usd"],
        "",
        totais["total_brl"],
        "",
    ]
    ws.append([])
    ws.append(row_total)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"medicao_{date.today().isoformat()}.xlsx"

    return HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =================================================================
# EXCLUSÃO / LIXEIRA
# =================================================================

@login_required
def excluir_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)
    motivo = request.POST.get("motivo", "") if request.method == "POST" else ""
    mover_para_lixeira(documento, request, motivo=motivo)
    messages.success(request, "Documento movido para a lixeira.")
    return redirect("documentos:listar_documentos")


@login_required
def excluir_selecionados(request):
    if request.method != "POST":
        return redirect("documentos:listar_documentos")

    ids = request.POST.getlist("ids") or []
    count = 0
    for _id in ids:
        try:
            doc = Documento.objects.get(id=_id)
            if mover_para_lixeira(doc, request, motivo="Exclusão em lote"):
                count += 1
        except Documento.DoesNotExist:
            continue

    messages.success(request, f"{count} documento(s) movido(s) para a lixeira.")
    return redirect("documentos:listar_documentos")


@login_required
def lixeira(request):
    docs = Documento.objects.filter(ativo=False).order_by("-deletado_em")
    return render(request, "documentos/lixeira.html", {"documentos": docs})


@login_required
def restaurar_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)
    restaurar_da_lixeira(documento, request)
    messages.success(request, "Documento restaurado com sucesso.")
    return redirect("documentos:lixeira")


# =================================================================
# CONFIGURAÇÕES BÁSICAS
# =================================================================

@login_required
@allow_admin
def configuracoes(request):
    projetos = Projeto.objects.all().order_by("nome")
    resp_disc = ResponsavelDisciplina.objects.all().order_by("disciplina")
    etapas = WorkflowEtapa.objects.all().order_by("ordem")

    return render(
        request,
        "documentos/configuracoes.html",
        {
            "projetos": projetos,
            "responsaveis": resp_disc,
            "etapas": etapas,
        },
    )

# =================================================================
# CONFIGURAÇÕES UNIFICADAS (Usuário + Administração)
# =================================================================

@login_required
def configuracoes(request):
    from apps.contas.models import UserConfig  # Importa aqui para evitar ciclos

    # Carrega ou cria configuração do usuário
    config, _ = UserConfig.objects.get_or_create(user=request.user)

    # Salvar preferências pessoais
    if request.method == "POST" and "salvar_prefs" in request.POST:
        config.tema = request.POST.get("tema", config.tema)
        config.animacoes = request.POST.get("animacoes", "True") == "True"
        config.notificacoes_email = request.POST.get("notificacoes_email", "True") == "True"
        config.dashboard_expandido = request.POST.get("dashboard_expandido", "True") == "True"
        config.save()
        messages.success(request, "Preferências atualizadas com sucesso!")

    # Dados para parte administrativa (apenas admins/master)
    projetos = Projeto.objects.all().order_by("nome")
    resp_disc = ResponsavelDisciplina.objects.all().order_by("disciplina")
    etapas = WorkflowEtapa.objects.all().order_by("ordem")

    return render(
        request,
        "documentos/configuracoes.html",
        {
            "projetos": projetos,
            "responsaveis": resp_disc,
            "etapas": etapas,
            "config": config,
        },
    )

# =================================================================
# BUSCA GLOBAL
# =================================================================

@login_required
def buscar_global(request):
    termo = request.GET.get("q", "").strip()
    resultados = []

    if termo:
        qs = Documento.objects.filter(ativo=True).select_related("projeto")
        qs = qs.filter(
            Q(codigo__icontains=termo)
            | Q(titulo__icontains=termo)
            | Q(disciplina__icontains=termo)
            | Q(projeto__nome__icontains=termo)
        )

        for doc in qs:
            resultados.append(
                {
                    "id": doc.id,
                    "codigo": highlight_text(doc.codigo, termo),
                    "titulo": highlight_text(doc.titulo, termo),
                    "disciplina": doc.disciplina or "",
                    "projeto": doc.projeto.nome if doc.projeto else "",
                }
            )

    return render(
        request,
        "documentos/buscar.html",
        {"resultados": resultados, "termo": termo},
    )

# =============================================
# 📊 DASHBOARD MASTER – FINANCE + STATUS ENGINE
# =============================================
from django.db.models import Count, Sum
from .models import Documento, LogAuditoria, ProjetoFinanceiro


def dashboard_master(request):
    projeto_nome = "TP25 NAVIOS HANDY CLASSE 80"
    taxa_brl = 5.7642

    # ============================
    # BASE DE DOCUMENTOS FILTRADOS
    # ============================
    docs = Documento.objects.filter(
        projeto__nome__icontains=projeto_nome,
        ativo=True
    )

    total_docs = docs.count()

    total_emitidos     = docs.filter(status_emissao="Emitido").count()
    total_aprovados    = docs.filter(status_emissao="Aprovado").count()
    total_em_revisao   = docs.filter(status_documento__icontains="Revisão").count()
    total_nao_recebidos = docs.filter(status_emissao="Não Recebido").count()
    total_excluidos    = docs.filter(deletado_em__isnull=False).count()
    total_restaurados  = docs.filter(motivo_exclusao__icontains="Restaurado").count()

    # ============================
    # FINANCEIRO POR FASE / STATUS
    # ============================
    valores = ProjetoFinanceiro.objects.filter(projeto__nome=projeto_nome)

    valor_basico   = valores.filter(fase="Básico").aggregate(total=Sum("valor_total_usd"))["total"] or 0
    valor_aprovado = valores.filter(fase="Aprovado").aggregate(total=Sum("valor_total_usd"))["total"] or 0
    valor_asbuilt  = valores.filter(fase="As Built").aggregate(total=Sum("valor_total_usd"))["total"] or 0


    total_basico_docs = docs.filter(fase="Básico").count() or 1

    # USD Unitários (valor / documentos)
    usd_basico_unit   = valor_basico   / total_basico_docs
    usd_aprovado_unit = valor_aprovado / total_basico_docs
    usd_asbuilt_unit  = valor_asbuilt  / total_basico_docs

    # Medição baseada no STATUS EMISSÃO
    valor_emitidos_usd = total_emitidos     * usd_basico_unit
    valor_nao_rec_usd  = total_nao_recebidos * usd_basico_unit
    valor_total_usd    = valor_emitidos_usd + valor_nao_rec_usd

    # Conversão BRL
    valor_emitidos_brl = valor_emitidos_usd * taxa_brl
    valor_nao_rec_brl  = valor_nao_rec_usd  * taxa_brl
    valor_total_brl    = valor_total_usd    * taxa_brl

    # =============
    # AUDITORIA
    # =============
    auditoria_acoes = LogAuditoria.objects.order_by('-data')[:10]

    top_excluidores = LogAuditoria.objects.filter(acao="Excluído") \
        .values('usuario') \
        .annotate(qtd=Count('id')) \
        .order_by('-qtd')[:5]

    # =====================
    # GRÁFICOS
    # =====================
    disc_labels = list(docs.values_list("disciplina", flat=True).distinct())
    disc_data   = [docs.filter(disciplina=d).count() for d in disc_labels]

    status_labels = ["Emitido", "Aprovado", "Revisão", "Não Recebido"]
    status_data   = [
        total_emitidos,
        total_aprovados,
        total_em_revisao,
        total_nao_recebidos
    ]

    # =====================
    # RENDERIZA O TEMPLATE
    # =====================
    context = {
        "total_docs": total_docs,
        "total_emitidos": total_emitidos,
        "total_aprovados": total_aprovados,
        "total_em_revisao": total_em_revisao,
        "total_nao_recebidos": total_nao_recebidos,
        "total_excluidos": total_excluidos,
        "total_restaurados": total_restaurados,

        "valor_emitidos_usd": valor_emitidos_usd,
        "valor_nao_rec_usd": valor_nao_rec_usd,
        "valor_total_usd": valor_total_usd,

        "valor_emitidos_brl": valor_emitidos_brl,
        "valor_nao_rec_brl": valor_nao_rec_brl,
        "valor_total_brl": valor_total_brl,

        # Graphs
        "disc_labels": disc_labels,
        "disc_data": disc_data,
        "status_labels": status_labels,
        "status_data": status_data,

        # Auditoria
        "top_excluidores": top_excluidores,
        "auditoria_acoes": auditoria_acoes,
    }

    return render(request, "documentos/dashboard_master.html", context)

# =================================================================
# BUSCA AJAX
# =================================================================

@login_required
def buscar_ajax(request):
    termo = request.GET.get("q", "").strip()
    items = []

    if termo:
        qs = Documento.objects.filter(ativo=True)
        qs = qs.filter(
            Q(codigo__icontains=termo)
            | Q(titulo__icontains=termo)
            | Q(disciplina__icontains=termo)
        )[:20]

        for d in qs:
            items.append(
                {
                    "id": d.id,
                    "codigo": d.codigo,
                    "titulo": d.titulo,
                    "disciplina": d.disciplina or "",
                }
            )

    return JsonResponse({"results": items})
