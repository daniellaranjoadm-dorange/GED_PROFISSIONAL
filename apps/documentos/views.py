from django.db.models import Q
from django.views.decorators.http import require_POST
# apps/documentos/views.py


import csv
import json
import logging
import os
import hashlib
import re
from django.apps import apps
from django.urls import reverse
from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect
from datetime import date, datetime
from difflib import unified_diff
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db.models import Count, Q, Sum, OuterRef, Subquery
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.cache import never_cache
from django.utils import timezone
from django.conf import settings

from apps.contas.decorators import allow_admin
from apps.contas.permissions import has_perm

from .models import (
    ArquivoDocumento,
    Documento,
    DocumentoAprovacao,
    DocumentoVersao,
    DocumentoWorkflowStatus,
    LogAuditoria,
    Projeto,
    ProjetoFinanceiro,
    ResponsavelDisciplina,
    WorkflowEtapa,
    registrar_log,
)
from .utils_email import enviar_email

logger = logging.getLogger(__name__)

# ==============================================================
# DEFINIÇÃO OFICIAL DAS ETAPAS DO WORKFLOW (CODIFICADAS)
# ==============================================================


ETAPAS_WORKFLOW = [
    "ELABORACAO",
    "REVISAO_INTERNA",
    "APROVACAO_TECNICA",
    "DOC_CONTROL",
    "ENVIADO_CLIENTE",
    "APROVACAO_CLIENTE",
    "EMISSAO_FINAL",
]

ETAPAS_LABEL = {
    "ELABORACAO": "Elaboração",
    "REVISAO_INTERNA": "Revisão Interna – Disciplina",
    "APROVACAO_TECNICA": "Aprovação Técnica – Coordenador",
    "DOC_CONTROL": "Doc Control",
    "ENVIADO_CLIENTE": "Envio ao Cliente",
    "APROVACAO_CLIENTE": "Aprovação do Cliente",
    "EMISSAO_FINAL": "Emissão Final",
}

# =================================================================
# CONFIG GLOBAL
# =================================================================

from decimal import Decimal, InvalidOperation  # <-- garante precisão nos valores em dólar/reais

VALOR_MEDICAO_USD = Decimal("979.00")  # ainda usado na medição genérica
TAXA_CAMBIO_REAIS = Decimal("5.7642")  # taxa fixa que você passou

# =================================================================
# FUNÇÕES AUXILIARES – ENTERPRISE S7
# =================================================================

def normalizar_revisao(rev_bruto):
    """Normaliza revisões e valida formatos aceitos."""
    if rev_bruto is None:
        return ""

    rev = str(rev_bruto).strip().upper()

    if rev == "":
        return ""
    if rev == "0":
        return "0"
    if rev.isdigit():
        return None
    if any(ch.isdigit() for ch in rev):
        return None

    if rev.startswith("R") and rev[1:].isalpha():
        return rev[1:]

    return rev if rev.isalpha() else None


def highlight_text(texto, termo):
    """Destaca termo com <mark> (case-insensitive)."""
    if not texto or not termo:
        return texto or ""

    regex = re.compile(re.escape(termo), re.IGNORECASE)

    def repl(match):
        return (
            "<mark style='background:#ffeb3b; padding:2px; border-radius:4px;'>"
            f"{match.group(0)}</mark>"
        )

    return regex.sub(repl, texto)


def usuario_em_grupos(user, grupos):
    """Verifica se usuario pertence a qualquer grupo da lista."""
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=grupos).exists()


# -------------------------------------------------------------
# PERMISSÕES – BASEADAS EM WorkflowEtapa (objeto), não strings
# -------------------------------------------------------------

def pode_avancar_etapa(user, documento: Documento):
    """Verifica se o usuário pode avançar o documento para a próxima etapa."""

    if user.is_superuser:
        return True

    # Controle central: grupo Doc-Control sempre pode avançar
    doc_control = usuario_em_grupos(user, ["DOC_CONTROL", "Doc-Control", "DOC CONTROL"])

    if doc_control:
        return True

    # Se não há etapa definida ainda → permitido avançar para a primeira etapa
    if documento.etapa is None:
        return True

    # Se a etapa tem grupos responsáveis definidos
    grupos = documento.etapa.grupos_responsaveis.all()

    if grupos:
        return user.groups.filter(id__in=grupos).exists()

    # Se não há restrição configurada → libera avanço
    return True


def pode_retornar_etapa(user, documento: Documento):
    """Permissão para retornar etapa."""
    if user.is_superuser:
        return True

    # Doc-Control sempre pode retornar
    if usuario_em_grupos(user, ["DOC_CONTROL", "DOC CONTROL", "Doc-Control"]):
        return True

    # Coordenadores (COORD_*) também podem
    if user.groups.filter(name__startswith="COORD_").exists():
        return True

    return False


# -------------------------------------------------------------
# WORKFLOW ENTERPRISE — PRÓXIMA / ANTERIOR
# -------------------------------------------------------------

def proxima_etapa(documento):
    """
    Retorna a próxima etapa configurada no Workflow Enterprise.
    """

    # Sem etapa → assume a primeira
    if documento.etapa is None:
        return WorkflowEtapa.objects.filter(ativa=True).order_by("ordem").first()

    # Próxima etapa declarada no admin
    if documento.etapa.proxima_etapa:
        return documento.etapa.proxima_etapa

    # Fallback: próxima pela ordem
    return (
        WorkflowEtapa.objects.filter(
            ativa=True, ordem__gt=documento.etapa.ordem
        ).order_by("ordem").first()
    )


def etapa_anterior(documento: Documento):
    """
    Retorna a etapa anterior à atual, baseada na ordem.
    """

    if documento.etapa is None:
        return None

    return (
        WorkflowEtapa.objects.filter(
            ativa=True, ordem__lt=documento.etapa.ordem
        ).order_by("-ordem").first()
    )


# -------------------------------------------------------------
# REGISTRO DE WORKFLOW (LOG)
# -------------------------------------------------------------

def registrar_workflow(documento: Documento, etapa, acao: str = "", request=None, observacao: str = "", **kwargs):
    """
    Registra movimentação no log enterprise.
    'etapa' agora é WorkflowEtapa ou string com nome.
    """

    # Compatibilidade: versões antigas chamavam registrar_workflow(..., status="...")
    if not acao:
        acao = (kwargs.get("status") or "").strip()

    if request is None:
        raise ValueError("registrar_workflow requer request")

    if isinstance(etapa, str):
        etapa_obj = WorkflowEtapa.objects.filter(nome=etapa).first()
    else:
        etapa_obj = etapa

    usuario = request.user if request.user.is_authenticated else None

    descricao = (
        f"[WORKFLOW] Documento {documento.codigo} "
        f"→ {etapa_obj.nome if etapa_obj else etapa} | "
        f"Ação: {acao} | {observacao or ''}"
    )

    registrar_log(
        usuario,
        documento,
        f"Workflow: {acao}",
        descricao
    )

# -------------------------------------------------------------
# NOTIFICAÇÕES
# -------------------------------------------------------------

import os
import logging

logger = logging.getLogger(__name__)
DEFAULT_NOTIF_EMAIL = os.getenv("DEFAULT_NOTIF_EMAIL", "")

def _destinatarios_padrao(request=None, extras=None):
    """Monta lista de destinatários (sem duplicados)."""
    emails = []
    if extras:
        emails.extend([e for e in extras if e])

    if request is not None:
        user_email = getattr(getattr(request, "user", None), "email", "") or ""
        if user_email:
            emails.append(user_email)

    if DEFAULT_NOTIF_EMAIL:
        emails.append(DEFAULT_NOTIF_EMAIL)

    # remove duplicados preservando ordem
    seen = set()
    out = []
    for e in emails:
        e = (e or "").strip()
        if e and e not in seen:
            out.append(e)
            seen.add(e)
    return out


def _etapa_code(etapa) -> str:
    """Tenta extrair o código da etapa (ex: 'EMISSAO_FINAL')."""
    if etapa is None:
        return ""
    for attr in ("codigo", "slug", "chave", "nome"):
        val = getattr(etapa, attr, "") or ""
        if isinstance(val, str) and val.strip():
            return val.strip().upper()
    return ""


# -------------------------------------------------------------
# NOTIFICAÇÕES
# -------------------------------------------------------------
def notificar_evento_documento(documento: Documento, tipo_evento: str, destinatarios=None, etapa=None) -> bool:
    """Centraliza notificações por e-mail. Retorna True/False."""

    tipo_evento = (tipo_evento or "").strip()

    # Destinatários
    if destinatarios is None:
        destinatarios = _destinatarios_padrao()

    if not destinatarios:
        logger.warning(
            "Sem destinatários para notificação | doc=%s | evento=%s",
            getattr(documento, "codigo", ""),
            tipo_evento,
        )
        return False

    assunto = ""
    mensagem = ""

    # 1) Notificação por ETAPA (principal)
    # Aceita: tipo_evento='etapa' + etapa=WorkflowEtapa / ou tipo_evento='etapa_EMISSAO_FINAL'
    etapa_code = ""

    if tipo_evento.lower() == "etapa":
        etapa_code = _etapa_code(etapa) or _etapa_code(getattr(documento, "etapa", None))
    elif tipo_evento.upper().startswith("ETAPA_"):
        etapa_code = tipo_evento.upper().replace("ETAPA_", "", 1)

    if etapa_code:
        label = ETAPAS_LABEL.get(etapa_code, etapa_code)
        assunto = f"[GED] Documento na etapa: {label} ({etapa_code}) — {documento.codigo}"
        mensagem = (
            f"O documento {documento.codigo} (Rev {documento.revisao}) agora está na etapa: {label} ({etapa_code}).\n"
            f"Status LDP: {getattr(documento, 'status_documento', '')}\n"
            f"Status Emissão: {getattr(documento, 'status_emissao', '')}\n"
        )

    # 2) Eventos legados (compatibilidade)
    if not assunto:
        if tipo_evento == "envio_revisao":
            assunto = f"[GED] Documento em Revisão: {documento.codigo}"
            mensagem = f"O documento {documento.codigo} (Rev {documento.revisao}) foi enviado para revisão."
        elif tipo_evento == "aprovacao":
            assunto = f"[GED] Documento Aprovado: {documento.codigo}"
            mensagem = f"O documento {documento.codigo} (Rev {documento.revisao}) foi aprovado."
        elif tipo_evento == "emissao":
            assunto = f"[GED] Documento Emitido: {documento.codigo}"
            mensagem = (
                f"O documento {documento.codigo} (Rev {documento.revisao}) foi emitido em {documento.data_emissao_grdt}."
            )
        elif tipo_evento == "cancelamento":
            assunto = f"[GED] Documento Cancelado: {documento.codigo}"
            mensagem = f"O documento {documento.codigo} (Rev {documento.revisao}) foi cancelado."
        else:
            return False

    try:
        ok = enviar_email(
            assunto=assunto,
            mensagem=mensagem,
            destinatarios=destinatarios,
        )
        if not ok:
            logger.warning("Falha ao enviar e-mail | doc=%s | evento=%s", documento.codigo, tipo_evento)
        return ok
    except Exception as e:
        logger.exception("Erro ao enviar e-mail | doc=%s | evento=%s | erro=%s", documento.codigo, tipo_evento, e)
        return False

# -------------------------------------------------------------
# LIXEIRA
# -------------------------------------------------------------

def mover_para_lixeira(documento: Documento, request, motivo: str = ""):
    """Soft delete enterprise."""
    if not documento.ativo:
        return False

    documento.ativo = False
    documento.deletado_em = timezone.now()
    documento.deletado_por = request.user.username
    documento.motivo_exclusao = motivo or ""
    documento.save()

    registrar_workflow(documento, documento.etapa, "Excluído", request, motivo)
    return True


def restaurar_da_lixeira(documento: Documento, request):
    documento.ativo = True
    documento.deletado_em = None
    documento.deletado_por = None
    documento.motivo_exclusao = ""
    documento.save()

    registrar_workflow(documento, documento.etapa, "Restaurado", request)


# =================================================================
# DASHBOARD SIMPLES
# =================================================================

@login_required
def dashboard(request):
    docs = Documento.objects.filter(ativo=True)

    total_docs = docs.count()
    total_aprovados = docs.filter(status_documento="Aprovado").count()
    total_em_revisao = docs.filter(status_documento="Em Revisão").count()
    total_emitidos = docs.filter(status_emissao="Emitido").count()
    total_nao_recebidos = docs.filter(status_emissao="Não Recebido").count()

    valor_emitidos_usd = total_emitidos * VALOR_MEDICAO_USD
    valor_nao_rec_usd = total_nao_recebidos * VALOR_MEDICAO_USD
    valor_total_usd = valor_emitidos_usd + valor_nao_rec_usd

    return render(
        request,
        "documentos/dashboard.html",
        {
            "total_docs": total_docs,
            "total_aprovados": total_aprovados,
            "total_em_revisao": total_em_revisao,
            "total_emitidos": total_emitidos,
            "total_nao_recebidos": total_nao_recebidos,
            "valor_total_usd": f"{valor_total_usd:,.2f}",
        },
    )


# =================================================================
# DASHBOARD ENTERPRISE
# =================================================================

@login_required
@has_perm("sistema.dashboard_enterprise")
def dashboard_enterprise(request):
    docs = Documento.objects.filter(ativo=True).select_related("projeto")

    filtros = {
        "projeto": request.GET.get("projeto") or "",
        "disciplina": request.GET.get("disciplina") or "",
        "tipo_doc": request.GET.get("tipo_doc") or "",
        "status_documento": request.GET.get("status_ldp") or "",
        "status_emissao": request.GET.get("status_emissao") or "",
        "dt_ini": request.GET.get("dt_ini") or "",
        "dt_fim": request.GET.get("dt_fim") or "",
    }

    if filtros["projeto"]:
        docs = docs.filter(projeto__nome__icontains=filtros["projeto"])

    if filtros["disciplina"]:
        docs = docs.filter(disciplina=filtros["disciplina"])

    if filtros["tipo_doc"]:
        docs = docs.filter(tipo_doc=filtros["tipo_doc"])

    if filtros["status_documento"]:
        docs = docs.filter(status_documento=filtros["status_documento"])

    if filtros["status_emissao"]:
        docs = docs.filter(status_emissao=filtros["status_emissao"])

    if filtros["dt_ini"]:
        dt_ini = datetime.strptime(filtros["dt_ini"], "%Y-%m-%d").date()
        docs = docs.filter(data_emissao_grdt__gte=dt_ini)

    if filtros["dt_fim"]:
        dt_fim = datetime.strptime(filtros["dt_fim"], "%Y-%m-%d").date()
        docs = docs.filter(data_emissao_grdt__lte=dt_fim)

    total_docs = docs.count()
    total_emitidos = docs.filter(status_emissao="Emitido").count()
    total_nao_recebidos = docs.filter(status_emissao="Não Recebido").count()
    total_em_revisao = docs.filter(status_documento="Em Revisão").count()
    total_aprovados = docs.filter(status_documento="Aprovado").count()

    valor_emitidos_usd = total_emitidos * VALOR_MEDICAO_USD
    valor_nao_rec_usd = total_nao_recebidos * VALOR_MEDICAO_USD
    valor_total_usd = valor_emitidos_usd + valor_nao_rec_usd

    valor_emitidos_brl = valor_emitidos_usd * TAXA_CAMBIO_REAIS
    valor_nao_rec_brl = valor_nao_rec_usd * TAXA_CAMBIO_REAIS
    valor_total_brl = valor_emitidos_brl + valor_nao_rec_brl

    por_disc = docs.values("disciplina").annotate(qtd=Count("id"))
    disc_labels = [d["disciplina"] or "Sem Disciplina" for d in por_disc]
    disc_data = [d["qtd"] for d in por_disc]

    por_status = docs.values("status_documento").annotate(qtd=Count("id"))
    status_labels = [d["status_documento"] or "Sem Status" for d in por_status]
    status_data = [d["qtd"] for d in por_status]

    med_raw = docs.values("tipo_doc").annotate(
        total=Count("id"),
        emitidos=Count("id", filter=Q(status_emissao="Emitido")),
        nr=Count("id", filter=Q(status_emissao="Não Recebido")),
    )

    medicao_linhas = []
    total_usd = 0
    total_brl = 0

    for m in med_raw:
        tipo = m["tipo_doc"] or "Sem Tipo"
        emit = m["emitidos"]
        nr = m["nr"]

        v_emit_usd = emit * VALOR_MEDICAO_USD
        v_nr_usd = nr * VALOR_MEDICAO_USD
        v_emit_brl = v_emit_usd * TAXA_CAMBIO_REAIS
        v_nr_brl = v_nr_usd * TAXA_CAMBIO_REAIS

        total_usd += v_emit_usd + v_nr_usd
        total_brl += v_emit_brl + v_nr_brl

        medicao_linhas.append(
            {
                "tipo_doc": tipo,
                "total": m["total"],
                "emitidos": emit,
                "nao_recebidos": nr,
                "valor_emitidos_usd": f"{v_emit_usd:,.2f}",
                "valor_nr_usd": f"{v_nr_usd:,.2f}",
                "valor_emitidos_brl": f"{v_emit_brl:,.2f}",
                "valor_nr_brl": f"{v_nr_brl:,.2f}",
            }
        )

    medicao_totais = {
        "total_docs": sum(m["total"] for m in med_raw),
        "total_usd": f"{total_usd:,.2f}",
        "total_brl": f"{total_brl:,.2f}",
    }

    base_qs = Documento.objects.filter(ativo=True).select_related("projeto")

    return render(
        request,
        "documentos/dashboard_enterprise.html",
        {
            "filtros": filtros,
            "total_docs": total_docs,
            "total_emitidos": total_emitidos,
            "total_nao_recebidos": total_nao_recebidos,
            "total_em_revisao": total_em_revisao,
            "total_aprovados": total_aprovados,
            "valor_emitidos_usd": f"{valor_emitidos_usd:,.2f}",
            "valor_emitidos_brl": f"{valor_emitidos_brl:,.2f}",
            "valor_nao_rec_usd": f"{valor_nao_rec_usd:,.2f}",
            "valor_nao_rec_brl": f"{valor_nao_rec_brl:,.2f}",
            "valor_total_usd": f"{valor_total_usd:,.2f}",
            "valor_total_brl": f"{valor_total_brl:,.2f}",
            "disc_labels": json.dumps(disc_labels),
            "disc_data": json.dumps(disc_data),
            "status_labels": json.dumps(status_labels),
            "status_data": json.dumps(status_data),
            "medicao_linhas": medicao_linhas,
            "medicao_totais": medicao_totais,
            "lista_projetos": base_qs.values_list("projeto__nome", flat=True).distinct(),
            "lista_disciplinas": base_qs.values_list("disciplina", flat=True).distinct(),
            "lista_tipos": base_qs.values_list("tipo_doc", flat=True).distinct(),
            "lista_status_ldp": base_qs.values_list("status_documento", flat=True).distinct(),
            "lista_status_emissao": base_qs.values_list("status_emissao", flat=True).distinct(),
        },
    )


# =================================================================
# PAINEL DE WORKFLOW
# =================================================================

@login_required
def painel_workflow(request):
    docs_base = Documento.objects.filter(ativo=True).select_related("projeto")

    etapa_filtro = request.GET.get("etapa") or ""
    disciplina_filtro = request.GET.get("disciplina") or ""
    projeto_filtro = request.GET.get("projeto") or ""
    status_ldp_filtro = request.GET.get("status_ldp") or ""
    status_emissao_filtro = request.GET.get("status_emissao") or ""

    docs = docs_base

    if etapa_filtro and etapa_filtro != "Sem etapa":
        docs = docs.filter(etapa_atual=etapa_filtro)
    elif etapa_filtro == "Sem etapa":
        docs = docs.filter(Q(etapa_atual__isnull=True) | Q(etapa_atual=""))

    if disciplina_filtro:
        docs = docs.filter(disciplina=disciplina_filtro)

    if projeto_filtro:
        docs = docs.filter(projeto__nome=projeto_filtro)

    if status_ldp_filtro:
        docs = docs.filter(status_documento=status_ldp_filtro)

    if status_emissao_filtro:
        docs = docs.filter(status_emissao=status_emissao_filtro)

    order = request.GET.get("order") or ""
    direction = request.GET.get("direction") or "asc"

    valid_fields = {
        "codigo": "codigo",
        "titulo": "titulo",
        "disciplina": "disciplina",
        "projeto": "projeto__nome",
        "etapa_atual": "etapa_atual",
        "status_ldp": "status_documento",
        "status_emissao": "status_emissao",
        "revisao": "revisao",
    }

    if order in valid_fields:
        field = valid_fields[order]
        if direction == "desc":
            field = "-" + field
        docs = docs.order_by(field)
    else:
        docs = docs.order_by("etapa__ordem", "codigo")

    disciplinas = (
        docs_base.exclude(disciplina__isnull=True)
        .exclude(disciplina__exact="")
        .values_list("disciplina", flat=True)
        .distinct()
        .order_by("disciplina")
    )

    projetos = (
        docs_base.filter(projeto__isnull=False)
        .values_list("projeto__nome", flat=True)
        .distinct()
        .order_by("projeto__nome")
    )

    status_ldp_lista = (
        docs_base.exclude(status_documento__isnull=True)
        .exclude(status_documento__exact="")
        .values_list("status_documento", flat=True)
        .distinct()
        .order_by("status_documento")
    )

    status_emissao_lista = (
        docs_base.exclude(status_emissao__isnull=True)
        .exclude(status_emissao__exact="")
        .values_list("status_emissao", flat=True)
        .distinct()
        .order_by("status_emissao")
    )

    per_page = request.GET.get("per_page") or 25
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 25

    total_filtrados = docs.count()

    paginator = Paginator(docs, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    documentos = page_obj

    por_etapa_qs = docs_base.values("etapa_atual").annotate(qtd=Count("id"))
    mapa_etapas = {row["etapa_atual"] or "Sem etapa": row["qtd"] for row in por_etapa_qs}

    etapas_kpi = []
    for nome in ETAPAS_WORKFLOW:
        url = (
            f"?etapa={nome}"
            f"&disciplina={disciplina_filtro}"
            f"&projeto={projeto_filtro}"
            f"&status_ldp={status_ldp_filtro}"
            f"&status_emissao={status_emissao_filtro}"
            f"&order={order}"
            f"&direction={direction}"
            f"&per_page={per_page}"
        )
        etapas_kpi.append(
            {
                "nome": nome,
                "qtd": mapa_etapas.get(nome, 0),
                "url": url,
            }
        )

    url_sem_etapa = (
        f"?etapa=Sem%20etapa"
        f"&disciplina={disciplina_filtro}"
        f"&projeto={projeto_filtro}"
        f"&status_ldp={status_ldp_filtro}"
        f"&status_emissao={status_emissao_filtro}"
        f"&order={order}"
        f"&direction={direction}"
        f"&per_page={per_page}"
    )

    etapas_kpi.append(
        {
            "nome": "Sem etapa definida",
            "qtd": mapa_etapas.get("Sem etapa", mapa_etapas.get(None, 0)),
            "url": url_sem_etapa,
        }
    )

    ultimos_movimentos = (
        DocumentoAprovacao.objects
        .select_related("documento", "etapa", "usuario")
        .order_by("-data")[:20]
    )

    return render(
        request,
        "documentos/painel_workflow.html",
        {
            "documentos": documentos,
            "page_obj": page_obj,
            "per_page": per_page,
            "total_docs": total_filtrados,
            "order": order,
            "direction": direction,
            "etapas_kpi": etapas_kpi,
            "etapas_workflow": ETAPAS_WORKFLOW,
            "etapa_filtro": etapa_filtro,
            "disciplinas": disciplinas,
            "disciplina_filtro": disciplina_filtro,
            "projetos": projetos,
            "projeto_filtro": projeto_filtro,
            "status_ldp_lista": status_ldp_lista,
            "status_ldp_filtro": status_ldp_filtro,
            "status_emissao_lista": status_emissao_lista,
            "status_emissao_filtro": status_emissao_filtro,
            "ultimos_movimentos": ultimos_movimentos,
        },
    )


# =================================================================
# EXPORTAR EXCEL DO PAINEL
# =================================================================

@login_required
def painel_workflow_exportar_excel(request):
    etapa = request.GET.get("etapa") or ""
    disciplina = request.GET.get("disciplina") or ""
    projeto = request.GET.get("projeto") or ""
    status_ldp = request.GET.get("status_ldp") or ""
    status_emissao = request.GET.get("status_emissao") or ""

    docs = Documento.objects.filter(ativo=True).select_related("projeto")

    if etapa and etapa != "Sem etapa":
        docs = docs.filter(etapa_atual=etapa)
    elif etapa == "Sem etapa":
        docs = docs.filter(Q(etapa_atual__isnull=True) | Q(etapa_atual=""))

    if disciplina:
        docs = docs.filter(disciplina=disciplina)
    if projeto:
        docs = docs.filter(projeto__nome=projeto)
    if status_ldp:
        docs = docs.filter(status_documento=status_ldp)
    if status_emissao:
        docs = docs.filter(status_emissao=status_emissao)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Workflow GED"

    headers = [
        "Código",
        "Revisão",
        "Título",
        "Disciplina",
        "Projeto",
        "Etapa Atual",
        "Status Documento",
        "Status Emissão",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="0052A2", end_color="0052A2", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in ws[1]:
        col.fill = header_fill
        col.font = Font(color="FFFFFF", bold=True)
        col.border = border
        col.alignment = Alignment(horizontal="center")

    for d in docs:
        ws.append(
            [
                d.codigo or "",
                d.revisao or "",
                d.titulo or "",
                d.disciplina or "",
                d.projeto.nome if d.projeto else "",
                d.etapa_atual or "",
                d.status_documento or "",
                d.status_emissao or "",
            ]
        )

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"workflow_{date.today().isoformat()}.xlsx"

    return HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
# =====================================================================
# 📁 LISTA DE DOCUMENTOS – com filtros funcionando 100%
# =====================================================================
@never_cache
def listar_documentos(request):
    # Base: só documentos ativos e não deletados
    base_qs = Documento.objects.filter(ativo=True, deletado_em__isnull=True)

    # ---- Filtros vindos da barra superior ----
    projeto_filtro        = request.GET.get("projeto", "") or ""
    disciplina_filtro     = request.GET.get("disciplina", "") or ""
    status_doc_filtro     = request.GET.get("status_documento", "") or ""
    status_emissao_filtro = request.GET.get("status_emissao", "") or ""
    busca                 = request.GET.get("busca", "") or ""

    latest_pk = (
        base_qs.filter(codigo=OuterRef("codigo"))
        .order_by("-criado_em", "-id")
        .values("pk")[:1]
    )
    documentos = (
        base_qs.filter(pk=Subquery(latest_pk))
        .select_related("projeto")
        .annotate(versoes_count=Count("versoes__numero_revisao", distinct=True))
    )

    # Aplica filtros sobre o conjunto "atual" por código
    if projeto_filtro:
        documentos = documentos.filter(projeto__nome__icontains=projeto_filtro)

    if disciplina_filtro:
        documentos = documentos.filter(disciplina__icontains=disciplina_filtro)

    if status_doc_filtro:
        documentos = documentos.filter(status_documento__icontains=status_doc_filtro)

    if status_emissao_filtro:
        documentos = documentos.filter(status_emissao__icontains=status_emissao_filtro)

    # 🔥 CORREÇÃO DEFINITIVA DA BUSCA GLOBAL
    if busca:
        documentos = documentos.filter(
            Q(codigo__icontains=busca) |
            Q(titulo__icontains=busca)
        )

    documentos = documentos.order_by("codigo")

    # ---- Combos dos selects (usados no template listar.html) ----
    projetos = (
        Documento.objects
        .values_list("projeto__nome", flat=True)
        .exclude(projeto__nome__isnull=True)
        .exclude(projeto__nome__exact="")
        .distinct()
        .order_by("projeto__nome")
    )

    disciplinas = (
        Documento.objects
        .values_list("disciplina", flat=True)
        .exclude(disciplina__isnull=True)
        .exclude(disciplina__exact="")
        .distinct()
        .order_by("disciplina")
    )

    status_documento_list = (
        Documento.objects
        .values_list("status_documento", flat=True)
        .exclude(status_documento__isnull=True)
        .exclude(status_documento__exact="")
        .distinct()
        .order_by("status_documento")
    )

    status_emissao_list = (
        Documento.objects
        .values_list("status_emissao", flat=True)
        .exclude(status_emissao__isnull=True)
        .exclude(status_emissao__exact="")
        .distinct()
        .order_by("status_emissao")
    )

    # ---- Contexto para o template listar.html ----
    context = {
        "documentos": documentos,
        "projetos": projetos,
        "disciplinas": disciplinas,
        "status_documento_list": status_documento_list,
        "status_emissao_list": status_emissao_list,
    }

    return render(request, "documentos/listar.html", context)


@never_cache
def revisoes(request):
    base_qs = (
        Documento.objects.filter(ativo=True, deletado_em__isnull=True)
        .exclude(revisao__isnull=True)
        .exclude(revisao__exact="")
    )

    latest_pk = (
        base_qs.filter(codigo=OuterRef("codigo"))
        .order_by("-criado_em", "-id")
        .values("pk")[:1]
    )

    versoes_count = (
        DocumentoVersao.objects.filter(documento__codigo=OuterRef("codigo"))
        .values("documento__codigo")
        .annotate(c=Count("numero_revisao", distinct=True))
        .values("c")[:1]
    )

    documentos = (
        base_qs.filter(pk=Subquery(latest_pk))
        .select_related("projeto")
        .annotate(versoes_count=Subquery(versoes_count))
        .filter(versoes_count__gt=0)
        .order_by("-criado_em", "codigo")
    )

    rows = []
    for doc in documentos:
        rows.append(
            {
                "codigo": doc.codigo,
                "rev_atual": (str(getattr(doc, "revisao", "") or "").strip() or "—"),
                "versoes_count": doc.versoes_count or 0,
                "projeto_nome": getattr(getattr(doc, "projeto", None), "nome", "") or "",
                "disciplina": getattr(doc, "disciplina", "") or "",
                "doc_id": doc.id,
            }
        )

    return render(
        request,
        "documentos/revisoes.html",
        {
            "rows": rows,
            "total": len(rows),
        },
    )

# ==============================
# Revisões permitidas no sistema
# ==============================
REVISOES_VALIDAS = ["0"] + [chr(i) for i in range(ord("A"), ord("Z")+1)]


def _documento_atual_por_codigo(qs_base, codigo):
    return qs_base.filter(codigo=codigo).order_by("-criado_em", "-id").first()

# =================================================================
# DETALHE DO DOCUMENTO
# =================================================================

@login_required
def detalhes_documento(request, documento_id):
    documento = get_object_or_404(
        Documento.objects.select_related("projeto", "etapa"), id=documento_id
    )

    anexos = list(documento.arquivos.all())
    for a in anexos:
        try:
            a.url_absoluta = request.build_absolute_uri(a.arquivo.url)
        except Exception:
            a.url_absoluta = a.arquivo.url

    versoes_qs = documento.versoes.all().order_by("-criado_em")
    versao_atual = versoes_qs.first()

    rev_atual = str(documento.revisao or "").strip().upper()
    if rev_atual not in REVISOES_VALIDAS:
        rev_atual = "0"

    try:
        idx = REVISOES_VALIDAS.index(rev_atual)
        proxima_revisao = (
            REVISOES_VALIDAS[idx + 1]
            if idx + 1 < len(REVISOES_VALIDAS)
            else rev_atual
        )
    except ValueError:
        proxima_revisao = "A"

    workflow_status = getattr(documento, "workflow_status", None)

    context = {
        "documento": documento,
        "anexos": anexos,
        "versoes": versoes_qs,
        "versao_atual": versao_atual,
        "proxima_revisao": proxima_revisao,
        "workflow_status": workflow_status,
        "pode_avancar_etapa": pode_avancar_etapa(request.user, documento),
        "proxima_etapa": proxima_etapa(documento),
        "pode_retornar_etapa": pode_retornar_etapa(request.user, documento),
        "etapa_anterior": etapa_anterior(documento),
        "historico_workflow": documento.historico_workflow.select_related(
            "etapa", "usuario"
        ).order_by("-data"),
    }

    return render(request, "documentos/detalhes.html", context)


# =================================================================
# HISTÓRICO POR CÓDIGO
# =================================================================

@login_required
def historico(request, codigo):
    base = Documento.objects.filter(codigo=codigo, ativo=True, deletado_em__isnull=True)
    documento = _documento_atual_por_codigo(base, codigo)
    if not documento:
        raise Http404("Documento não encontrado")

    versoes = DocumentoVersao.objects.filter(documento__codigo=codigo).select_related(
        "criado_por"
    ).order_by("-criado_em", "-id")

    versoes_resumo = []
    versoes_por_numero = {}
    for v in versoes:
        numero = (
            str(getattr(v, "numero_revisao", None) or getattr(v, "revisao", "") or "")
            .strip()
        )
        if not numero:
            numero = "-"

        entry = versoes_por_numero.get(numero)
        if not entry:
            entry = {
                "numero_revisao": numero,
                "criado_em": getattr(v, "criado_em", None),
                "criado_por": getattr(v, "criado_por", None),
                "status_revisao": getattr(v, "status_revisao", "") or "",
                "arquivos": [],
            }
            versoes_por_numero[numero] = entry
            versoes_resumo.append(entry)

        arquivo = getattr(v, "arquivo", None)
        if arquivo:
            try:
                url = arquivo.url
            except Exception:
                url = ""
            nome = getattr(v, "nome_original", "") or os.path.basename(
                getattr(arquivo, "name", "") or ""
            )
            entry["arquivos"].append(
                {
                    "nome": nome or "arquivo",
                    "url": url,
                }
            )

    return render(
        request,
        "documentos/historico.html",
        {
            "documento": documento,
            "versoes_resumo": versoes_resumo,
        },
    )


# =================================================================
# AVANÇAR / RETORNAR ETAPA — WORKFLOW ENTERPRISE (COM OBS + ANEXOS)
# =================================================================

@login_required
def enviar_proxima_etapa(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    if request.method != "POST":
        messages.error(request, "Requisição inválida.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    if not pode_avancar_etapa(request.user, documento):
        messages.error(request, "Você não tem permissão para avançar esta etapa.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    observacao = (request.POST.get("observacao") or "").strip()
    anexos = request.FILES.getlist("anexos")

    if not observacao:
        messages.warning(request, "Informe o motivo/observação para avançar a etapa.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    nova_etapa = documento.enviar_para_proxima_etapa(
        usuario=request.user,
        observacao=observacao,
        anexos=anexos,
    )

    if not nova_etapa:
        messages.warning(request, "Este documento já está na última etapa do workflow.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    registrar_workflow(documento, nova_etapa, "Avançou etapa", request, observacao)

    # Notificação por e-mail: sempre que muda de etapa
    notificar_evento_documento(
        documento,
        "etapa",
        destinatarios=_destinatarios_padrao(request),
        etapa=nova_etapa,
    )
    messages.success(request, f"Documento avançado para: {nova_etapa.nome}")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)


@login_required
def retornar_etapa(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    if request.method != "POST":
        messages.error(request, "Requisição inválida.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    if not pode_retornar_etapa(request.user, documento):
        messages.error(request, "Sem permissão para retornar etapa.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    observacao = (request.POST.get("observacao") or request.POST.get("motivo") or "").strip()
    anexos = request.FILES.getlist("anexos")

    if not observacao:
        messages.warning(request, "Informe o motivo/observação para retornar a etapa.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    destino = etapa_anterior(documento)
    if not destino:
        messages.warning(request, "Documento já está na primeira etapa.")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    documento.retornar_etapa(
        etapa_destino=destino,
        usuario=request.user,
        motivo=observacao,
        anexos=anexos,
    )

    registrar_workflow(documento, destino, "Retornou etapa", request, observacao)

    # Notificação por e-mail: sempre que muda de etapa
    notificar_evento_documento(
        documento,
        "etapa",
        destinatarios=_destinatarios_padrao(request),
        etapa=destino,
    )
    messages.success(request, f"Documento retornado para: {destino.nome}")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)
def nova_versao(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    rev_atual = str(documento.revisao or "").strip().upper()
    if rev_atual not in REVISOES_VALIDAS:
        rev_atual = "0"
    try:
        idx = REVISOES_VALIDAS.index(rev_atual)
        proxima_revisao = (
            REVISOES_VALIDAS[idx + 1] if idx + 1 < len(REVISOES_VALIDAS) else rev_atual
        )
    except ValueError:
        proxima_revisao = "A"

    if request.method == "POST":
        numero_revisao = request.POST.get("numero_revisao", "").strip().upper()
        observacao = request.POST.get("observacao", "").strip()
        arquivo = request.FILES.get("arquivo")

        if not arquivo:
            messages.error(request, "Selecione um arquivo.")
            return redirect("documentos:detalhes_documento", documento_id=documento.id)

        if not numero_revisao:
            numero_revisao = proxima_revisao

        if numero_revisao not in REVISOES_VALIDAS:
            messages.error(request, "Revisão inválida.")
            return redirect("documentos:detalhes_documento", documento_id=documento.id)

        DocumentoVersao.objects.create(
            documento=documento,
            numero_revisao=numero_revisao,
            arquivo=arquivo,
            observacao=observacao,
            criado_por=request.user,
            status_revisao="REVISAO",
        )

        documento.revisao = numero_revisao
        documento.save(update_fields=["revisao"])

        registrar_workflow(
            documento,
            etapa="Nova Versão",
            acao="Criada",
            request=request,
            observacao=f"Versão {numero_revisao} adicionada.",
        )

        messages.success(request, "Nova versão adicionada com sucesso!")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    return redirect("documentos:detalhes_documento", documento_id=documento.id)

# =================================================================
# UPLOAD DOCUMENTO (CRIAR)
# =================================================================
import logging
logger = logging.getLogger(__name__)

@login_required
@has_perm("documento.criar")
def upload_documento(request):
    if request.method == "POST":
        revisao = normalizar_revisao(request.POST.get("revisao"))
        if revisao is None:
            messages.error(request, "Revisão inválida!")
            return redirect("documentos:upload_documento")
        revisao = revisao or "0"

        titulo = (request.POST.get("titulo") or "").strip()
        codigo = (request.POST.get("codigo") or "").strip()
        if not titulo or not codigo:
            messages.error(request, "Preencha Título e Código.")
            return redirect("documentos:upload_documento")

        # projeto (opcional)
        projeto = None
        projeto_id = (request.POST.get("projeto") or "").strip()
        if projeto_id:
            projeto = Projeto.objects.filter(id=projeto_id, ativo=True).first()

        if not projeto:
            projeto = Projeto.objects.filter(ativo=True).order_by("id").first()

        if not projeto:
            messages.error(request, "Cadastre ao menos 1 Projeto antes de criar documentos.")
            return redirect("documentos:upload_documento")

        # cria documento (independente do upload)
        doc = Documento.objects.create(
            projeto=projeto,
            titulo=titulo,
            codigo=codigo,
            revisao=revisao,
            disciplina=((request.POST.get("disciplina") or "").strip() or None),
            tipo_doc=((request.POST.get("tipo_doc") or "").strip() or None),
            fase=((request.POST.get("fase") or "").strip() or ""),
            ativo=True,
            deletado_em=None,
        )

        # arquivos: "arquivos" (multiple) ou "arquivo" (single)
        arquivos = request.FILES.getlist("arquivos") or []
        if not arquivos:
            unico = request.FILES.get("arquivo")
            if unico:
                arquivos = [unico]

        anexados = 0
        falha_storage = False

        for arq in arquivos:
            try:
                ArquivoDocumento.objects.create(
                    documento=doc,
                    arquivo=arq,  # R2/S3 PutObject
                    nome_original=getattr(arq, "name", None),
                    tipo=(arq.name.split(".")[-1].lower() if getattr(arq, "name", "") else None),
                )
                anexados += 1
            except Exception:
                falha_storage = True
                logger.exception("Falha ao salvar anexo no storage (R2/S3)")
                break

        # workflow (não pode derrubar)
        try:
            registrar_workflow(doc, "Criação", "Criado", request)
        except Exception:
            logger.exception("Falha ao registrar workflow")

        # auditoria (não usa montar_descricao_log -> evita NameError)
        try:
            usuario = getattr(request.user, "username", None) or getattr(request.user, "email", None) or str(request.user)
            descricao = f"{usuario} criou o documento {doc.codigo} (Rev {doc.revisao}) - {doc.titulo}"
            registrar_log(request.user, doc, "Criação de Documento", descricao)
        except Exception:
            logger.exception("Falha ao registrar log de auditoria")

        if falha_storage:
            messages.warning(
                request,
                "Documento criado, mas falhou o upload do arquivo (R2 Unauthorized). "
                "Corrija as credenciais/permissões do R2 no Railway."
            )
        elif anexados == 0 and arquivos:
            messages.warning(request, "Documento criado, mas sem anexos.")
        else:
            messages.success(request, f"Documento criado com sucesso! ({anexados} arquivo(s))")

        return redirect("documentos:listar_documentos")

    projetos = Projeto.objects.filter(ativo=True).order_by("nome")
    return render(
        request,
        "documentos/upload.html",
        {"projetos": projetos, "REVISOES_VALIDAS": REVISOES_VALIDAS},
    )


# =================================================================
# EDITAR DOCUMENTO
# =================================================================

def _set_if_exists(obj, field, value):
    if value is None:
        return
    if hasattr(obj, field):
        setattr(obj, field, value)


def _parse_date(value):
    if not value:
        return None
    value = str(value).strip()
    try:
        if len(value) == 10 and value[4] == "-" and value[7] == "-":
            return datetime.strptime(value, "%Y-%m-%d").date()
        if len(value) == 10 and value[2] == "/" and value[5] == "/":
            return datetime.strptime(value, "%d/%m/%Y").date()
    except Exception:
        return None
    return None


def _first_post_value(request, keys):
    for key in keys:
        val = request.POST.get(key)
        if val is None:
            continue
        val = str(val).strip()
        if val != "":
            return val
    return None


def _first_attr_value(obj, fields):
    for field in fields:
        if hasattr(obj, field):
            val = getattr(obj, field)
            if val not in (None, ""):
                return val
    return None


def _format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value) if value else ""

@login_required
@has_perm("documento.editar")
def editar_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    if request.method == "POST":
        revisao = normalizar_revisao(request.POST.get("revisao"))
        if revisao is None:
            messages.error(request, "Revisão inválida.")
            return redirect("documentos:editar_documento", documento_id=documento.id)

        projeto_id = (request.POST.get("projeto") or "").strip()
        if projeto_id:
            try:
                if projeto_id.isdigit():
                    projeto = get_object_or_404(Projeto, id=projeto_id)
                else:
                    projeto = get_object_or_404(Projeto, nome__iexact=projeto_id)
                documento.projeto = projeto
            except Http404:
                messages.error(request, "Projeto inválido.")
                return redirect("documentos:editar_documento", documento_id=documento.id)

        documento.fase = request.POST.get("fase") or ""
        documento.tipo_doc = request.POST.get("tipo_doc") or ""
        documento.codigo = request.POST.get("codigo") or ""
        documento.disciplina = request.POST.get("disciplina") or ""
        documento.titulo = request.POST.get("titulo") or ""
        status_doc = (request.POST.get("status_documento")
                      or request.POST.get("status_ldp")
                      or request.POST.get("status")
                      or "")
        documento.status_documento = status_doc
        documento.status_emissao = request.POST.get("status_emissao") or ""
        documento.ged_interna = request.POST.get("ged_interna") or ""
        documento.revisao = revisao

        grdt_val = _first_post_value(
            request,
            ["num_grdt", "numero_grdt", "grdt", "n_grdt", "grdt_cliente"],
        )
        pcf_val = _first_post_value(
            request,
            ["num_pcf", "numero_pcf", "pcf", "n_pcf", "resposta_cliente"],
        )
        dt_raw = _first_post_value(
            request,
            [
                "data_emissao_tp",
                "data_emissao_tp_doc",
                "data_emissao_grdt",
                "data_emissao",
                "data",
            ],
        )
        dt_val = _parse_date(dt_raw)

        _set_if_exists(documento, "num_grdt", grdt_val)
        _set_if_exists(documento, "numero_grdt", grdt_val)
        _set_if_exists(documento, "grdt", grdt_val)
        _set_if_exists(documento, "n_grdt", grdt_val)
        _set_if_exists(documento, "grdt_cliente", grdt_val)

        _set_if_exists(documento, "num_pcf", pcf_val)
        _set_if_exists(documento, "numero_pcf", pcf_val)
        _set_if_exists(documento, "pcf", pcf_val)
        _set_if_exists(documento, "n_pcf", pcf_val)
        _set_if_exists(documento, "resposta_cliente", pcf_val)

        _set_if_exists(documento, "data_emissao_tp", dt_val)
        _set_if_exists(documento, "data_emissao_tp_doc", dt_val)
        _set_if_exists(documento, "data_emissao_grdt", dt_val)
        _set_if_exists(documento, "data_emissao", dt_val)
        _set_if_exists(documento, "data", dt_val)

        documento.save()
        documento.refresh_from_db()

        grdt_saved = _first_attr_value(
            documento,
            ["num_grdt", "numero_grdt", "grdt", "n_grdt", "grdt_cliente"],
        )
        pcf_saved = _first_attr_value(
            documento,
            ["num_pcf", "numero_pcf", "pcf", "n_pcf", "resposta_cliente"],
        )
        dt_saved = _first_attr_value(
            documento,
            [
                "data_emissao_tp",
                "data_emissao_tp_doc",
                "data_emissao_grdt",
                "data_emissao",
                "data",
            ],
        )

        messages.success(
            request,
            f"Salvo: GRDT={grdt_saved or '—'} PCF={pcf_saved or '—'} DATA={_format_date(dt_saved) or '—'}",
        )
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    projetos = Projeto.objects.filter(ativo=True).order_by("nome")
    return render(
        request,
        "documentos/editar.html",
        {
            "documento": documento,
            "projetos": projetos,
            "REVISOES_VALIDAS": REVISOES_VALIDAS,
        },
    )

# =================================================================
# NOVA REVISÃO (WORKFLOW)
# =================================================================

def _reset_campos_emissao(doc):
    """
    Zera campos que NÃO podem ser herdados ao criar nova revisão.
    Feito com hasattr para suportar variações de nomes no model.
    """
    for fname in [
        "status_emissao",
        "status_emissao_tp",
        "status_emissao_cliente",
        "grdt_cliente",
        "resposta_cliente",
    ]:
        if hasattr(doc, fname):
            setattr(doc, fname, "")

    for fname in [
        "num_grdt", "numero_grdt", "grdt", "n_grdt",
        "num_pcf", "numero_pcf", "pcf", "n_pcf",
    ]:
        if hasattr(doc, fname):
            setattr(doc, fname, "")

    for fname in [
        "data_emissao_tp",
        "data_emissao_tp_doc",
        "data_emissao",
        "data_emissao_grdt",
        "data",
    ]:
        if hasattr(doc, fname):
            setattr(doc, fname, None)

    return doc


@login_required
@has_perm("documento.revisar")
def nova_revisao(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    idx = REVISOES_VALIDAS.index(documento.revisao) if documento.revisao in REVISOES_VALIDAS else -1
    nova_rev = REVISOES_VALIDAS[idx + 1] if idx + 1 < len(REVISOES_VALIDAS) else "A"

    if request.method == "POST":
        arquivos = request.FILES.getlist("arquivo") or request.FILES.getlist("arquivos")
        observacao = request.POST.get("observacao", "")

        if not arquivos:
            messages.error(request, "Envie ao menos 1 arquivo da nova revisão.")
            return redirect("documentos:detalhes_documento", documento_id=documento.id)

        with transaction.atomic():
            documento = Documento.objects.select_for_update().get(pk=documento_id)

            for arquivo in arquivos:
                try:
                    DocumentoVersao.objects.create(
                        documento=documento,
                        numero_revisao=nova_rev,
                        arquivo=arquivo,  # chama storage aqui (R2/S3)
                        criado_por=request.user,
                        observacao=observacao,
                        status_revisao="REVISAO",
                    )
                except Exception:
                    logger.exception("Falha ao salvar nova revisão no storage (R2/S3)")
                    messages.error(
                        request,
                        "Falha ao enviar o arquivo da nova revisão (storage/R2 Unauthorized). "
                        "A revisão do documento NÃO foi alterada."
                    )
                    return redirect("documentos:detalhes_documento", documento_id=documento.id)

            documento.revisao = nova_rev
            documento.status_documento = "Em Revisão"
            _reset_campos_emissao(documento)
            update_fields = ["revisao", "status_documento"]
            for fname in [
                "status_emissao",
                "status_emissao_tp",
                "status_emissao_cliente",
                "grdt_cliente",
                "resposta_cliente",
                "num_grdt",
                "numero_grdt",
                "grdt",
                "n_grdt",
                "num_pcf",
                "numero_pcf",
                "pcf",
                "n_pcf",
                "data_emissao_tp",
                "data_emissao_tp_doc",
                "data_emissao",
                "data_emissao_grdt",
                "data",
            ]:
                if hasattr(documento, fname):
                    update_fields.append(fname)
            documento.save(update_fields=update_fields)

        registrar_workflow(documento, "Nova Revisão", "Criado", request)

        # Notificação por e-mail (usa email do usuário; se não tiver, cai no DEFAULT_NOTIF_EMAIL)
        notificar_evento_documento(documento, "envio_revisao", destinatarios=_destinatarios_padrao(request))

        messages.success(request, f"Revisão {nova_rev} criada com sucesso!")
        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    return render(
        request,
        "documentos/nova_revisao.html",
        {"documento": documento, "proxima_revisao": nova_rev},
    )

# =================================================================
# UPLOAD DE ANEXOS
# =================================================================

@login_required
def adicionar_arquivos(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    if request.method == "POST":
        arquivos = request.FILES.getlist("arquivos") or []

        if not arquivos:
            messages.error(request, "Selecione ao menos 1 arquivo.")
            return redirect("documentos:detalhes_documento", documento_id=documento.id)

        anexados = 0
        falhas = 0

        for arq in arquivos:
            try:
                ArquivoDocumento.objects.create(
                    documento=documento,
                    arquivo=arq,  # aqui é onde o S3/R2 tenta fazer PutObject
                    nome_original=getattr(arq, "name", None),
                    tipo=(arq.name.split(".")[-1].lower() if getattr(arq, "name", "") else None),
                )
                anexados += 1
            except Exception:
                falhas += 1
                logger.exception("Falha ao salvar anexo no storage (R2/S3)")

        # Log de workflow (ajusta a observação para refletir sucesso/falha)
        registrar_workflow(
            documento,
            "Upload de anexos",
            "Arquivos adicionados",
            request,
            observacao=f"{anexados} enviado(s), {falhas} falha(s)",
        )

        if anexados > 0 and falhas == 0:
            messages.success(request, "Arquivos enviados com sucesso!")
        elif anexados > 0 and falhas > 0:
            messages.warning(
                request,
                f"{anexados} arquivo(s) enviado(s), mas {falhas} falharam (storage/R2). Verifique credenciais/permissões.",
            )
        else:
            messages.error(
                request,
                "Nenhum arquivo foi anexado (falha no storage/R2). Verifique credenciais/permissões do Cloudflare R2.",
            )

        return redirect("documentos:detalhes_documento", documento_id=documento.id)

    return render(request, "documentos/adicionar_arquivos.html", {"documento": documento})

# =================================================================
# EXCLUIR ARQUIVO INDIVIDUAL
# =================================================================

@login_required
def excluir_arquivo(request, arquivo_id):
    arq = get_object_or_404(ArquivoDocumento, id=arquivo_id)
    documento_id = arq.documento.id

    try:
        default_storage.delete(arq.arquivo.name)
    except Exception:
        pass

    arq.delete()

    messages.success(request, "Arquivo excluído com sucesso.")
    return redirect("documentos:detalhes_documento", documento_id=documento_id)

# =================================================================
# WORKFLOW DE EMISSÃO / APROVAÇÃO
# =================================================================

@login_required
def enviar_para_revisao(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    documento.status_documento = "Em Revisão"
    documento.save(update_fields=["status_documento"])

    registrar_workflow(documento, "Revisão Interna – Disciplina", "Enviado", request)

    notificar_evento_documento(documento, "envio_revisao", destinatarios=_destinatarios_padrao(request))

    messages.success(request, "Documento enviado para revisão.")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)

@login_required
def aprovar_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    documento.status_documento = "Aprovado"
    documento.save(update_fields=["status_documento"])

    registrar_workflow(documento, "Aprovação Técnica – Coordenador", "Aprovado", request)

    notificar_evento_documento(documento, "aprovacao", destinatarios=_destinatarios_padrao(request))

    messages.success(request, "Documento aprovado.")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)

@login_required
def emitir_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    documento.status_emissao = "Emitido"
    documento.data_emissao_grdt = date.today()
    documento.save(update_fields=["status_emissao", "data_emissao_grdt"])

    registrar_workflow(documento, "Emissão Final", "Emitido", request)

    notificar_evento_documento(documento, "emissao", destinatarios=_destinatarios_padrao(request))

    messages.success(request, "Documento emitido.")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)

@login_required
def cancelar_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)

    documento.status_documento = "Cancelado"
    documento.status_emissao = "Cancelado"
    documento.save(update_fields=["status_documento", "status_emissao"])

    registrar_workflow(documento, "Emissão Final", "Cancelado", request)

    notificar_evento_documento(documento, "cancelamento", destinatarios=_destinatarios_padrao(request))

    messages.success(request, "Documento cancelado.")
    return redirect("documentos:detalhes_documento", documento_id=documento.id)

# =================================================================
# GERAR DIFF ENTRE DUAS REVISÕES
# =================================================================


@login_required
def gerar_diff(request, documento_id, revA, revB):
    documento = get_object_or_404(Documento, id=documento_id)

    versao_a = get_object_or_404(
        DocumentoVersao, documento=documento, numero_revisao=revA
    )
    versao_b = get_object_or_404(
        DocumentoVersao, documento=documento, numero_revisao=revB
    )

    def ler_arquivo_texto(file_field):
        try:
            path = file_field.path
        except Exception:
            return None
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read().splitlines()
        except Exception:
            return None

    linhas_a = ler_arquivo_texto(versao_a.arquivo)
    linhas_b = ler_arquivo_texto(versao_b.arquivo)

    diff_text = []
    if linhas_a is not None and linhas_b is not None:
        diff_text = list(
            unified_diff(
                linhas_a,
                linhas_b,
                fromfile=f"Rev {revA}",
                tofile=f"Rev {revB}",
                lineterm="",
            )
        )

    return render(
        request,
        "documentos/diff.html",
        {
            "documento": documento,
            "versao_a": versao_a,
            "versao_b": versao_b,
            "diff": diff_text,
        },
    )
# === IMPORTADOR LDP (TOP FIVE) - sem quebrar backend ===
import re
from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect
from django.apps import apps

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("ç", "c").replace("ã", "a").replace("á", "a").replace("à", "a").replace("â", "a")
    s = s.replace("é", "e").replace("ê", "e").replace("í", "i").replace("ó", "o").replace("ô", "o").replace("õ", "o")
    s = s.replace("ú", "u")
    s = re.sub(r"[^a-z0-9 _/-]", "", s)
    return s


def _find_header_row(ws, max_scan=25):
    """
    Procura uma linha que pareça cabeçalho (contenha "codigo" ou "documento" etc).
    Retorna (row_index, headers_list).
    """
    for r in range(1, max_scan + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        norms = [_norm(str(v)) if v is not None else "" for v in values]
        joined = " | ".join(norms)

        # heurística: cabeçalho costuma ter "codigo" e "titulo"/"descricao" ou "rev"
        if ("codigo" in joined or "documento" in joined or "doc" in joined) and (
            "titulo" in joined or "descricao" in joined or "rev" in joined or "revisao" in joined
        ):
            return r, norms

    # fallback: primeira linha
    values = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    norms = [_norm(str(v)) if v is not None else "" for v in values]
    return 1, norms


def _colmap(headers_norm):
    """
    Monta mapa campo->índice de coluna (1-based) por variações de nome.
    """
    candidates = {
        "codigo": ["codigo", "cod", "documento", "doc", "numero documento", "n documento", "document no", "doc no"],
        "titulo": ["titulo", "descricao", "descrição", "document title", "title", "nome documento", "nome"],
        "revisao": ["rev", "revisao", "revisão", "revision"],
        "disciplina": ["disciplina", "discipline"],
        "fase": ["fase", "phase"],
        "tipo_doc": ["tipo documento", "tipo doc", "tipo", "document type"],
    }

    out = {}
    for field, keys in candidates.items():
        for i, h in enumerate(headers_norm, start=1):
            if not h:
                continue
            for k in keys:
                if k in h:
                    out[field] = i
                    break
            if field in out:
                break
    return out


def _find_history_model():
    """
    Encontra automaticamente um model de histórico no app documentos, se existir,
    baseado na presença de campos esperados (arquivo_nome, total_sucesso, total_erros).
    """
    try:
        cfg = apps.get_app_config("documentos")
    except Exception:
        return None

    for m in cfg.get_models():
        fn = {f.name for f in m._meta.fields}
        if {"arquivo_nome", "total_sucesso", "total_erros"}.issubset(fn):
            return m
    return None


def importar_ldp(request):
    """Importa uma planilha LDP (.xlsx/.xlsm) e cria/atualiza Documentos.

    ✅ Mantém a lógica principal (update_or_create) intacta, mas agora:
    - valida colunas/campos e registra erros por LINHA + COLUNA;
    - gera um resumo de quais colunas estão com problema;
    - alimenta um histórico simples em session (sem migrations).
    """
    import pandas as pd
    from io import BytesIO
    from collections import Counter
    from django.db import transaction

    # =========================
    # Helpers
    # =========================
    def _norm(s: str) -> str:
        s = (s or "").strip().lower()
        import unicodedata
        s = "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
        s = re.sub(r"\s+", " ", s)
        return s

    def _clean(v):
        if v is None:
            return ""
        # pandas NaN
        try:
            import math
            if isinstance(v, float) and math.isnan(v):
                return ""
        except Exception:
            pass
        s = str(v).strip()
        return "" if s.lower() in {"nan", "none", "null"} else s

    def _find_header_row(file_bytes: bytes, sheet_name: str = "LDP", max_scan: int = 25) -> int:
        """Retorna a linha (1-index) do cabeçalho. Cai em 1 se não encontrar."""
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

        for r in range(1, max_scan + 1):
            vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 80) + 1)]
            norm = [_norm(str(v)) for v in vals if v is not None and str(v).strip() != ""]
            joined = " | ".join(norm)
            if "codigo" in joined and ("titulo" in joined or "title" in joined):
                return r
        return 1

    def _pick_column(df_columns, candidates):
        """Dado df.columns e lista de nomes candidatos, retorna o nome real da coluna (ou None)."""
        norm_map = {}
        for c in df_columns:
            norm_map.setdefault(_norm(str(c)), str(c))
        for cand in candidates:
            key = _norm(cand)
            if key in norm_map:
                return norm_map[key]
        return None

    def _add_error(items, excel_row, codigo, coluna, valor, mensagem):
        items.append({
            "linha": int(excel_row),
            "codigo": (codigo or "").strip() or "—",
            "coluna": (coluna or "").strip() or "—",
            "valor": ("" if valor is None else str(valor))[:220],
            "mensagem": str(mensagem)[:400],
        })

    def _error_line(e):
        base = f"L{e['linha']} · {e['coluna']}: {e['mensagem']}"
        if e.get("codigo") and e["codigo"] != "—":
            base += f" (cód: {e['codigo']})"
        return base

    # =========================
    # Session history (sem model)
    # =========================
    history = request.session.get("ldp_import_history", [])
    if not isinstance(history, list):
        history = []

    if request.method != "POST":
        return render(request, "documentos/importar_ldp.html", {
            "history": list(reversed(history))[:10],
        })

    arquivo = request.FILES.get("arquivo")
    if not arquivo:
        messages.error(request, "Selecione uma planilha (.xlsx/.xlsm) para importar.")
        return render(request, "documentos/importar_ldp.html", {"history": list(reversed(history))[:10]})

    if not (arquivo.name.lower().endswith(".xlsx") or arquivo.name.lower().endswith(".xlsm")):
        messages.error(request, "Formato inválido. Envie um arquivo .xlsx ou .xlsm.")
        return render(request, "documentos/importar_ldp.html", {"history": list(reversed(history))[:10]})

    file_bytes = arquivo.read()
    header_row = _find_header_row(file_bytes, sheet_name="LDP")

    # lê via pandas a partir do header encontrado
    try:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name="LDP", header=header_row - 1)
    except Exception:
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=header_row - 1)

    # Mapeamento de colunas (sinônimos)
    COLS = {
        "codigo": ["Código", "Codigo", "CODIGO", "Code"],
        "revisao": ["Rev", "Revisão", "Revisao"],
        "titulo": ["Título", "Titulo", "Title"],
        "disciplina": ["Disciplina"],
        "fase": ["Fase"],
        "tipo_documento": ["Tipo", "Tipo Documento", "Tipo Doc", "Tipo de Documento"],
        "projeto": ["Projeto", "Project"],
    }

    col_codigo = _pick_column(df.columns, COLS["codigo"])
    col_revisao = _pick_column(df.columns, COLS["revisao"])
    col_titulo = _pick_column(df.columns, COLS["titulo"])
    col_disc = _pick_column(df.columns, COLS["disciplina"])
    col_fase = _pick_column(df.columns, COLS["fase"])
    col_tipo = _pick_column(df.columns, COLS["tipo_documento"])
    col_proj = _pick_column(df.columns, COLS["projeto"])

    error_items = []

    # coluna mínima
    if not col_codigo:
        messages.error(request, "Não encontrei a coluna 'Código' na planilha. Verifique o cabeçalho.")
        _add_error(error_items, header_row, "", "Código", "", "Coluna obrigatória ausente no cabeçalho")
        error_lines = [_error_line(e) for e in error_items]
        return render(request, "documentos/importar_ldp.html", {
            "errors": error_lines,
            "error_items": error_items,
            "errors_by_column": {"Código": 1},
            "history": list(reversed(history))[:10],
        })

    # Campos do model Documento (pra não quebrar se mudar)
    try:
        doc_fields = {f.name for f in Documento._meta.get_fields()}
    except Exception:
        doc_fields = set()

    # Projeto pode ser FK ou texto — tenta resolver com segurança
    def _resolve_projeto(val: str):
        val = (val or "").strip()
        if not val:
            return None
        try:
            f = Documento._meta.get_field("projeto")
        except Exception:
            return val  # se não existe field, retorna string
        # FK
        if getattr(f, "many_to_one", False):
            try:
                proj_fields = {pf.name for pf in Projeto._meta.get_fields()}
                key_field = None
                for cand in ("nome", "titulo", "descricao", "name"):
                    if cand in proj_fields:
                        key_field = cand
                        break
                if not key_field:
                    return None
                lookup = {f"{key_field}__iexact": val}
                obj = Projeto.objects.filter(**lookup).first()
                if obj:
                    return obj
                obj = Projeto.objects.create(**{key_field: val})
                return obj
            except Exception as e:
                return e
        return val

    # Regras de obrigatoriedade (mínimo + título se a coluna existir)
    required_cols = [("Código", col_codigo)]
    if col_titulo:
        required_cols.append(("Título", col_titulo))

    total_lidas = 0
    criadas = 0
    atualizadas = 0
    sem_mudanca = 0
    ignoradas = 0

    # Processa linhas
    for idx, row in df.iterrows():
        excel_row = header_row + 1 + int(idx)
        total_lidas += 1

        codigo = _clean(row.get(col_codigo))
        if not codigo:
            ignoradas += 1
            _add_error(error_items, excel_row, "", "Código", row.get(col_codigo), "Código vazio")
            continue

        revisao = _clean(row.get(col_revisao)) if col_revisao else ""

        # Monta defaults com segurança (só se o field existir)
        defaults = {}

        if col_titulo and "titulo" in doc_fields:
            titulo = _clean(row.get(col_titulo))
            if not titulo:
                ignoradas += 1
                _add_error(error_items, excel_row, codigo, "Título", row.get(col_titulo), "Título vazio")
                continue
            defaults["titulo"] = titulo

        if col_disc and "disciplina" in doc_fields:
            defaults["disciplina"] = _clean(row.get(col_disc))

        if col_fase and "fase" in doc_fields:
            defaults["fase"] = _clean(row.get(col_fase))

        tipo_val = _clean(row.get(col_tipo)) if col_tipo else ""
        if tipo_val:
            if "tipo_documento" in doc_fields:
                defaults["tipo_documento"] = tipo_val
            elif "tipo" in doc_fields:
                defaults["tipo"] = tipo_val

        if col_proj and "projeto" in doc_fields:
            proj_val = _clean(row.get(col_proj))
            if proj_val:
                resolved = _resolve_projeto(proj_val)
                if isinstance(resolved, Exception):
                    ignoradas += 1
                    _add_error(error_items, excel_row, codigo, "Projeto", proj_val, f"Falha ao resolver Projeto: {resolved}")
                    continue
                if resolved is None:
                    _add_error(error_items, excel_row, codigo, "Projeto", proj_val, "Não consegui mapear Projeto (model não compatível)")
                else:
                    defaults["projeto"] = resolved

        # calcula sem_mudanca (comparando antes/depois)
        try:
            before = None
            if defaults:
                before = Documento.objects.filter(codigo=codigo, revisao=revisao).values(*defaults.keys()).first()

            with transaction.atomic():
                doc, created = Documento.objects.update_or_create(
                    codigo=codigo,
                    revisao=revisao,
                    defaults=defaults
                )

            if created:
                criadas += 1
            else:
                if before is None:
                    atualizadas += 1
                else:
                    def _cmp(v):
                        return getattr(v, "pk", v)
                    same = True
                    for k, v in defaults.items():
                        if before.get(k) != _cmp(v):
                            same = False
                            break
                    if same:
                        sem_mudanca += 1
                    else:
                        atualizadas += 1

        except Exception as e:
            ignoradas += 1
            _add_error(error_items, excel_row, codigo, "—", "", f"Erro ao importar linha: {e}")

    # Resumo por coluna
    by_col = Counter([e["coluna"] for e in error_items]) if error_items else Counter()
    errors_by_column = dict(sorted(by_col.items(), key=lambda kv: (-kv[1], kv[0])))

    error_lines = [_error_line(e) for e in error_items]

    # Mensagens (top banner)
    msg_base = f"Lidas {total_lidas} | Criadas {criadas} | Atualizadas {atualizadas} | Sem mudança {sem_mudanca} | Erros {len(error_items)}"
    if error_items:
        preview = error_lines[:8]
        cols_top = ", ".join([f"{c}({n})" for c, n in list(errors_by_column.items())[:5]])
        messages.warning(request, msg_base + f". Colunas com problemas: {cols_top}. Primeiros erros: " + " | ".join(preview))
    else:
        messages.success(request, msg_base)

    # Session history
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    history.append({
        "ts": now,
        "arquivo": arquivo.name,
        "lidas": total_lidas,
        "criadas": criadas,
        "atualizadas": atualizadas,
        "sem_mudanca": sem_mudanca,
        "erros": len(error_items),
    })
    request.session["ldp_import_history"] = history[-10:]
    request.session.modified = True

    return render(request, "documentos/importar_ldp.html", {
        "arquivo_nome": arquivo.name,
        "total_lidas": total_lidas,
        "criadas": criadas,
        "atualizadas": atualizadas,
        "sem_mudanca": sem_mudanca,
        # compat: lista simples de strings
        "errors": error_lines,
        # detalhado: linha/coluna/valor/mensagem
        "error_items": error_items,
        "errors_by_column": errors_by_column,
        "history": list(reversed(history[-10:])),
    })

def importar_ldp_legacy(request):
    if request.method == "POST":
        arquivo = request.FILES.get("arquivo")
        if not arquivo:
            messages.error(request, "Selecione um arquivo CSV ou XLSX.")
            return redirect("documentos:importar_ldp")

        nome = arquivo.name.lower()
        total_linhas = 0

        # Aqui está uma implementação segura e genérica.
        # Você pode adaptar o mapeamento de colunas depois.
        try:
            if nome.endswith(".csv"):
                decoded = arquivo.read().decode("utf-8", errors="ignore").splitlines()
                reader = csv.reader(decoded, delimiter=";")
                for row in reader:
                    total_linhas += 1
                    # TODO: mapear colunas → Documento
            elif nome.endswith(".xlsx"):
                wb = openpyxl.load_workbook(arquivo)
                ws = wb.active
                for _ in ws.iter_rows(min_row=2):  # pula cabeçalho
                    total_linhas += 1
                    # TODO: mapear colunas → Documento
            else:
                messages.error(request, "Formato não suportado. Use CSV ou XLSX.")
                return redirect("documentos:importar_ldp")
        except Exception as e:
            messages.error(request, f"Erro ao processar arquivo: {e}")
            return redirect("documentos:importar_ldp")

        messages.success(
            request,
            f"Arquivo processado com sucesso ({total_linhas} linhas lidas). Ajuste a lógica de importação conforme necessário.",
        )
        return redirect("documentos:listar_documentos")

    return render(request, "documentos/importar_ldp.html")


# =================================================================
# MEDIÇÃO / EXPORTAÇÃO MEDIÇÃO
# =================================================================

def _aplicar_filtros_medicao(qs, request):
    filtros = {
        "projeto": (request.GET.get("projeto") or "").strip(),
        "disciplina": (request.GET.get("disciplina") or "").strip(),
        "fase": (request.GET.get("fase") or "").strip(),
        "tipo_doc": (request.GET.get("tipo_doc") or "").strip(),
        "status_documento": (request.GET.get("status_documento") or "").strip(),
        "status_emissao": (request.GET.get("status_emissao") or "").strip(),
        "q": (request.GET.get("q") or "").strip(),
    }

    if filtros["projeto"]:
        qs = qs.filter(projeto__nome__icontains=filtros["projeto"])
    if filtros["disciplina"]:
        qs = qs.filter(disciplina__icontains=filtros["disciplina"])
    if filtros["fase"]:
        qs = qs.filter(fase__icontains=filtros["fase"])
    if filtros["tipo_doc"]:
        qs = qs.filter(tipo_doc__icontains=filtros["tipo_doc"])
    if filtros["status_documento"]:
        qs = qs.filter(status_documento__icontains=filtros["status_documento"])
    if filtros["status_emissao"]:
        qs = qs.filter(status_emissao__icontains=filtros["status_emissao"])
    if filtros["q"]:
        qs = qs.filter(
            Q(codigo__icontains=filtros["q"]) | Q(titulo__icontains=filtros["q"])
        )

    return qs, filtros


def _calcular_medicao_queryset(docs_qs):
    med_raw = docs_qs.values("tipo_doc").annotate(
        total=Count("id"),
        emitidos=Count("id", filter=Q(status_emissao="Emitido")),
        nr=Count("id", filter=Q(status_emissao="Não Recebido")),
    )

    linhas = []
    total_usd = 0
    total_brl = 0

    for m in med_raw:
        tipo = m["tipo_doc"] or "Sem Tipo"
        emit = m["emitidos"]
        nr = m["nr"]

        v_emit_usd = emit * VALOR_MEDICAO_USD
        v_nr_usd = nr * VALOR_MEDICAO_USD
        v_emit_brl = v_emit_usd * TAXA_CAMBIO_REAIS
        v_nr_brl = v_nr_usd * TAXA_CAMBIO_REAIS

        total_usd += v_emit_usd + v_nr_usd
        total_brl += v_emit_brl + v_nr_brl

        linhas.append(
            {
                "tipo_doc": tipo,
                "total": m["total"],
                "emitidos": emit,
                "nao_recebidos": nr,
                "valor_emitidos_usd": f"{v_emit_usd:,.2f}",
                "valor_nr_usd": f"{v_nr_usd:,.2f}",
                "valor_emitidos_brl": f"{v_emit_brl:,.2f}",
                "valor_nr_brl": f"{v_nr_brl:,.2f}",
            }
        )

    totais = {
        "total_docs": sum(m["total"] for m in med_raw),
        "total_usd": f"{total_usd:,.2f}",
        "total_brl": f"{total_brl:,.2f}",
    }

    return linhas, totais


@login_required
def medicao(request):
    base_qs = Documento.objects.filter(ativo=True).select_related("projeto")
    docs, filtros = _aplicar_filtros_medicao(base_qs, request)

    linhas, totais = _calcular_medicao_queryset(docs)

    def _to_decimal(value):
        if value is None:
            return Decimal("0")
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        s = str(value).strip()
        if not s:
            return Decimal("0")
        s = s.replace("R$", "").replace("US$", "").strip()
        s = s.replace(",", "")
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return Decimal("0")

    total_count = sum(int(m.get("total") or 0) for m in linhas)
    emitidos = sum(int(m.get("emitidos") or 0) for m in linhas)
    nao_recebidos = sum(int(m.get("nao_recebidos") or 0) for m in linhas)

    total_emit_usd = sum(_to_decimal(m.get("valor_emitidos_usd")) for m in linhas)
    total_nr_usd = sum(_to_decimal(m.get("valor_nr_usd")) for m in linhas)
    total_emit_brl = sum(_to_decimal(m.get("valor_emitidos_brl")) for m in linhas)
    total_nr_brl = sum(_to_decimal(m.get("valor_nr_brl")) for m in linhas)

    total_usd = total_emit_usd + total_nr_usd
    total_brl = total_emit_brl + total_nr_brl
    ticket_medio_usd = (total_usd / emitidos) if emitidos > 0 else Decimal("0")

    totais_gerais = {
        "total": total_count,
        "emitidos": emitidos,
        "nao_recebidos": nao_recebidos,
        "valor_emitidos_usd": f"{total_emit_usd:,.2f}",
        "valor_emitidos_brl": f"{total_emit_brl:,.2f}",
        "valor_nao_recebidos_usd": f"{total_nr_usd:,.2f}",
        "valor_nao_recebidos_brl": f"{total_nr_brl:,.2f}",
    }

    total_geral = totais_gerais

    def _normalize_tipo(label):
        label = str(label or "").strip()
        return label if label else "Sem Tipo"

    def _build_chart_data(rows, value_key, is_currency=False):
        data = {}
        for row in rows:
            label = _normalize_tipo(row.get("tipo_doc"))
            raw = row.get(value_key)
            if is_currency:
                value = _to_decimal(raw)
            else:
                try:
                    value = int(raw or 0)
                except (TypeError, ValueError):
                    value = 0
            data[label] = data.get(label, 0) + value

        if not data:
            return [], [], False

        items = sorted(data.items(), key=lambda kv: kv[1], reverse=True)
        if len(items) > 10:
            top = items[:10]
            rest = items[10:]
            if is_currency:
                outros = sum((v for _, v in rest), Decimal("0"))
            else:
                outros = sum((int(v) for _, v in rest), 0)
            if outros:
                top.append(("Outros", outros))
            items = top

        labels = [k for k, _ in items]
        values = [v for _, v in items]
        if not values or all(v == 0 for v in values):
            return labels, values, False
        return labels, values, True

    charts_ok = False
    chart1_url = None
    chart2_url = None

    labels_totais, values_totais, has_totais = _build_chart_data(
        linhas, "total", is_currency=False
    )
    labels_usd, values_usd, has_usd = _build_chart_data(
        linhas, "valor_emitidos_usd", is_currency=True
    )
    charts_have_data = has_totais or has_usd

    if total_count > 0 and charts_have_data and settings.MEDIA_ROOT:
        try:
            import matplotlib

            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

            charts_dir = os.path.join(settings.MEDIA_ROOT, "medicao_charts")
            os.makedirs(charts_dir, exist_ok=True)

            def _serialize_values(vals):
                out = []
                for v in vals:
                    if isinstance(v, Decimal):
                        out.append(f"{v:.2f}")
                    else:
                        out.append(str(int(v)))
                return out

            payload = json.dumps(
                {
                    "totais": {
                        "labels": labels_totais,
                        "values": _serialize_values(values_totais),
                    },
                    "usd": {
                        "labels": labels_usd,
                        "values": _serialize_values(values_usd),
                    },
                },
                ensure_ascii=False,
                sort_keys=True,
            )
            digest = hashlib.md5(payload.encode("utf-8")).hexdigest()[:12]
            chart1_name = f"chart_totais_{digest}.png"
            chart2_name = f"chart_usd_{digest}.png"
            chart1_path = os.path.join(charts_dir, chart1_name)
            chart2_path = os.path.join(charts_dir, chart2_name)

            if has_totais and not os.path.exists(chart1_path):
                fig, ax = plt.subplots(figsize=(12, 5), dpi=100)
                fig.patch.set_facecolor("#ffffff")
                ax.set_facecolor("#ffffff")
                vals = [int(v) for v in values_totais]
                y_pos = range(len(labels_totais))
                ax.barh(y_pos, vals, color="#2563eb")
                ax.set_yticks(y_pos)
                ax.set_yticklabels(labels_totais)
                ax.invert_yaxis()
                ax.set_xlabel("Total")
                ax.set_title("Totais por Tipo de Documento", color="#0f172a")
                ax.grid(axis="x", color="#e2e8f0", linestyle="--", linewidth=0.8)
                ax.tick_params(colors="#334155")
                for spine in ax.spines.values():
                    spine.set_color("#e2e8f0")
                fig.tight_layout()
                fig.savefig(chart1_path, bbox_inches="tight")
                plt.close(fig)

            if has_usd and not os.path.exists(chart2_path):
                fig, ax = plt.subplots(figsize=(12, 5), dpi=100)
                fig.patch.set_facecolor("#ffffff")
                ax.set_facecolor("#ffffff")
                vals = [float(v) for v in values_usd]
                ax.bar(labels_usd, vals, color="#0ea5e9")
                ax.set_ylabel("USD")
                ax.set_title("Valores Emitidos por Tipo (USD)", color="#0f172a")
                ax.grid(axis="y", color="#e2e8f0", linestyle="--", linewidth=0.8)
                ax.tick_params(colors="#334155")
                ax.tick_params(axis="x", rotation=25)
                for spine in ax.spines.values():
                    spine.set_color("#e2e8f0")
                fig.tight_layout()
                fig.savefig(chart2_path, bbox_inches="tight")
                plt.close(fig)

            base_media = (settings.MEDIA_URL or "/media/").rstrip("/")
            if has_totais:
                chart1_url = f"{base_media}/medicao_charts/{chart1_name}"
                charts_ok = True
            if has_usd:
                chart2_url = f"{base_media}/medicao_charts/{chart2_name}"
                charts_ok = True
        except Exception:
            charts_ok = False

    tem_dados_medicao = bool(linhas) and (
        total_count > 0 or total_emit_usd > 0 or total_emit_brl > 0
    )

    return render(
        request,
        "documentos/medicao.html",
        {
            "linhas": linhas,
            "resumo": linhas,
            "page": "medicao",
            "totais": totais,
            "totais_gerais": totais_gerais,
            "total_geral": total_geral,
            "total_docs": total_count,
            "emitidos_total": emitidos,
            "nao_recebidos_total": nao_recebidos,
            "total_usd": f"{total_usd:,.2f}",
            "total_brl": f"{total_brl:,.2f}",
            "ticket_medio_usd": f"{ticket_medio_usd:,.2f}",
            "tem_dados_medicao": tem_dados_medicao,
            "charts_ok": charts_ok,
            "chart1_url": chart1_url,
            "chart2_url": chart2_url,
            "filtros": filtros,
            "projetos": (
                Documento.objects.filter(ativo=True)
                .values_list("projeto__nome", flat=True)
                .exclude(projeto__nome__isnull=True)
                .exclude(projeto__nome__exact="")
                .distinct()
                .order_by("projeto__nome")
            ),
            "disciplinas": (
                Documento.objects.filter(ativo=True)
                .values_list("disciplina", flat=True)
                .exclude(disciplina__isnull=True)
                .exclude(disciplina__exact="")
                .distinct()
                .order_by("disciplina")
            ),
            "fases": (
                Documento.objects.filter(ativo=True)
                .values_list("fase", flat=True)
                .exclude(fase__isnull=True)
                .exclude(fase__exact="")
                .distinct()
                .order_by("fase")
            ),
            "tipos": (
                Documento.objects.filter(ativo=True)
                .values_list("tipo_doc", flat=True)
                .exclude(tipo_doc__isnull=True)
                .exclude(tipo_doc__exact="")
                .distinct()
                .order_by("tipo_doc")
            ),
            "status_docs": (
                Documento.objects.filter(ativo=True)
                .values_list("status_documento", flat=True)
                .exclude(status_documento__isnull=True)
                .exclude(status_documento__exact="")
                .distinct()
                .order_by("status_documento")
            ),
            "status_emissoes": (
                Documento.objects.filter(ativo=True)
                .values_list("status_emissao", flat=True)
                .exclude(status_emissao__isnull=True)
                .exclude(status_emissao__exact="")
                .distinct()
                .order_by("status_emissao")
            ),
        },
    )


@login_required
def exportar_medicao_excel(request):
    base_qs = Documento.objects.filter(ativo=True).select_related("projeto")
    docs, _filtros = _aplicar_filtros_medicao(base_qs, request)

    linhas, totais = _calcular_medicao_queryset(docs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medição GED"

    headers = [
        "Tipo de Documento",
        "Total",
        "Emitidos",
        "Não Recebidos",
        "Valor Emitidos (USD)",
        "Valor Não Recebidos (USD)",
        "Valor Emitidos (BRL)",
        "Valor Não Recebidos (BRL)",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="0052A2", end_color="0052A2", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in ws[1]:
        col.fill = header_fill
        col.font = Font(color="FFFFFF", bold=True)
        col.border = border
        col.alignment = Alignment(horizontal="center")

    for linha in linhas:
        ws.append(
            [
                linha["tipo_doc"],
                linha["total"],
                linha["emitidos"],
                linha["nao_recebidos"],
                linha["valor_emitidos_usd"],
                linha["valor_nr_usd"],
                linha["valor_emitidos_brl"],
                linha["valor_nr_brl"],
            ]
        )

    row_total = [
        "TOTAL",
        totais["total_docs"],
        "",
        "",
        totais["total_usd"],
        "",
        totais["total_brl"],
        "",
    ]
    ws.append([])
    ws.append(row_total)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 3

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"medicao_{date.today().isoformat()}.xlsx"

    return HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =================================================================
# EXCLUSÃO / LIXEIRA
# =================================================================

@login_required
def excluir_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)
    motivo = request.POST.get("motivo", "") if request.method == "POST" else ""
    mover_para_lixeira(documento, request, motivo=motivo)
    messages.success(request, "Documento movido para a lixeira.")
    return redirect("documentos:listar_documentos")


@login_required
def excluir_selecionados(request):
    if request.method != "POST":
        return redirect("documentos:listar_documentos")

    ids = request.POST.getlist("selecionados") or []
    count = 0
    for _id in ids:
        try:
            doc = Documento.objects.get(id=_id)
            if mover_para_lixeira(doc, request, motivo="Exclusão em lote"):
                count += 1
        except Documento.DoesNotExist:
            continue

    messages.success(request, f"{count} documento(s) movido(s) para a lixeira.")
    return redirect("documentos:listar_documentos")


@login_required
def lixeira(request):
    docs = Documento.objects.filter(ativo=False).order_by("-deletado_em")
    return render(request, "documentos/lixeira.html", {"documentos": docs})


@login_required
def restaurar_documento(request, documento_id):
    documento = get_object_or_404(Documento, id=documento_id)
    restaurar_da_lixeira(documento, request)
    messages.success(request, "Documento restaurado com sucesso.")
    return redirect("documentos:lixeira")


# =================================================================
# CONFIGURAÇÕES UNIFICADAS (Usuário + Administração)
# =================================================================

@login_required
def configuracoes(request):
    from apps.contas.models import UserConfig  # Importa aqui para evitar ciclos
    from apps.contas.forms import UserConfigForm  # Importa aqui para evitar ciclos

    # Carrega ou cria configuração do usuário
    config, _ = UserConfig.objects.get_or_create(user=request.user)

    # Salvar preferências pessoais
    if request.method == "POST" and "salvar_prefs" in request.POST:
        config.tema = request.POST.get("tema", config.tema)
        config.animacoes = request.POST.get("animacoes", "True") == "True"
        config.notificacoes_email = request.POST.get("notificacoes_email", "True") == "True"
        config.dashboard_expandido = request.POST.get("dashboard_expandido", "True") == "True"
        config.save()
        messages.success(request, "Preferências atualizadas com sucesso!")

    # Dados para parte administrativa (apenas admins/master)
    projetos = Projeto.objects.all().order_by("nome")
    resp_disc = ResponsavelDisciplina.objects.all().order_by("disciplina")
    etapas = WorkflowEtapa.objects.all().order_by("ordem")

    return render(
        request,
        "documentos/configuracoes.html",
        {
            "projetos": projetos,
            "responsaveis": resp_disc,
            "etapas": etapas,
            "config": config,
        },
    )

# =================================================================
# BUSCA GLOBAL
# =================================================================

@login_required
def buscar_global(request):
    termo = request.GET.get("q", "").strip()
    resultados = []

    if termo:
        qs = Documento.objects.filter(ativo=True).select_related("projeto")
        qs = qs.filter(
            Q(codigo__icontains=termo)
            | Q(titulo__icontains=termo)
            | Q(disciplina__icontains=termo)
            | Q(projeto__nome__icontains=termo)
        )

        for doc in qs:
            resultados.append(
                {
                    "id": doc.id,
                    "codigo": highlight_text(doc.codigo, termo),
                    "titulo": highlight_text(doc.titulo, termo),
                    "disciplina": doc.disciplina or "",
                    "projeto": doc.projeto.nome if doc.projeto else "",
                }
            )

    return render(
        request,
        "documentos/buscar.html",
        {"resultados": resultados, "termo": termo},
    )

# =============================================
# 📊 DASHBOARD MASTER – FINANCE + STATUS ENGINE
# =============================================
from django.db.models import Count, Sum
from .models import Documento, LogAuditoria, ProjetoFinanceiro


def dashboard_master(request):
    projeto_nome = "TP25 NAVIOS HANDY CLASSE 80"
    taxa_brl = 5.7642

    # ============================
    # BASE DE DOCUMENTOS FILTRADOS
    # ============================
    docs = Documento.objects.filter(
        projeto__nome__icontains=projeto_nome,
        ativo=True
    )

    total_docs = docs.count()

    total_emitidos     = docs.filter(status_emissao="Emitido").count()
    total_aprovados    = docs.filter(status_emissao="Aprovado").count()
    total_em_revisao   = docs.filter(status_documento__icontains="Revisão").count()
    total_nao_recebidos = docs.filter(status_emissao="Não Recebido").count()
    total_excluidos    = docs.filter(deletado_em__isnull=False).count()
    total_restaurados  = docs.filter(motivo_exclusao__icontains="Restaurado").count()

    # ============================
    # FINANCEIRO POR FASE / STATUS
    # ============================
    valores = ProjetoFinanceiro.objects.filter(projeto__nome=projeto_nome)

    valor_basico   = valores.filter(fase="Básico").aggregate(total=Sum("valor_total_usd"))["total"] or 0
    valor_aprovado = valores.filter(fase="Aprovado").aggregate(total=Sum("valor_total_usd"))["total"] or 0
    valor_asbuilt  = valores.filter(fase="As Built").aggregate(total=Sum("valor_total_usd"))["total"] or 0


    total_basico_docs = docs.filter(fase="Básico").count() or 1

    # USD Unitários (valor / documentos)
    usd_basico_unit   = valor_basico   / total_basico_docs
    usd_aprovado_unit = valor_aprovado / total_basico_docs
    usd_asbuilt_unit  = valor_asbuilt  / total_basico_docs

    # Medição baseada no STATUS EMISSÃO
    valor_emitidos_usd = total_emitidos     * usd_basico_unit
    valor_nao_rec_usd  = total_nao_recebidos * usd_basico_unit
    valor_total_usd    = valor_emitidos_usd + valor_nao_rec_usd

    # Conversão BRL
    valor_emitidos_brl = valor_emitidos_usd * taxa_brl
    valor_nao_rec_brl  = valor_nao_rec_usd  * taxa_brl
    valor_total_brl    = valor_total_usd    * taxa_brl

    # =============
    # AUDITORIA
    # =============
    auditoria_acoes = LogAuditoria.objects.order_by('-data')[:10]

    top_excluidores = LogAuditoria.objects.filter(acao="Excluído") \
        .values('usuario') \
        .annotate(qtd=Count('id')) \
        .order_by('-qtd')[:5]

    # =====================
    # GRÁFICOS
    # =====================
    disc_labels = list(docs.values_list("disciplina", flat=True).distinct())
    disc_data   = [docs.filter(disciplina=d).count() for d in disc_labels]

    status_labels = ["Emitido", "Aprovado", "Revisão", "Não Recebido"]
    status_data   = [
        total_emitidos,
        total_aprovados,
        total_em_revisao,
        total_nao_recebidos
    ]

    # =====================
    # RENDERIZA O TEMPLATE
    # =====================
    context = {
        "total_docs": total_docs,
        "total_emitidos": total_emitidos,
        "total_aprovados": total_aprovados,
        "total_em_revisao": total_em_revisao,
        "total_nao_recebidos": total_nao_recebidos,
        "total_excluidos": total_excluidos,
        "total_restaurados": total_restaurados,

        "valor_emitidos_usd": valor_emitidos_usd,
        "valor_nao_rec_usd": valor_nao_rec_usd,
        "valor_total_usd": valor_total_usd,

        "valor_emitidos_brl": valor_emitidos_brl,
        "valor_nao_rec_brl": valor_nao_rec_brl,
        "valor_total_brl": valor_total_brl,

        # Graphs
        "disc_labels": disc_labels,
        "disc_data": disc_data,
        "status_labels": status_labels,
        "status_data": status_data,

        # Auditoria
        "top_excluidores": top_excluidores,
        "auditoria_acoes": auditoria_acoes,
    }

    return render(request, "documentos/dashboard_master.html", context)

# =================================================================
# BUSCA AJAX
# =================================================================

@login_required
def buscar_ajax(request):
    termo = request.GET.get("q", "").strip()
    items = []

    if termo:
        qs = Documento.objects.filter(ativo=True)
        qs = qs.filter(
            Q(codigo__icontains=termo)
            | Q(titulo__icontains=termo)
            | Q(disciplina__icontains=termo)
        )[:20]

        for d in qs:
            items.append(
                {
                    "id": d.id,
                    "codigo": d.codigo,
                    "titulo": d.titulo,
                    "disciplina": d.disciplina or "",
                }
            )

    return JsonResponse({"results": items})

# ==============================
# AUTO-FIX v2 - views ausentes
# ==============================

@login_required
@require_POST
def esvaziar_lixeira(request, *args, **kwargs):
    """
    Esvazia a lixeira (hard delete) dos documentos marcados como deletados.
    Regra: Documento deletado => deletado_em != None OU ativo == False.
    """
    from django.contrib import messages
    from django.shortcuts import redirect
    from django.db import transaction
    from django.db.models import Q
    from .models import Documento

    qs = Documento.objects.filter(Q(deletado_em__isnull=False) | Q(ativo=False))
    total = qs.count()

    with transaction.atomic():
        qs.delete()

    messages.success(request, f"Lixeira esvaziada com sucesso ({total} itens removidos).")
    return redirect("documentos:lixeira")

