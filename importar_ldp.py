import openpyxl
import os
import django
from datetime import datetime
import difflib

# ------------------------------
# CONFIGURAR DJANGO
# ------------------------------
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ged.settings')
django.setup()

from apps.documentos.models import Documento

# ------------------------------
# Caminho da planilha
# ------------------------------
CAMINHO_ARQUIVO = r"C:\Meus Projetos\I-LD-4880.00-9311-000-CZ1-001_R0.xlsx"
print("ðŸ“„ Carregando planilha:", CAMINHO_ARQUIVO)

wb = openpyxl.load_workbook(CAMINHO_ARQUIVO, data_only=True)
ws = wb.active

# ------------------------------
# Mapa desejado (nomes do Django)
# ------------------------------
CAMPOS_DJANGO = {
    "projeto",
    "fase",
    "tipo_doc",
    "codigo",
    "revisao",
    "disciplina",
    "titulo",
    "status_ldp",
    "status_emissao",
    "numero_grdt",
    "numero_pcf",
    "data_emissao_tp",
}

# ------------------------------
# PossÃ­veis nomes do cabeÃ§alho por coluna
# (suporte a variaÃ§Ãµes)
# ------------------------------
ALIAS = {
    "projeto": ["projeto", "project"],
    "fase": ["fase", "fa", "stg"],
    "tipo_doc": ["tipo de doc", "tipo doc", "documento", "doc type", "tipo"],
    "codigo": ["codigo", "cÃ³digo", "code", "id"],
    "revisao": ["revisao", "rev", "revision"],
    "disciplina": ["disciplina", "discipline", "dept"],
    "titulo": ["titulo", "tÃ­tulo", "title", "nome"],
    "status_ldp": ["status ldp", "ldp", "ldp status"],
    "status_emissao": ["status emissÃ£o", "emissÃ£o", "status emissao"],
    "numero_grdt": ["numero grdt", "nÂº grdt", "grdt"],
    "numero_pcf": ["numero pcf", "nÂº pcf", "pcf"],
    "data_emissao_tp": ["data emissÃ£o tp", "data emissao tp", "emissao tp", "release date"],
}

# ------------------------------
# FunÃ§Ã£o inteligente para mapear cabeÃ§alho
# ------------------------------
def identificar_campo_excel(nome_coluna):
    if not nome_coluna:
        return None
    nome = nome_coluna.strip().lower()

    # 1. Match exato com alias
    for campo, apelidos in ALIAS.items():
        if nome in [a.lower() for a in apelidos]:
            return campo

    # 2. Match aproximado (difflib)
    for campo, apelidos in ALIAS.items():
        melhor = difflib.get_close_matches(nome, apelidos, n=1, cutoff=0.7)
        if melhor:
            return campo

    return None  # desconhecido

# ---------------------------------------
# LER CABEÃ‡ALHO
# ---------------------------------------
header = [cell.value for cell in ws[1]]
print("\nðŸ“Œ CabeÃ§alho detectado:", header)

index = {}

for i, coluna in enumerate(header):
    campo = identificar_campo_excel(coluna)
    if campo:
        index[i] = campo

print("ðŸ” Ãndice mapeado:", index)

# ---------------------------------------
# FunÃ§Ã£o para converter datas
# ---------------------------------------
def converter_data(valor):
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor.date()

    try:
        return datetime.strptime(str(valor), "%d/%m/%Y").date()
    except:
        return None

# ---------------------------------------
# Processar linhas
# ---------------------------------------
total = 0
erros = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    dados = {}

    for col_index, campo in index.items():
        valor = row[col_index]

        # converter data
        if campo == "data_emissao_tp":
            dados[campo] = converter_data(valor)
        else:
            dados[campo] = valor

    # Documento precisa ter CÃ“DIGO
    if not dados.get("codigo"):
        continue

    try:
        Documento.objects.update_or_create(
            codigo=dados["codigo"],
            defaults=dados
        )
        total += 1

    except Exception as e:
        erros += 1
        print("âŒ Erro na linha:", row)
        print("   â†’", e)

print("\nâœ” ImportaÃ§Ã£o Inteligente Finalizada!")
print("ðŸ“ Sucesso:", total)
print("âš  Erros:", erros)

